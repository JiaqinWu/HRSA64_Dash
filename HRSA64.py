import streamlit as st
import re
import pandas as pd
from millify import millify # shortens values (10_000 ---> 10k)
from streamlit_extras.metric_cards import style_metric_cards # beautify metric card with css
from datetime import datetime, timedelta
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import altair as alt
import json
import time
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import io
from mailjet_rest import Client
import openpyxl
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from PIL import Image as PILImage, ImageDraw, ImageFont
import base64
import urllib.request

st.set_page_config(
        page_title="GU TAP System",
        page_icon="https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png", 
        layout="centered"
    ) 

scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
#creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
# Use Streamlit's secrets management
creds_dict = st.secrets["gcp_service_account"]
# Extract individual attributes needed for ServiceAccountCredentials
credentials = {
    "type": creds_dict.type,
    "project_id": creds_dict.project_id,
    "private_key_id": creds_dict.private_key_id,
    "private_key": creds_dict.private_key,
    "client_email": creds_dict.client_email,
    "client_id": creds_dict.client_id,
    "auth_uri": creds_dict.auth_uri,
    "token_uri": creds_dict.token_uri,
    "auth_provider_x509_cert_url": creds_dict.auth_provider_x509_cert_url,
    "client_x509_cert_url": creds_dict.client_x509_cert_url,
}

# Create JSON string for credentials
creds_json = json.dumps(credentials)

# Load credentials and authorize gspread
creds = ServiceAccountCredentials.from_json_keyfile_dict(json.loads(creds_json), scope)
client = gspread.authorize(creds)


def send_email_mailjet(to_email, subject, body):
    """Send email via Mailjet and return success status"""
    api_key = st.secrets["mailjet"]["api_key"]
    api_secret = st.secrets["mailjet"]["api_secret"]
    sender = st.secrets["mailjet"]["sender"]

    mailjet = Client(auth=(api_key, api_secret), version='v3.1')

    data = {
        'Messages': [
            {
                "From": {
                    "Email": sender,
                    "Name": "GU-TAP System"
                },
                "To": [
                    {
                        "Email": to_email,
                        "Name": to_email.split("@")[0]
                    }
                ],
                "Subject": subject,
                "TextPart": body
            }
        ]
    }

    try:
        result = mailjet.send.create(data=data)
        if result.status_code == 200:
            return True
        else:
            st.warning(f"❌ Failed to email {to_email}: Status {result.status_code}")
            return False
    except Exception as e:
        st.error(f"❗ Mailjet error: {e}")
        return False

# Student schedule data
STUDENT_SCHEDULE = {
    "Asha Patel": {
        "email": "ap2349@georgetown.edu",
        "schedule": {
            "Monday": "9am - 5pm",
            "Tuesday": "9am - 12pm", 
            "Wednesday": "9am - 5pm",
            "Thursday": "9am - 5pm",
            "Friday": ""
        }
    },
    "Hang Nguyen": {
        "email": "htn16@georgetown.edu",
        "schedule": {
            "Monday": "9am - 4pm",
            "Tuesday": "9am - 4pm",
            "Wednesday": "9am - 4pm", 
            "Thursday": "9am - 4pm",
            "Friday": "9am - 4pm"
        }
    },
    "Olayinka Adedeji": {
        "email": "ooa36@georgetown.edu",
        "schedule": {
            "Monday": "1pm - 5pm",
            "Tuesday": "9am - 5pm",
            "Wednesday": "9am - 11am, 3pm - 5pm",
            "Thursday": "9am - 4pm",
            "Friday": ""
        }
    },
    "Saara Bidiwala": {
        "email": "ssb120@georgetown.edu", 
        "schedule": {
            "Monday": "",
            "Tuesday": "1:15pm - 5pm",
            "Wednesday": "",
            "Thursday": "1:15pm - 5pm", 
            "Friday": ""
        }
    },
    "Ziqiao Shan": {
        "email": "zs352@georgetown.edu",
        "schedule": {
            "Monday": "1pm - 5pm",
            "Tuesday": "10am - 12pm",
            "Wednesday": "10am - 5pm",
            "Thursday": "10am - 5pm",
            "Friday": "10am - 5pm"
        }
    },
    "Yannis Ying": {
        "email": "sy803@georgetown.edu",
        "schedule": {
            "Monday": "",
            "Tuesday": "",
            "Wednesday": "9am - 4pm",
            "Thursday": "9am - 5pm",
            "Friday": "9am - 5pm"
        }
    },
    "Shedrack Osuji": {
        "email": "sco47@georgetown.edu",
        "schedule": {
            "Monday": "11am - 5pm",
            "Tuesday": "",
            "Wednesday": "11am - 5pm",
            "Thursday": "",
            "Friday": "9am - 5pm"
        }
    }
}

def parse_time_range(time_str):
    """Parse time range string like '9am - 5pm' or '10am - 12pm, 2pm - 5pm'"""
    if not time_str or time_str.strip() == "":
        return []
    
    time_ranges = []
    # Handle multiple ranges separated by comma
    ranges = time_str.split(',')
    
    for range_str in ranges:
        range_str = range_str.strip()
        if ' - ' in range_str:
            start_time, end_time = range_str.split(' - ')
            time_ranges.append((start_time.strip(), end_time.strip()))
    
    return time_ranges

def time_to_24h(time_str):
    """Convert time string like '9am' or '2pm' to 24-hour format"""
    time_str = time_str.strip().lower()
    
    # Remove 'am'/'pm' and convert
    if 'am' in time_str:
        hour = int(time_str.replace('am', '').strip())
        if hour == 12:
            hour = 0
    elif 'pm' in time_str:
        hour = int(time_str.replace('pm', '').strip())
        if hour != 12:
            hour += 12
    else:
        hour = int(time_str)
    
    return hour

def is_time_overlap(request_time, student_availability):
    """Check if request time overlaps with student availability"""
    if not student_availability:
        return False
    
    # Parse request time (format: "09:00 - 17:00")
    try:
        request_start, request_end = request_time.split(' - ')
        req_start_hour = int(request_start.split(':')[0])
        req_end_hour = int(request_end.split(':')[0])
    except:
        return False
    
    for avail_start, avail_end in student_availability:
        avail_start_hour = time_to_24h(avail_start)
        avail_end_hour = time_to_24h(avail_end)
        
        # Check for overlap
        if req_start_hour < avail_end_hour and req_end_hour > avail_start_hour:
            return True
    
    return False

def get_available_students(date_str, time_str):
    """Get list of students available for the given date and time"""
    try:
        # Get day of week from date
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        day_name = date_obj.strftime("%A")  # Monday, Tuesday, etc.
        
        available_students = []
        
        for student_name, student_info in STUDENT_SCHEDULE.items():
            schedule = student_info["schedule"]
            if day_name in schedule:
                availability = parse_time_range(schedule[day_name])
                if is_time_overlap(time_str, availability):
                    available_students.append({
                        "name": student_name,
                        "email": student_info["email"],
                        "availability": schedule[day_name]
                    })
        
        return available_students
    except Exception as e:
        st.error(f"Error getting available students: {e}")
        return []

def send_support_request_notifications(date_str, time_str, request_description, anticipated_delivery, tap_name, tap_email):
    """Send email notifications to available students"""
    available_students = get_available_students(date_str, time_str)
    
    if not available_students:
        st.info("No students are available during the requested time slot.")
        return True
    
    subject = f"New Support Request Available - {date_str} at {time_str}"
    success_count = 0
    total_count = len(available_students)
    
    for i, student in enumerate(available_students, 1):
        body = f"""
Dear {student['name']},

A new support request has been submitted and you are available during the requested time slot.

Request Details:
- Date: {date_str}
- Time: {time_str}
- TAP Name: {tap_name}
- TAP Email: {tap_email}
- Request Description: {request_description}
- Anticipated Deliverable: {anticipated_delivery}

If you are interested in taking this request, please log into the GU-TAP System and assign it to yourself.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
        """
        
        try:
            status = send_email_mailjet(
                to_email=student['email'],
                subject=subject,
                body=body.strip()
            )
            if status:
                success_count += 1
                st.success(f"📧 ({i}/{total_count}) Sent to {student['name']} ({student['email']})")
            
            # Add delay between emails to avoid rate limiting (except after the last email)
            if i < total_count:
                time.sleep(0.8)  # 0.8 second delay between emails
                
        except Exception as e:
            st.warning(f"⚠️ Failed to send notification to {student['name']}: {e}")
    
    if success_count == total_count:
        return True
    else:
        st.warning(f"⚠️ Sent {success_count}/{total_count} emails successfully")
        return False

def send_project_request_notifications(request_description, anticipated_delivery, time_commitment, anticipated_deadline, tap_name, tap_email):
    """Send email notifications to all students for non-meeting project requests"""
    subject = f"New Project Request Available - {anticipated_delivery}"
    
    student_list = list(STUDENT_SCHEDULE.items())
    success_count = 0
    total_count = len(student_list)
    
    for i, (student_name, student_info) in enumerate(student_list, 1):
        body = f"""
Dear {student_name},

A new project support request has been submitted and is available for assignment.

Project Details:
- Project Type: {anticipated_delivery}
- Time Commitment: {time_commitment}
- Anticipated Deadline: {anticipated_deadline}
- TAP Name: {tap_name}
- TAP Email: {tap_email}
- Project Description: {request_description}

This is a flexible project that can be completed according to your schedule. If you are interested in taking this project, please log into the GU-TAP System and assign it to yourself.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
        """
        
        try:
            status = send_email_mailjet(
                to_email=student_info['email'],
                subject=subject,
                body=body.strip()
            )
            if status:
                success_count += 1
                st.success(f"📧 ({i}/{total_count}) Sent to {student_name} ({student_info['email']})")
            
            # Add delay between emails to avoid rate limiting (except after the last email)
            if i < total_count:
                time.sleep(0.8)  # 0.8 second delay between emails
                
        except Exception as e:
            st.warning(f"⚠️ Failed to send project notification to {student_name}: {e}")
    
    if success_count == total_count:
        return True
    else:
        st.warning(f"⚠️ Sent {success_count}/{total_count} emails successfully")
        return False

# Travel Form Generator Functions
@st.cache_data
def load_excel_template():
    """Load the Excel template and identify form structure"""
    import os
    # Try current directory first
    template_path = 'Georgetown Domestic Travel Authorization Form.xlsx'
    if not os.path.exists(template_path):
        # Try alternative path relative to this file
        alt_path = os.path.join(os.path.dirname(__file__), '..', 'Georgetown_Travel_Form_Generator', 'Georgetown Domestic Travel Authorization Form.xlsx')
        if os.path.exists(alt_path):
            template_path = alt_path
        else:
            raise FileNotFoundError(f"Excel template not found. Tried: {template_path} and {alt_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Reimbursement Form']
    return wb, ws

def chunk_list(items, chunk_size):
    """Yield successive chunks of size chunk_size from items."""
    if items is None:
        return []
    return [items[i:i+chunk_size] for i in range(0, len(items), chunk_size)]

def pad_to_length(items, length, pad_value=''):
    """Return a copy padded to given length."""
    items = list(items)
    if len(items) < length:
        items.extend([pad_value] * (length - len(items)))
    return items

def number_text_input(label, key, value=0.0, min_value=0.0, placeholder="0.00"):
    """Text input that accepts numeric values only, with validation.
    Returns the numeric value and shows inline warnings if invalid."""
    # Initialize session state if not exists
    if key not in st.session_state:
        st.session_state[key] = str(value) if value else ""
    
    # Track validation state for this specific input
    validation_key = f"{key}_has_error"
    
    text_val = st.text_input(label, key=key, placeholder=placeholder)
    
    # If empty, return 0.0 (no validation needed, clear any previous errors)
    if not text_val or not text_val.strip():
        st.session_state[validation_key] = False
        return 0.0
    
    # Try to extract numeric value from input
    # Remove common non-numeric characters like $, commas, spaces, etc.
    cleaned_text = text_val.strip().replace('$', '').replace(',', '').replace(' ', '')
    
    # Try to parse as float
    has_error = False
    error_message = None
    try:
        num_val = float(cleaned_text)
        if num_val < min_value:
            num_val = min_value
        # Input is valid, clear error state
        st.session_state[validation_key] = False
        return num_val
    except (ValueError, AttributeError):
        # Check if there are any invalid characters
        # Allow: integers, decimals, negative numbers
        if not re.match(r'^-?\d+(\.\d+)?$', cleaned_text):
            has_error = True
            error_message = "⚠️ Invalid input. Please enter a valid number."
    
    # Update error state
    st.session_state[validation_key] = has_error
    
    # Show warning inline if there's an error
    if has_error and error_message:
        st.warning(error_message)
        return 0.0
    
    return 0.0

def generate_signature_image(text, width=600, height=120, scale_factor=3):
    """Generate a signature-style image from text with high resolution"""
    if not text or not text.strip():
        return None
    
    # Use scale factor for high-resolution rendering (render at 3x, then scale down)
    scaled_width = width * scale_factor
    scaled_height = height * scale_factor
    
    # Create an image with white background (blank/transparent-looking) at high resolution
    img = PILImage.new('RGB', (scaled_width, scaled_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    # Try to use a cursive/signature-style font, fallback to default
    # Include more common paths and try PIL's built-in fonts
    font_paths = [
        '/System/Library/Fonts/Supplemental/SnellRoundhand.ttc',  # macOS
        '/System/Library/Fonts/Supplemental/Chalkduster.ttf',    # macOS alternative
        'C:/Windows/Fonts/brushsc.ttf',                           # Windows
        'C:/Windows/Fonts/BRUSHSCI.TTF',                          # Windows
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf', # Linux
        '/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf', # Linux alternative
        '/usr/share/fonts/truetype/noto/NotoSans-Italic.ttf',     # Linux alternative
    ]
    
    # Start with larger font size (scale it too)
    font_size = 72 * scale_factor  # Larger base font size
    font = None
    font_path_used = None
    
    for font_path in font_paths:
        try:
            font = ImageFont.truetype(font_path, font_size)
            font_path_used = font_path
            break
        except (OSError, IOError, Exception):
            continue
    
    # If no system font found, try to use PIL's default font or create a simple signature style
    if font is None:
        try:
            # Try common font names that might be available
            import platform
            system = platform.system()
            if system == 'Windows':
                # Try Windows common fonts
                for font_name in ['arial', 'calibri', 'times']:
                    try:
                        font = ImageFont.truetype(f"{font_name}.ttf", font_size)
                        font_path_used = font_name
                        break
                    except:
                        continue
            elif system == 'Linux':
                # Try Linux common fonts
                for font_path_linux in [
                    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                ]:
                    try:
                        font = ImageFont.truetype(font_path_linux, font_size)
                        font_path_used = font_path_linux
                        break
                    except:
                        continue
        except:
            pass
        
        # Final fallback: use PIL's default font
        if font is None:
            try:
                # Try to load default font with larger size
                font = ImageFont.load_default()
                font_size = 36 * scale_factor
            except:
                # Ultimate fallback
                font = ImageFont.load_default()
                font_size = 36 * scale_factor
    
    # Calculate text dimensions first
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except:
        # Fallback if textbbox fails
        text_width = len(text) * font_size * 0.6
        text_height = font_size * 1.2
    
    # Adjust font size if text is too wide to fit in available width
    min_font_size = 30 * scale_factor
    while text_width > scaled_width - (40 * scale_factor) and font_size > min_font_size:
        font_size -= 3 * scale_factor
        try:
            if font_path_used and font_path_used not in ['arial', 'calibri', 'times']:
                font = ImageFont.truetype(font_path_used, font_size)
            else:
                font = ImageFont.load_default()
        except:
            font = ImageFont.load_default()
        try:
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
        except:
            text_width = len(text) * font_size * 0.6
            text_height = font_size * 1.2
    
    # Calculate position (left-aligned with padding, vertically centered)
    padding = 20 * scale_factor
    x = padding
    y = (scaled_height - text_height) / 2
    
    # Draw the signature text in black with antialiasing
    try:
        draw.text((x, y), text, fill=(0, 0, 0), font=font)
    except Exception as e:
        # If font drawing fails, try with default font
        try:
            font = ImageFont.load_default()
            draw.text((x, y), text, fill=(0, 0, 0), font=font)
            # Recalculate dimensions with default font
            try:
                bbox = draw.textbbox((0, 0), text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                text_width = len(text) * font_size * 0.6
                text_height = font_size * 1.2
        except:
            # Ultimate fallback - draw text without font specification
            draw.text((x, y), text, fill=(0, 0, 0))
    
    # Add a thicker underline for signature effect (also scaled)
    line_y = y + text_height + (8 * scale_factor)
    line_width = 3 * scale_factor
    draw.line([(x - 5 * scale_factor, line_y), (x + text_width + 5 * scale_factor, line_y)], 
              fill=(0, 0, 0), width=int(line_width))
    
    # Calculate the actual bounds including underline
    actual_bottom = line_y + 5 * scale_factor
    actual_right = x + text_width + 30 * scale_factor
    
    # Crop to actual content with some padding, but ensure full signature is visible
    img = img.crop((0, 0, min(scaled_width, max(int(actual_right), int(text_width) + padding * 2)), 
                    min(scaled_height, max(int(actual_bottom), int(text_height) + padding * 2))))
    
    # Scale down using high-quality resampling for sharp, clear output
    final_width = img.size[0] // scale_factor
    final_height = img.size[1] // scale_factor
    img = img.resize((final_width, final_height), PILImage.Resampling.LANCZOS)
    
    return img

def generate_date_range(start_date, end_date, max_days=7):
    """Generate a list of dates from start_date to end_date, formatted as MM/DD/YY"""
    if not start_date or not end_date:
        return [''] * max_days
    
    dates = []
    current_date = start_date
    while current_date <= end_date and len(dates) < max_days:
        # Format as MM/DD/YY
        formatted_date = current_date.strftime('%m/%d/%y')
        dates.append(formatted_date)
        current_date += timedelta(days=1)
    
    # Fill remaining slots with empty strings
    while len(dates) < max_days:
        dates.append('')
    
    return dates

def create_pdf(form_data, ws):
    """Create PDF with form data and red highlighting"""
    meal_deductions = {
        68: { 'breakfast': 16, 'lunch': 19, 'dinner': 28, 'incidental': 5, 'first_last': 51.00 },
        74: { 'breakfast': 18, 'lunch': 20, 'dinner': 31, 'incidental': 5, 'first_last': 55.50 },
        80: { 'breakfast': 20, 'lunch': 22, 'dinner': 33, 'incidental': 5, 'first_last': 60.00 },
        86: { 'breakfast': 22, 'lunch': 23, 'dinner': 36, 'incidental': 5, 'first_last': 64.50 },
        92: { 'breakfast': 23, 'lunch': 26, 'dinner': 38, 'incidental': 5, 'first_last': 69.00 },
    }
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        alignment=1  # Center
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    # Helper to load, trim and whiten logo from URL and size by target height
    def load_logo_image(url: str, target_height_inch: float):
        try:
            with urllib.request.urlopen(url) as resp:
                data = resp.read()
            img_pil = PILImage.open(io.BytesIO(data)).convert('RGB')
            # Replace near-black background with white and trim borders
            pixels = img_pil.load()
            width, height = img_pil.size
            # Replace very dark pixels with white to avoid giant black boxes
            for y in range(height):
                for x in range(width):
                    r, g, b = pixels[x, y]
                    if r < 20 and g < 20 and b < 20:
                        pixels[x, y] = (255, 255, 255)
            # Create trim mask for white background to crop extra whitespace
            gray = img_pil.convert('L')
            # Inverse mask of non-white areas
            mask = gray.point(lambda p: 0 if p > 250 else 255)
            bbox = mask.getbbox()
            if bbox:
                img_pil = img_pil.crop(bbox)
            # Scale by target height
            target_h = target_height_inch * inch
            w, h = img_pil.size
            aspect = w / h if h else 1.0
            target_w = target_h * aspect
            buf = io.BytesIO()
            img_pil.save(buf, format='PNG')
            buf.seek(0)
            return Image(buf, width=target_w, height=target_h)
        except Exception:
            return None

    # Logos beside title
    georgetown_logo_url = 'https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png'
    advance_logo_url = 'https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/ADVANCE%20Logo_Horizontal%20Blue.png'

    left_logo = load_logo_image(georgetown_logo_url, target_height_inch=0.8)
    right_logo = load_logo_image(advance_logo_url, target_height_inch=0.4)

    title_para = Paragraph("Domestic Travel Authorization Form", title_style)

    title_block = [[title_para]]
    title_table = Table(title_block)
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Build a clean 3-column header row
    # Reserve ~1.3in for each logo, center column takes remaining width
    content_width = 8.5*inch - (0.5*inch + 0.5*inch)
    left_w = 1.3*inch
    right_w = 1.0*inch
    center_w = max(content_width - (left_w + right_w), 3.5*inch)
    header_row = [left_logo if left_logo else '', title_table, right_logo if right_logo else '']
    header_table = Table([header_row], colWidths=[left_w, center_w, right_w])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    story.append(header_table)
    # Thin rule below header
    story.append(Spacer(1, 0.05*inch))
    story.append(Table([["" ]], colWidths=[content_width], rowHeights=[0.5]))    
    story.append(Spacer(1, 0.1*inch))
    
    # Traveler Information Section
    story.append(Paragraph("<b>Traveler Information</b>", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    # Create traveler info table
    traveler_data = [
        ['Name', form_data.get('name', ''), 'Organization', form_data.get('organization', 'Georgetown University')],
        ['Address Line 1', form_data.get('address1', ''), 'Destination', form_data.get('destination', '')],
        ['Address Line 2', form_data.get('address2', ''), 'Departure Date', form_data.get('departure_date', '')],
        ['City', form_data.get('city', ''), 'Return Date', form_data.get('return_date', '')],
        ['State', form_data.get('state', ''), 'Email Address', form_data.get('email', '')],
        ['Zip', form_data.get('zip', ''), '', '']
    ]
    
    traveler_table = Table(traveler_data, colWidths=[1.5*inch, 1.8*inch, 1.5*inch, 1.8*inch])
    traveler_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E0E0E0')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.red),  # Red text for input fields
        ('TEXTCOLOR', (3, 0), (3, -1), colors.red),  # Red text for input fields
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FFEBEE')),  # Light red background
        ('BACKGROUND', (3, 0), (3, -1), colors.HexColor('#FFEBEE')),  # Light red background
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(traveler_table)
    story.append(Spacer(1, 0.1*inch))
    
    # Purpose of Travel Section
    story.append(Paragraph("<b>Purpose of Travel</b>", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    purpose_of_travel = form_data.get('purpose_of_travel', '')
    objective = form_data.get('objective', '')
    attendees = form_data.get('attendees', '')
    deliverables = form_data.get('deliverables', '')
    support_files = form_data.get('support_files', '')
    
    # Create purpose of travel table with Paragraph for text wrapping
    purpose_data = []
    purpose_style = ParagraphStyle(
        'PurposeStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.red,
        leading=11,
    )
    
    if purpose_of_travel:
        purpose_data.append(['Purpose of Travel', Paragraph(purpose_of_travel, purpose_style)])
    if objective and objective.strip():
        purpose_data.append(['Objective', Paragraph(objective, purpose_style)])
    if attendees:
        purpose_data.append(['Attendees', Paragraph(attendees, purpose_style)])
    if deliverables:
        purpose_data.append(['Deliverables', Paragraph(deliverables, purpose_style)])
    if support_files and support_files.strip():
        purpose_data.append(['Support Materials', Paragraph(support_files, purpose_style)])
    
    if purpose_data:
        purpose_table = Table(purpose_data, colWidths=[2*inch, 5.5*inch])
        purpose_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FFEBEE')),  # Light red background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (0, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(purpose_table)
        story.append(Spacer(1, 0.1*inch))
    
    # Traveler Paid Expenses Section
    story.append(Paragraph("<b>Traveler Paid Expenses</b>", styles['Heading2']))
    
    # Mileage Section
    story.append(Paragraph("<b>Mileage</b>", styles['Heading3']))
    story.append(Paragraph("The Mileage (Per Day) should be rounded to the nearest mile.", styles['Normal']))
    story.append(Paragraph("Mileage for 2025 is $0.70 per mile.", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    
    # Mileage: build multiple tables, 7 days per table
    all_mileage_dates = form_data.get('mileage_dates', [])
    all_mileage_amounts = form_data.get('mileage_amounts', [])
    # Grand total for mileage rate across all days
    grand_mileage_rate_total = 0.0
    for amount in all_mileage_amounts:
        if amount and str(amount).strip():
            try:
                grand_mileage_rate_total += round(float(amount) * 0.70, 2)
            except:
                pass
    grand_mileage_rate_total = round(grand_mileage_rate_total, 0)
    mileage_tables = []
    mileage_dates_chunks = chunk_list(all_mileage_dates, 7)
    mileage_amount_chunks = chunk_list(all_mileage_amounts, 7)
    total_mileage_chunks = len(mileage_dates_chunks)
    for idx in range(total_mileage_chunks):
        dates_chunk = mileage_dates_chunks[idx] if idx < len(mileage_dates_chunks) else []
        amounts_chunk = mileage_amount_chunks[idx] if idx < len(mileage_amount_chunks) else []
        dates_chunk = pad_to_length(dates_chunk, 7, '')
        amounts_chunk = pad_to_length(amounts_chunk, 7, '')
        mileage_data = [['Date (MM/DD/YY)'] + dates_chunk + ['Total']]
        mileage_data.append(['MILEAGE (Per Day)'] + [str(x) if x else '' for x in amounts_chunk] + [''])
        # Calculate mileage rates per chunk
        mileage_rates = []
        for amount in amounts_chunk:
            if amount and str(amount).strip():
                try:
                    rate = round(float(amount) * 0.70, 2)
                    mileage_rates.append(f"${int(rate)}")
                except:
                    mileage_rates.append('')
            else:
                mileage_rates.append('')
        # Only last table shows grand total; others blank
        mileage_total_cell = f"${int(grand_mileage_rate_total)}" if idx == total_mileage_chunks - 1 else ''
        mileage_data.append(['Mileage Rate'] + mileage_rates + [mileage_total_cell])
        mileage_table = Table(mileage_data, colWidths=[1.3*inch] + [0.7*inch]*7 + [0.75*inch])
        mileage_table.setStyle(TableStyle([
            # Headers and date row set to white for print
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (1, 1), (7, 1), colors.red), 
            ('BACKGROUND', (1, 1), (7, 1), colors.HexColor('#FFF5F5')),
            ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (1, 2), (7, 2), colors.red), 
            ('BACKGROUND', (1, 2), (7, 2), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TEXTCOLOR', (8, 2), (8, 2), colors.red),
            ('BACKGROUND', (8, 2), (8, 2), colors.HexColor('#FFF5F5')),
        ]))
        mileage_tables.append(mileage_table)
    for t in mileage_tables:
        story.append(t)
        story.append(Spacer(1, 0.1*inch))
    
    # Expenses Section - 7 days + total column
    story.append(Paragraph("<b>Airfare, Transportation, Parking, Lodging, Miscellaneous.</b>", styles['Heading3']))
    story.append(Paragraph("Ground Transportation Includes: Taxi, Uber, etc.", styles['Normal']))
    story.append(Paragraph("Miscellaneous/Other: Pre-approved travel expenses not listed in this form", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    expense_dates = form_data.get('expense_dates', [])
    airfare = form_data.get('airfare', [])
    ground_transport = form_data.get('ground_transport', [])
    parking = form_data.get('parking', [])
    lodging = form_data.get('lodging', [])
    baggage = form_data.get('baggage', [])
    misc = form_data.get('misc', [])
    misc2 = form_data.get('misc2', [])  # Second row for misc expenses
    
    # Build labels; only use if descriptions are actually provided
    misc_desc1_val = form_data.get('misc_desc1', '').strip() if form_data.get('misc_desc1', '') else ''
    misc_desc2_val = form_data.get('misc_desc2', '').strip() if form_data.get('misc_desc2', '') else ''
    
    # Create Paragraph style for misc labels that allows text wrapping
    # Column width is 1.3*inch, so set width slightly less to account for padding
    misc_label_style = ParagraphStyle(
        'MiscLabelStyle',
        parent=styles['Normal'],
        fontSize=8,
        fontName='Helvetica',
        alignment=0,  # LEFT
        leading=10,  # Line spacing
        leftIndent=0,
        rightIndent=0,
    )
    
    # Convert misc labels to Paragraph objects for text wrapping
    # Only show misc rows that have actual descriptions
    misc_label1 = Paragraph(misc_desc1_val, misc_label_style) if misc_desc1_val else None
    misc_label2 = Paragraph(misc_desc2_val, misc_label_style) if misc_desc2_val else None
    
    # Grand totals across all days
    grand_af = sum(x for x in airfare if x)
    grand_gt = sum(x for x in ground_transport if x)
    grand_pk = sum(x for x in parking if x)
    grand_lg = sum(x for x in lodging if x)
    grand_bg = sum(x for x in baggage if x)
    grand_m1 = sum(x for x in misc if x)
    grand_m2 = sum(x for x in misc2 if x)

    expense_tables = []
    expense_chunks = chunk_list(expense_dates, 7)
    total_expense_chunks = len(expense_chunks)
    for i, dates_chunk in enumerate(expense_chunks):
        chunk_len = len(pad_to_length(dates_chunk, 7))
        pad_len = 7
        af = pad_to_length(airfare[i*7:(i+1)*7], pad_len, 0)
        gt = pad_to_length(ground_transport[i*7:(i+1)*7], pad_len, 0)
        pk = pad_to_length(parking[i*7:(i+1)*7], pad_len, 0)
        lg = pad_to_length(lodging[i*7:(i+1)*7], pad_len, 0)
        bg = pad_to_length(baggage[i*7:(i+1)*7], pad_len, 0)
        m1 = pad_to_length(misc[i*7:(i+1)*7], pad_len, 0)
        m2 = pad_to_length(misc2[i*7:(i+1)*7], pad_len, 0)
        # Build expenses data - only include misc rows if descriptions are provided
        expenses_data = [
            ['Date (MM/DD/YY)'] + pad_to_length(dates_chunk, 7, '') + ['Total'],
            ['Airfare'] + [f"${x:.2f}" if x else '' for x in af] + ([f"${grand_af:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Ground Transportation'] + [f"${x:.2f}" if x else '' for x in gt] + ([f"${grand_gt:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Parking'] + [f"${x:.2f}" if x else '' for x in pk] + ([f"${grand_pk:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Lodging'] + [f"${x:.2f}" if x else '' for x in lg] + ([f"${grand_lg:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Baggage Fees'] + [f"${x:.2f}" if x else '' for x in bg] + ([f"${grand_bg:.2f}"] if i == total_expense_chunks - 1 else ['']),
            ['Miscellaneous/Other\n(Provide Description)'] + [''] * 7 + [''],
        ]
        
        # Only add misc rows if descriptions are provided
        misc_row1_idx = None
        misc_row2_idx = None
        if misc_label1 is not None:
            expenses_data.append([misc_label1] + [f"${x:.2f}" if x else '' for x in m1] + ([f"${grand_m1:.2f}"] if i == total_expense_chunks - 1 else ['']))
            misc_row1_idx = len(expenses_data) - 1
        if misc_label2 is not None:
            expenses_data.append([misc_label2] + [f"${x:.2f}" if x else '' for x in m2] + ([f"${grand_m2:.2f}"] if i == total_expense_chunks - 1 else ['']))
            misc_row2_idx = len(expenses_data) - 1
        
        expenses_table = Table(expenses_data, colWidths=[1.3*inch] + [0.65*inch]*7 + [0.75*inch])
        
        # Build table style - dynamically handle misc rows
        table_style = [
            # Left label column light gray; headers white
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('TEXTCOLOR', (1, 1), (7, 5), colors.red),
            ('BACKGROUND', (1, 1), (7, 5), colors.HexColor('#FFF5F5')),
            ('SPAN', (0, 6), (7, 6)),  # Span across all day columns
            ('TEXTCOLOR', (0, 6), (0, 6), colors.black),
            ('BACKGROUND', (0, 6), (7, 6), colors.white),
        ]
        
        # Add styling for misc rows only if they exist
        if misc_row1_idx is not None:
            table_style.extend([
                ('TEXTCOLOR', (0, misc_row1_idx), (0, misc_row1_idx), colors.black),
                ('BACKGROUND', (0, misc_row1_idx), (0, misc_row1_idx), colors.HexColor('#E0E0E0')),
                ('TEXTCOLOR', (1, misc_row1_idx), (7, misc_row1_idx), colors.red),
                ('BACKGROUND', (1, misc_row1_idx), (7, misc_row1_idx), colors.HexColor('#FFF5F5')),
                ('TEXTCOLOR', (8, misc_row1_idx), (8, misc_row1_idx), colors.red),
                ('BACKGROUND', (8, misc_row1_idx), (8, misc_row1_idx), colors.HexColor('#FFF5F5')),
            ])
        if misc_row2_idx is not None:
            table_style.extend([
                ('TEXTCOLOR', (0, misc_row2_idx), (0, misc_row2_idx), colors.black),
                ('BACKGROUND', (0, misc_row2_idx), (0, misc_row2_idx), colors.HexColor('#E0E0E0')),
                ('TEXTCOLOR', (1, misc_row2_idx), (7, misc_row2_idx), colors.red),
                ('BACKGROUND', (1, misc_row2_idx), (7, misc_row2_idx), colors.HexColor('#FFF5F5')),
                ('TEXTCOLOR', (8, misc_row2_idx), (8, misc_row2_idx), colors.red),
                ('BACKGROUND', (8, misc_row2_idx), (8, misc_row2_idx), colors.HexColor('#FFF5F5')),
            ])
        
        # Add common styling for totals column
        last_row = len(expenses_data) - 1
        table_style.extend([
            ('TEXTCOLOR', (8, 1), (8, last_row), colors.red),
            ('BACKGROUND', (8, 1), (8, last_row), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (7, -1), 'CENTER'),
            ('ALIGN', (8, 0), (8, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
        ])
        
        expenses_table.setStyle(TableStyle(table_style))
        expense_tables.append(expenses_table)
    for t in expense_tables:
        story.append(t)
        story.append(Spacer(1, 0.1*inch))
    
    # Meals and Incidentals Section
    story.append(Paragraph("<b>Meals and Incidentals Per Diem</b>", styles['Heading3']))
    story.append(Paragraph("Federal Guidelines: On the first and last travel day, travelers are only eligible for 75 percent of the total M&amp;IE rate.", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    
    per_diem_dates = form_data.get('per_diem_dates', [])
    per_diem_amounts = form_data.get('per_diem_amounts', [])  # Will be one of PER_DIEM_OPTIONS
    breakfast_checks = form_data.get('breakfast_checks', [])
    lunch_checks = form_data.get('lunch_checks', [])
    dinner_checks = form_data.get('dinner_checks', [])
    
    # Calculate adjusted per diem for each day using dollar-based deductions
    adjusted_per_diem = []
    daily_totals = []
    
    # Find which days have dates (non-empty)
    days_with_dates = [i for i, d in enumerate(per_diem_dates) if d and str(d).strip()]
    num_days = len(days_with_dates)
    first_day_idx = days_with_dates[0] if days_with_dates else 0
    last_day_idx = days_with_dates[-1] if days_with_dates else 0
    
    for i in range(len(per_diem_dates)):
        if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
            base_per_diem = int(per_diem_amounts[i]) if (i < len(per_diem_amounts) and per_diem_amounts[i]) else 80
            deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
            deduction_total = 0.0
            if i < len(breakfast_checks) and breakfast_checks[i]:
                deduction_total += deducts['breakfast']
            if i < len(lunch_checks) and lunch_checks[i]:
                deduction_total += deducts['lunch']
            if i < len(dinner_checks) and dinner_checks[i]:
                deduction_total += deducts['dinner']
            # Base already includes incidentals; do not add +$5 here
            pre75_total = max(0.0, float(base_per_diem) - deduction_total)
            # Apply 75% for first and last day
            if i == first_day_idx or i == last_day_idx:
                final_per_diem = round(pre75_total * 0.75, 2)
            else:
                final_per_diem = round(pre75_total, 2)
            
            adjusted_per_diem.append(final_per_diem)
            daily_totals.append(final_per_diem)
        else:
            adjusted_per_diem.append(0.0)
            daily_totals.append(0.0)
    
    total_per_diem_calculated = sum(daily_totals)
    
    # Calculate daily meal totals (before 75% reduction) and total dollar reductions
    daily_meal_totals = []
    total_reductions = []
    for i in range(len(per_diem_dates)):
        if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
            base_per_diem = int(per_diem_amounts[i]) if per_diem_amounts[i] else 80
            deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
            deduction_total = 0.0
            if i < len(breakfast_checks) and breakfast_checks[i]:
                deduction_total += deducts['breakfast']
            if i < len(lunch_checks) and lunch_checks[i]:
                deduction_total += deducts['lunch']
            if i < len(dinner_checks) and dinner_checks[i]:
                deduction_total += deducts['dinner']
            total_reductions.append(round(deduction_total, 2))
            pre75_total = max(0.0, float(base_per_diem) - deduction_total)
            daily_meal_totals.append(round(pre75_total, 2))
        else:
            total_reductions.append(0.0)
            daily_meal_totals.append(0.0)
    
    # Build per diem tables per 7-day chunk
    per_diem_tables = []
    for i, dates_chunk in enumerate(chunk_list(per_diem_dates, 7)):
        idx_start = i * 7
        pd = pad_to_length(dates_chunk, 7, '')
        amounts = pad_to_length(per_diem_amounts[idx_start:idx_start+7], 7, 80)
        bchk = pad_to_length(breakfast_checks[idx_start:idx_start+7], 7, False)
        lchk = pad_to_length(lunch_checks[idx_start:idx_start+7], 7, False)
        dchk = pad_to_length(dinner_checks[idx_start:idx_start+7], 7, False)
        # Map totals slice
        red = pad_to_length(total_reductions[idx_start:idx_start+7], 7, 0.0)
        meal_tot = pad_to_length(daily_meal_totals[idx_start:idx_start+7], 7, 0.0)
        adj = pad_to_length(adjusted_per_diem[idx_start:idx_start+7], 7, 0.0)
        # Determine labels for deductions in this chunk: use the per diem amount from actual dates
        # Find the first per diem amount that corresponds to a date (non-empty date)
        common_amount = None
        for j in range(len(pd)):
            if pd[j] and str(pd[j]).strip() and j < len(amounts) and amounts[j]:
                try:
                    common_amount = int(amounts[j])
                    break
                except (ValueError, TypeError):
                    continue
        # If no date found in this chunk, try to get from the actual per_diem_amounts (not padded)
        if common_amount is None and idx_start < len(per_diem_amounts):
            for j in range(idx_start, min(idx_start + 7, len(per_diem_amounts))):
                if j < len(per_diem_dates) and per_diem_dates[j] and str(per_diem_dates[j]).strip():
                    try:
                        common_amount = int(per_diem_amounts[j])
                        break
                    except (ValueError, TypeError):
                        continue
        # Default to 80 if still not found
        if common_amount is None:
            common_amount = 80
        
        if common_amount in meal_deductions:
            b_lbl = f"Breakfast -${meal_deductions[common_amount]['breakfast']}"
            l_lbl = f"Lunch -${meal_deductions[common_amount]['lunch']}"
            d_lbl = f"Dinner -${meal_deductions[common_amount]['dinner']}"
        else:
            b_lbl = "Breakfast -$"
            l_lbl = "Lunch -$"
            d_lbl = "Dinner -$"

        per_diem_data = [
            ['Date (MM/DD/YY)'] + [d if d and str(d).strip() else '' for d in pd] + [''],
            ['Per Diem Allowance'] + [f"${int(x)}" if (x and pd[j]) else '' for j, x in enumerate(amounts)] + [''],
            ['ADJUSTED PER DIEM', 'If meals were provided (Place "x" in box)', '', '', '', '', '', ''],
            [b_lbl] + ['X' if (bchk[j] and pd[j]) else '' for j in range(7)] + [''],
            [l_lbl] + ['X' if (lchk[j] and pd[j]) else '' for j in range(7)] + [''],
            [d_lbl] + ['X' if (dchk[j] and pd[j]) else '' for j in range(7)] + [''],
            ['Total Reduction ($)'] + [f"${x:.2f}" if x != 0 else '' for x in red] + [''],
            ['Daily Meal Total'] + [f"${x:.2f}" if x > 0 else '' for x in meal_tot] + [''],
            ['Total Per Diem'] + [f"${x:.2f}" if x > 0 else '' for x in adj] + [f"${total_per_diem_calculated:.2f}" if i == len(chunk_list(per_diem_dates,7)) - 1 else ''],
        ]
        per_diem_table = Table(per_diem_data, colWidths=[1.3*inch] + [0.6*inch]*7 + [0.75*inch])
        per_diem_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E0E0')),
            # Date row and header white
            ('BACKGROUND', (1, 0), (7, 0), colors.white),
            ('TEXTCOLOR', (1, 0), (7, 0), colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (1, 1), (-1, -1), colors.red),
            ('BACKGROUND', (1, 1), (-1, -1), colors.HexColor('#FFF5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('SPAN', (0, 2), (7, 2)),
            ('BACKGROUND', (0, 2), (7, 2), colors.HexColor('#FFF5F5')),
            ('TEXTCOLOR', (0, 2), (7, 2), colors.red),
            ('TEXTCOLOR', (8, -1), (8, -1), colors.red),
            ('BACKGROUND', (8, -1), (8, -1), colors.HexColor('#FFF5F5')),
        ]))
        per_diem_tables.append(per_diem_table)
    for t in per_diem_tables:
        story.append(t)
        story.append(Spacer(1, 0.1*inch))
    
    # Totals Section
    total_mileage = form_data.get('total_mileage', 0)
    total_airfare = form_data.get('total_airfare', 0)
    total_ground_transport = form_data.get('total_ground_transport', 0)
    total_parking = form_data.get('total_parking', 0)
    total_lodging = form_data.get('total_lodging', 0)
    total_baggage = form_data.get('total_baggage', 0)
    total_misc = form_data.get('total_misc', 0)
    total_per_diem = form_data.get('total_per_diem', 0)
    
    # Calculate subtotal
    subtotal = total_mileage + total_airfare + total_ground_transport + total_parking + total_lodging + total_baggage + total_misc + total_per_diem
    total_amount_due = subtotal
    
    # Ensure total_amount_due is not negative
    total_amount_due = max(0, total_amount_due)

    story.append(Paragraph("<b>Sub-Totals</b>", styles['Heading3']))

    totals_data = [
        ['Mileage', f"${int(total_mileage)}"],
        ['Airfare', f"${total_airfare:.2f}"],
        ['Ground Transportation', f"${total_ground_transport:.2f}"],
        ['Parking', f"${total_parking:.2f}"],
        ['Lodging', f"${total_lodging:.2f}"],
        ['Baggage Fees', f"${total_baggage:.2f}"],
        ['Miscellaneous/Other', f"${total_misc:.2f}"],
        ['Per Diem', f"${total_per_diem:.2f}"],
        [Paragraph('<b>Total Amount Due</b>', styles['Normal']), Paragraph(f'<b>${total_amount_due:.2f}</b>', styles['Normal'])]
    ]
    
    totals_table = Table(totals_data, colWidths=[3*inch, 1.5*inch])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (1, 7), colors.red),
        ('BACKGROUND', (1, 0), (1, 7), colors.HexColor('#FFF5F5')),
        ('BACKGROUND', (0, 8), (0, 8), colors.white),
        ('BACKGROUND', (1, 8), (1, 8), colors.HexColor('#FFF5F5')),
        ('TEXTCOLOR', (1, 8), (1, 8), colors.red),
        ('FONTNAME', (0, 8), (0, 8), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.white]),
        ('TEXTCOLOR', (0, 9), (1, 9), colors.red),
        ('BACKGROUND', (0, 9), (1, 9), colors.white),
        ('FONTNAME', (0, 9), (1, 9), 'Helvetica-Bold'),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 0.1*inch))


    story.append(Paragraph("<b>Approval Signatures</b>", styles['Heading2']))
    # Signature section
    signature_text = form_data.get('signature', '').strip()
    
    # Create signature cell with image or text
    signature_cell_value = ''
    
    # Generate signature image from text
    if signature_text:
        try:
            # Generate signature image from text with high resolution (3x scale)
            signature_img_pil = generate_signature_image(signature_text, width=800, height=150, scale_factor=3)
            
            if signature_img_pil:
                # Ensure it's RGB with white background (blank)
                if signature_img_pil.mode != 'RGB':
                    rgb_img = PILImage.new('RGB', signature_img_pil.size, (255, 255, 255))
                    if signature_img_pil.mode == 'RGBA':
                        rgb_img.paste(signature_img_pil, mask=signature_img_pil.split()[3])
                    else:
                        rgb_img.paste(signature_img_pil)
                    signature_img_pil = rgb_img
                
                # Resize signature to fit the table cell (cell width is 2 inches, accounting for padding)
                # Cell has 6pt left/right padding, so available width is ~1.88 inches
                max_width = 1.88 * inch
                max_height = 0.5 * inch  # Reduced height to fit better in cell
                
                img_width, img_height = signature_img_pil.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1
                
                # Calculate size maintaining aspect ratio but respecting both max width and height
                new_width = min(img_width, max_width)
                new_height = new_width * aspect_ratio
                
                if new_height > max_height:
                    new_height = max_height
                    new_width = new_height / aspect_ratio
                
                # Ensure width doesn't exceed cell width
                new_width = min(new_width, max_width)
                
                # Save to buffer for ReportLab with high quality
                img_buffer = io.BytesIO()
                # Save at full resolution for maximum clarity
                signature_img_pil.save(img_buffer, format='PNG', optimize=False, compress_level=1)
                img_buffer.seek(0)
                
                # Create ReportLab Image - use the calculated dimensions
                # The image will be high-res internally but displayed at the correct size
                signature_img = Image(img_buffer, width=new_width, height=new_height)
                signature_cell_value = signature_img
            else:
                signature_cell_value = signature_text
        except Exception as e:
            # Fallback to text if image generation fails
            signature_cell_value = signature_text
    
    # Combined Approval Signatures and Operations Use Only table
    # Use Paragraph for all labels to ensure consistent font size and width
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica',
        alignment=0,  # LEFT
    )
    
    traveler_label = Paragraph("Traveler Signature", label_style)
    program_assistant_label = Paragraph("Program Assistant", label_style)
    lead_provider_text = Paragraph("Lead Technical\nAssistance Provider", label_style)
    
    # Helper function to format dates consistently to MM/DD/YYYY
    def format_date_for_pdf(date_value):
        """Format date to MM/DD/YYYY format consistently"""
        if not date_value or date_value == '':
            return ''
        try:
            # If it's already in MM/DD/YYYY format, return as-is
            if isinstance(date_value, str) and '/' in date_value:
                # Check if it's already MM/DD/YYYY
                parts = date_value.split('/')
                if len(parts) == 3:
                    # Validate it's MM/DD/YYYY (not DD/MM/YYYY)
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2 and len(parts[2]) == 4:
                        return date_value
                    elif len(parts[2]) == 4:
                        # Might be MM/DD/YYYY, return as-is
                        return date_value
            # Try parsing as YYYY-MM-DD format
            if isinstance(date_value, str) and '-' in date_value:
                try:
                    parsed_date = datetime.strptime(date_value, '%Y-%m-%d')
                    return parsed_date.strftime('%m/%d/%Y')
                except:
                    pass
            # Try parsing as date object
            if hasattr(date_value, 'strftime'):
                return date_value.strftime('%m/%d/%Y')
            # Try parsing various string formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d']:
                try:
                    parsed_date = datetime.strptime(str(date_value), fmt)
                    return parsed_date.strftime('%m/%d/%Y')
                except:
                    continue
            # If all parsing fails, return as string
            return str(date_value)
        except:
            return str(date_value) if date_value else ''
    
    # Get coordinator signatures and dates if available
    mabintou_sig_text = form_data.get('mabintou_signature', '').strip()
    kemisha_sig_text = form_data.get('kemisha_signature', '').strip()
    mabintou_date = format_date_for_pdf(form_data.get('mabintou_approval_date', ''))
    kemisha_date = format_date_for_pdf(form_data.get('kemisha_approval_date', ''))
    traveler_date = format_date_for_pdf(form_data.get('signature_date', ''))
    
    # Generate Mabintou signature image (Program Assistant)
    mabintou_signature_cell = ''
    if mabintou_sig_text:
        try:
            mabintou_img_pil = generate_signature_image(mabintou_sig_text, width=800, height=150, scale_factor=3)
            if mabintou_img_pil:
                if mabintou_img_pil.mode != 'RGB':
                    rgb_mabintou = PILImage.new('RGB', mabintou_img_pil.size, (255, 255, 255))
                    if mabintou_img_pil.mode == 'RGBA':
                        rgb_mabintou.paste(mabintou_img_pil, mask=mabintou_img_pil.split()[3])
                    else:
                        rgb_mabintou.paste(mabintou_img_pil)
                    mabintou_img_pil = rgb_mabintou
                
                max_width = 1.88 * inch
                max_height = 0.5 * inch
                img_width, img_height = mabintou_img_pil.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1
                new_width = min(img_width, max_width)
                new_height = new_width * aspect_ratio
                if new_height > max_height:
                    new_height = max_height
                    new_width = new_height / aspect_ratio
                new_width = min(new_width, max_width)
                
                mabintou_buffer = io.BytesIO()
                mabintou_img_pil.save(mabintou_buffer, format='PNG', optimize=False, compress_level=1)
                mabintou_buffer.seek(0)
                mabintou_signature_cell = Image(mabintou_buffer, width=new_width, height=new_height)
        except Exception:
            mabintou_signature_cell = mabintou_sig_text
    
    # Generate Kemisha signature image (Lead Technical Assistance Provider)
    kemisha_signature_cell = ''
    if kemisha_sig_text:
        try:
            kemisha_img_pil = generate_signature_image(kemisha_sig_text, width=800, height=150, scale_factor=3)
            if kemisha_img_pil:
                if kemisha_img_pil.mode != 'RGB':
                    rgb_kemisha = PILImage.new('RGB', kemisha_img_pil.size, (255, 255, 255))
                    if kemisha_img_pil.mode == 'RGBA':
                        rgb_kemisha.paste(kemisha_img_pil, mask=kemisha_img_pil.split()[3])
                    else:
                        rgb_kemisha.paste(kemisha_img_pil)
                    kemisha_img_pil = rgb_kemisha
                
                max_width = 1.88 * inch
                max_height = 0.5 * inch
                img_width, img_height = kemisha_img_pil.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1
                new_width = min(img_width, max_width)
                new_height = new_width * aspect_ratio
                if new_height > max_height:
                    new_height = max_height
                    new_width = new_height / aspect_ratio
                new_width = min(new_width, max_width)
                
                kemisha_buffer = io.BytesIO()
                kemisha_img_pil.save(kemisha_buffer, format='PNG', optimize=False, compress_level=1)
                kemisha_buffer.seek(0)
                kemisha_signature_cell = Image(kemisha_buffer, width=new_width, height=new_height)
        except Exception:
            kemisha_signature_cell = kemisha_sig_text
    
    combined_data = [
        [traveler_label, signature_cell_value, 'DATE', traveler_date],
        [program_assistant_label, mabintou_signature_cell, 'DATE', mabintou_date],
        [lead_provider_text, kemisha_signature_cell, 'DATE', kemisha_date],
        ['AWD', 'AWD-7776588', 'GR', 'GR426936'],
    ]
    
    combined_table = Table(combined_data, colWidths=[1.5*inch, 2*inch, 0.8*inch, 1.5*inch])
    combined_table.setStyle(TableStyle([
        # Grid and alignment for all rows
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        # Padding for all rows
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        # Traveler Signature row (row 0) - all white background, signature/date cells red text
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.red),
        ('TEXTCOLOR', (3, 0), (3, 0), colors.red),
        # Program Assistant row (row 1) - signature and date cells red if signature exists
        ('BACKGROUND', (0, 1), (-1, 1), colors.white),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.red),
        ('TEXTCOLOR', (3, 1), (3, 1), colors.red),
        # Lead Technical Assistance Provider row (row 2) - signature and date cells red if signature exists
        ('BACKGROUND', (0, 2), (-1, 2), colors.white),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.red),
        ('TEXTCOLOR', (3, 2), (3, 2), colors.red),
        # Operations row (row 3) - all white background
        ('BACKGROUND', (0, 3), (-1, 3), colors.white),
    ]))
    story.append(combined_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def upload_file_to_drive(file, filename, folder_id, creds_dict):
    # Convert Streamlit secret dict into Google Credentials object
    drive_creds = Credentials.from_service_account_info(creds_dict, scopes=[
        "https://www.googleapis.com/auth/drive"
    ])

    # Build the Drive API service
    drive_service = build('drive', 'v3', credentials=drive_creds)

    # Prepare file metadata
    file_metadata = {
        'name': filename,
        'parents': [folder_id]
    }

    # Read uploaded file and prepare media
    media = MediaIoBaseUpload(io.BytesIO(file.read()), mimetype=file.type)

    # Upload file
    uploaded = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()

    # Make the file public
    drive_service.permissions().create(
        fileId=uploaded['id'],
        body={'type': 'anyone', 'role': 'reader'}
    ).execute()

    # Return the sharable link
    return f"https://drive.google.com/file/d/{uploaded['id']}/view"


def _get_records_with_retry(spreadsheet_name, worksheet_name, retries=3, base_delay=0.5):
    """Fetch worksheet records with simple exponential backoff to mitigate 429s."""
    attempt = 0
    last_exc = None
    while attempt < retries:
        try:
            spreadsheet = client.open(spreadsheet_name)
            worksheet = spreadsheet.worksheet(worksheet_name)
            return worksheet.get_all_records()
        except Exception as exc:
            last_exc = exc
            delay = base_delay * (2 ** attempt)
            time.sleep(delay)
            attempt += 1
    # If all retries failed, re-raise last exception
    raise last_exc

@st.cache_data(ttl=600)
def load_main_sheet():
    df = pd.DataFrame(_get_records_with_retry('HRSA64_TA_Request', 'Main'))
    df['Submit Date'] = pd.to_datetime(df['Submit Date'], errors='coerce')
    df["Phone Number"] = df["Phone Number"].astype(str)
    return df

df = load_main_sheet()

# Ensure transfer-related columns exist
transfer_columns = [
    "Last Transfer From",
    "Last Transfer To",
    "Last Transfer Date",
    "Last Transfer By",
    "Transfer History",
]
for _col in transfer_columns:
    if _col not in df.columns:
        df[_col] = ""

# Ensure comment history columns exist
comment_history_columns = [
    "Coordinator Comment History",
    "Staff Comment History",
]
for _col in comment_history_columns:
    if _col not in df.columns:
        df[_col] = ""

@st.cache_data(ttl=600)
def load_interaction_sheet():
    return pd.DataFrame(_get_records_with_retry('HRSA64_TA_Request', 'Interaction'))

df_int = load_interaction_sheet()

# Ensure Interaction sheet has Jurisdiction column for no-ticket logs
if "Jurisdiction" not in df_int.columns:
    df_int["Jurisdiction"] = ""

@st.cache_data(ttl=600)
def load_delivery_sheet():
    return pd.DataFrame(_get_records_with_retry('HRSA64_TA_Request', 'Delivery'))

df_del = load_delivery_sheet()

@st.cache_data(ttl=600)
def load_support_sheet():
    return pd.DataFrame(_get_records_with_retry('HRSA64_TA_Request', 'GA_Support'))

df_support = load_support_sheet()

@st.cache_data(ttl=600)
def load_travel_sheet():
    return pd.DataFrame(_get_records_with_retry('HRSA64_TA_Request', 'Travel'))

df_travel = load_travel_sheet()

# Extract last Ticket ID from the existing sheet
last_ticket = df["Ticket ID"].dropna().astype(str).str.extract(r"GU(\d+)", expand=False).astype(int).max()
next_ticket_number = 1 if pd.isna(last_ticket) else last_ticket + 1
new_ticket_id = f"GU{next_ticket_number:04d}"


def format_phone(phone_str):
    # Remove non-digit characters
    digits = re.sub(r'\D', '', phone_str)

    # Format if it's 10 digits long
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    elif len(digits) == 11 and digits.startswith("1"):
        return f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    else:
        return phone_str  # Return original if not standard length

# Apply formatting
df["Phone Number"] = df["Phone Number"].astype(str).apply(format_phone)

# --- Demo user database
USERS = {
    "jw2104@georgetown.edu": {
        "Coordinator": {"password": "Qin88251216", "name": "Jiaqin Wu"},
        "Assignee/Staff": {"password": "Qin88251216", "name": "Jiaqin Wu"}
    },
    "Jenevieve.Opoku@georgetown.edu": {
        "Coordinator": {"password": "Tootles82!", "name": "Jenevieve Opoku"},
        "Assignee/Staff": {"password": "Tootles82!", "name": "Jenevieve Opoku"}
    },
    "me735@georgetown.edu": {
        "Coordinator": {"password": "me735hrsa64", "name": "Martine Etienne-Mesubi"},
        "Assignee/Staff": {"password": "me735hrsa64", "name": "Martine Etienne-Mesubi"}
    },
    "kd802@georgetown.edu": {
        "Coordinator": {"password": "kd802hrsa!!", "name": "Kemisha Denny"},
        "Assignee/Staff": {"password": "kd802hrsa!!", "name": "Kemisha Denny"}
    },
    "mo887@georgetown.edu": {
        "Coordinator": {"password": "Mabintou123!", "name": "Mabintou Ouattara"},
    },
    "lm1353@georgetown.edu": {
        "Coordinator": {"password": "LM1353hrsa64?", "name": "Lauren Mathae"}
    },
    "katherine.robsky@georgetown.edu": {
        "Coordinator": {"password": "Georgetown1", "name": "Katherine Robsky"},
        "Assignee/Staff": {"password": "Georgetown1", "name": "Katherine Robsky"}
    },
    "db1432@georgetown.edu": {
        "Assignee/Staff": {"password": "Deus123!", "name": "Deus Bazira"}
    },
    "sk2046@georgetown.edu": {
        "Assignee/Staff": {"password": "Sharon123!", "name": "Sharon Kibwana"}
    },
    "sgk23@georgetown.edu": {
        "Assignee/Staff": {"password": "Seble123!", "name": "Seble Kassaye"}
    },
    "weijun.yu@georgetown.edu": {
        "Assignee/Staff": {"password": "Weijun123!", "name": "Weijun Yu"}
    },
    "temesgen.zelalem@mayo.edu": {
        "Assignee/Staff": {"password": "Zelalem123!", "name": "Zelalem Temesgen"}
    },
    "carod@bu.edu": {
        "Assignee/Staff": {"password": "Carlos123!", "name": "Carlos Rodriguez-Diaz"}
    },
    "km2079@georgetown.edu": {
        "Assignee/Staff": {"password": "Kiah123!", "name": "Kiah Moorehead"}
    },
    "vd294@georgetown.edu": {
        "Assignee/Staff": {"password": "Vanessa123!", "name": "Vanessa Da Costa"}
    },
    "tm1649@georgetown.edu": {
        "Assignee/Staff": {"password": "Trena123!", "name": "Trena Mukherjee"}
    },
    "aj1202@georgetown.edu": {
        "Assignee/Staff": {"password": "Abby123!", "name": "Abby Jordan"}
    },
    "mh2504@georgetown.edu": {
        "Assignee/Staff": {"password": "Megan123!", "name": "Megan Highland"}
    },
    "jh2861@georgetown.edu": {
        "Assignee/Staff": {"password": "Jesus123!", "name": "Jesus Hernandez Burgos"}
    },
    "sc2710@georgetown.edu": {
        "Assignee/Staff": {"password": "Samantha123!", "name": "Samantha Cinnick"}
    },
    "bryan.shaw@georgetown.edu": {
        "Assignee/Staff": {"password": "Bryan123!", "name": "Bryan Shaw"}
    },
    "th1089@georgetown.edu": {
        "Assignee/Staff": {"password": "Tara123!", "name": "Tara Hixson"}
    },
    "da988@georgetown.edu":{
        "Assignee/Staff": {"password": "Dzifa123!", "name": "Dzifa Awunyo-Akaba"}
    },
    "mm5674@georgetown.edu":{
        "Assignee/Staff": {"password": "Masill123!", "name": "Masill Miranda"}
    },
    'jb3512@georgetown.edu':{
        "Assignee/Staff": {"password": "Joy123!", "name": "Joy Berry"}
    },
    'ac2992@georgetown.edu':{
        "Assignee/Staff": {"password": "Ashley123!", "name": "Ashley Clonchmore"}
    },
    'gh674@georgetown.edu':{
        "Assignee/Staff": {"password": "Grace123!", "name": "Grace Hazlett"}
    },
    'lw1035@georgetown.edu':{
        "Assignee/Staff": {"password": "Lauren123!", "name": "Lauren Wagner"}
    },
    'ew898@georgetown.edu':{
        "Assignee/Staff": {"password": "Eric123!", "name": "Eric Wagner"}
    },
    'htn16@georgetown.edu':{
        "Research Assistant": {"password": "Hang123!", "name": "Hang Nguyen"}
    },
    'ooa36@georgetown.edu':{
        "Research Assistant": {"password": "Olayinka123!", "name": "Olayinka Adedeji"}
    },
    'zs352@georgetown.edu':{
        "Research Assistant": {"password": "Ziqiao123!", "name": "Ziqiao Shan"}
    },
    'ap2349@georgetown.edu':{
        "Research Assistant": {"password": "Asha123!", "name": "Asha Patel"}
    },
    'sy803@georgetown.edu':{
        "Research Assistant": {"password": "Yannis123!", "name": "Yannis Ying"}
    },
    'ssb120@georgetown.edu':{
        "Research Assistant": {"password": "Saara123!", "name": "Saara Bidiwala"}
    },
    'sco47@georgetown.edu':{
        "Research Assistant": {"password": "Shedrack123!", "name": "Shedrack Osuji"}
    }
}

lis_location = ["Maricopa Co. - Arizona", "Alameda Co. - California", "Los Angeles Co. - California", "Orange Co. - California", "Riverside Co. - California",\
                "Sacramento Co. - California", "San Bernadino Co. -California", "San Diego Co. - California", "San Francisco Co. - California",\
                "Broward Co. - Florida", "Duval Co. - Florida", "Hillsborough Co. - Florida", "Miami-Dade Co. - Florida","Orange Co. - Florida",\
                "Palm Beach Co. - Florida", "Pinellas Co. - Florida", "Cobb Co. - Georgia", "Dekalb Co. - Georgia", "Fulton Co. - Georgia",\
                "Gwinnett Co. - Georgia", "Cook Co. - Illinois", "Marion Co. - Indiana", "East Baton Rough Parish - Louisiana",\
                "Orleans Parish - Louisiana", "Baltimore City - Maryland", "Montgomery Co. - Maryland", "Prince George's Co. - Maryland",\
                "Suffolk Co. - Massachusetts", "Wayne Co. - Michigan", "Clark Co. - Nevada", "Essex Co. - New Jersey","Hudson Co. - New Jersey",\
                "Bronx Co. - New York", "Kings Co. - New York", "New York Co. - New York", "Queens Co. - New York", "Mecklenburg Co. - North Carolina",\
                "Cuyahoga Co. - Ohio", "Franklin Co. - Ohio", "Hamilton Co. - Ohio", "Philadelphia Co. - Pennsylvania", "Shelby Co. - Tennessee",\
                "Bexar Co. - Texas", "Dallas Co. - Texas","Harris Co. - Texas", "Tarrant Co. - Texas","Travis Co. - Texas","King Co. - Washington",\
                "Washington, DC", "San Juan Municipio - Puerto Rico", "Alabama", "Arkansas","Kentucky","Mississippi","Missouri","Oklahoma","South Carolina"]

lis_organization = ["Maricopa County Public Health Department","Alameda County Public Health Department","Los Angeles County Department of Public Health",\
    "Orange County Health Care Agency","Riverside County Department of Public Health","Sacramento County Public Health","San Bernadino County Public Health",\
    "County of San Diego Health and Human Services Agency","San Francisco Department of Public Health","Broward County, Florida",\
    "Duval County, Florida","Hillsborough County, Florida","Miami-Dade County, Florida",\
    "Orange County, Florida","Palm Beach County, Florida","Pinellas County, Florida","Clark County Social Services, Office of HIV",\
    "Cobb and Douglas Public Health","Dekalb Public Health","Fulton County Government","Gwinnett County Board of Health","Cook County Health",\
    "Marion County Public Health Department","East Baton Rouge Parish Health Unit","New Orleans Health Department","Baltimore City Health Department",\
    "Montgomery County Health Department","Prince George's County Health Department","Boston Public Health Commission","Wayne County Health, Human and Veterans Services",\
    "Southern Nevada Health District","Essex County Department of Health","Hudson Regional Health Commission","New York City Department of Health and Mental Hygiene",\
    "Mecklenburg County Public Health","Cuyahoga County Board of Health","Franklin County Public Health","Ohio Department of Health","Philadelphia Department of Public Health",\
    "Shelby County Health Department","Bexar County Health and Human Services","Dallas County Health and Human Services","Harris County Public Health","Tarrant County Public Health",\
    "Travis County Health and Human Services","King Couty Public Health","San Juan Municipality Department of Health","Alabama Department of Public Health","Arkansas Department of Public Health",\
    "Kentucky Department for Public Health","Mississippi State Department of Health","Missouri Department of Health and Senior Services","Oklahoma State Department of Health","South Carolina Department of Public Health"]


# --- Initialize session state
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "role" not in st.session_state:
    st.session_state.role = None
if "user_email" not in st.session_state:
    st.session_state.user_email = ""

# --- Role selection
if st.session_state.role is None:
    #st.image("Georgetown_logo_blueRGB.png",width=200)
    #st.title("Welcome to the GU Technical Assistance Provider System")
    st.markdown(
        """
        <div style='
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            background: #f8f9fa;
            padding: 2em 0 1em 0;
            border-radius: 18px;
            box-shadow: 0 4px 24px rgba(0,0,0,0.07);
            margin-bottom: 2em;
        '>
            <img src='https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png' width='200' style='margin-bottom: 1em;'/>
            <h1 style='
                color: #1a237e;
                font-family: "Segoe UI", "Arial", sans-serif;
                font-weight: 700;
                margin: 0;
                font-size: 2.2em;
                text-align: center;
            '>Welcome to the GU Technical Assistance Provider System</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

    role = st.selectbox(
        "Select your role",
        ["Requester", "Coordinator", "Assignee/Staff","Research Assistant"],
        index=None,
        placeholder="Select option..."
    )

    if role:
        st.session_state.role = role
        st.rerun()

# --- Show view based on role
else:
    st.sidebar.markdown(
        f"""
        <div style='
            background: #f8f9fa;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
            padding: 1.2em 1em 1em 1em;
            margin-bottom: 1.5em;
            text-align: center;
            font-family: Arial, "Segoe UI", sans-serif;
        '>
            <span style='
                font-size: 1.15em;
                font-weight: 700;
                color: #1a237e;
                letter-spacing: 0.5px;
            '>
                Role: {st.session_state.role}
            </span>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.sidebar.button("🔄 Switch Role", on_click=lambda: st.session_state.update({
        "authenticated": False,
        "role": None,
        "user_email": ""
    }))

    # Sidebar: refresh cached datasets
    if st.sidebar.button("🔁 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

    st.markdown("""
        <style>
        .stButton > button {
            width: 100%;
            background-color: #cdb4db;
            color: black;
            font-family: Arial, "Segoe UI", sans-serif;
            font-weight: 600;
            border-radius: 8px;
            padding: 0.6em;
            margin-top: 1em;
            transition: background 0.2s;
        }
        .stButton > button:hover {
            background-color: #b197fc;
            color: #222;
        }
        </style>
    """, unsafe_allow_html=True)

    # Requester: No login needed
    if st.session_state.role == "Requester":
        st.markdown(
            """
            <div style='
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                background: #f8f9fa;
                padding: 2em 0 1em 0;
                border-radius: 18px;
                box-shadow: 0 4px 24px rgba(0,0,0,0.07);
                margin-bottom: 2em;
            '>
                <img src='https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png' width='200' style='margin-bottom: 1em;'/>
                <h1 style='
                    color: #1a237e;
                    font-family: "Segoe UI", "Arial", sans-serif;
                    font-weight: 700;
                    margin: 0;
                    font-size: 2.2em;
                    text-align: center;
                '>📥 Georgetown University Technical Assistance Form</h1>
            </div>
            """,
            unsafe_allow_html=True
        )
        #st.header("📥 Georgetown University Technical Assistance Form ")
        st.write("Please complete this form to request Technical Assistance from Georgetown University's Technical Assistance Provider (GU-TAP) team. We will review your request and will be in touch within 1-2 business days. You will receive an email from a TA Coordinator to schedule a time to gather more details about your needs. Once we have this information, we will assign a TA Lead to support you.")
        # Requester form
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Name *",placeholder="Enter text", key="requester_name")
        with col2:
            title = st.text_input("Title/Position *",placeholder='Enter text', key="requester_title")
        col3, col4 = st.columns(2)
        with col3:
            organization = st.selectbox(
                "Organization *",
                lis_organization,
                index=None,
                placeholder="Select option...",
                key="requester_organization"
            )
        with col4:
            location = st.selectbox(
                "Location *",
                lis_location,
                index=None,
                placeholder="Select option...",
                key="requester_location"
            )
        col5, col6 = st.columns(2)
        with col5:
            email = st.text_input("Email Address *",placeholder="Enter email", key="requester_email")
        with col6:
            phone = st.text_input("Phone Number *",placeholder="(201) 555-0123", key="requester_phone")    
        col7, col8 = st.columns(2)
        with col7:
            focus_area_options = [
                "Housing", "Prevention", "Substance Abuse", "Rapid Start",
                "Telehealth/Telemedicine", "Data Sharing","Evaluation","Implementation Science",
                "Continuous Quality Improvement", "Sustainability", "Sub-recipient Monitoring",
                "Procurement", "Work Plan and Budget", "Other"
            ]

            focus_area = st.selectbox(
                "TA Focus Area *",
                focus_area_options,
                index=None,
                placeholder="Select option...",
                key="requester_focus_area"
            )

            # If "Other" is selected, show a text input for custom value
            if focus_area == "Other":
                focus_area_other = st.text_input("Please specify the TA Focus Area *", key="requester_focus_area_other")
                if focus_area_other:
                    focus_area = focus_area_other 
        with col8:
            type_TA = st.selectbox(
                "What Style of TA is needed *",
                ["In-Person","Virtual","Hybrid (Combination of in-person and virtual)","Unsure"],
                index=None,
                placeholder="Select option...",
                key="requester_type_ta"
            )
        col9, col10 = st.columns(2)
        with col9:
            due_date = st.date_input(
                "Target Due Date *",
                value=None,
                key="requester_due_date"
            )
            #if not due_date: 
                #st.error("Target Due Date is required.")

            # Add required check: due_date must be after today
            #if due_date and due_date <= datetime.today().date():
                #st.error("Target Due Date must be after today.")

        ta_description = st.text_area("TA Description *", placeholder='Enter text', height=150, key="requester_ta_description") 
        document = st.file_uploader(
            "Upload any files or attachments that are relevant to this request.",accept_multiple_files=True, key="requester_document"
        )
        priority_status = st.selectbox(
                "Priority Status *",
                ["Critical","High","Normal","Low"],
                index=None,
                placeholder="Select option...",
                key="requester_priority"
            )

        # Submit button
        st.markdown("""
            <style>
            .stButton > button {
                width: 100%;
                background-color: #cdb4db;
                color: black;
                font-family: Arial, "Segoe UI", sans-serif;
                font-weight: 600;
                border-radius: 8px;
                padding: 0.6em;
                margin-top: 1em;
            }
            </style>
        """, unsafe_allow_html=True)

        # Submit logic
        if st.button("Submit", key="requester_submit"):
            email_pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
            def clean_and_format_us_phone(phone_input):
                digits = re.sub(r'\D', '', phone_input)
                if len(digits) == 10:
                    return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
                else:
                    return None

            errors = []

            formatted_phone = clean_and_format_us_phone(phone)


            drive_links = ""
            # Required field checks
            if not name: errors.append("Name is required.")
            if not title: errors.append("Title/Position is required.")
            if not organization: errors.append("Organization must be selected.")
            if not location: errors.append("Location must be selected.")
            if not email or not re.match(email_pattern, email):
                errors.append("Please enter a valid email address.")
            if not phone or not formatted_phone:
                errors.append("Please enter a valid U.S. phone number (10 digits).")
            if not focus_area: errors.append("TA Focus Area must be selected.")
            if not type_TA: errors.append("TA Style must be selected.")
            if not due_date: errors.append("Target Due Date is required.")
            elif due_date <= datetime.today().date():
                errors.append("Target Due Date must be after today.")
            if not ta_description: errors.append("TA Description is required.")
            if not priority_status: errors.append("Priority Status must be selected.")

            # Show warnings or success
            if errors:
                for error in errors:
                    st.warning(error)
            else:
                # Only upload files if all validation passes
                if document:
                    try:
                        folder_id = "1fy1CZSs_t6E6IF68rxblY6YeBihrSPNT" 
                        links = []
                        upload_count = 0
                        for file in document:
                            # Rename file as: GU0001_filename.pdf
                            renamed_filename = f"{new_ticket_id}_{file.name}"
                            link = upload_file_to_drive(
                                file=file,
                                filename=renamed_filename,
                                folder_id=folder_id,
                                creds_dict=st.secrets["gcp_service_account"]
                            )
                            links.append(link)
                            upload_count += 1
                            st.success(f"✅ Successfully uploaded: {file.name}")
                        drive_links = ", ".join(links)
                        if upload_count > 0:
                            st.success(f"✅ All {upload_count} file(s) uploaded successfully to Google Drive!")    
                    except Exception as e:
                        st.error(f"❌ Error uploading file(s) to Google Drive: {str(e)}")

                new_row = {
                    'Ticket ID': new_ticket_id,
                    'Jurisdiction': location,
                    'Organization': organization,
                    'Name': name,
                    'Title/Position': title,
                    'Email Address': email,
                    "Phone Number": formatted_phone,
                    "Focus Area": focus_area,
                    "TA Type": type_TA,
                    "Targeted Due Date": due_date.strftime("%Y-%m-%d"),
                    "TA Description":ta_description,
                    "Priority": priority_status,
                    "Submit Date": datetime.today().strftime("%Y-%m-%d"),
                    "Status": "Submitted",
                    "Assigned Date": pd.NA,
                    "Close Date": pd.NA,
                    "Assigned Coach": pd.NA,
                    "Coordinator Comment": pd.NA,
                    "Staff Comment": pd.NA,
                    "Document": drive_links
                }
                new_data = pd.DataFrame([new_row])

                try:
                    # Append new data to Google Sheet
                    updated_sheet = pd.concat([df, new_data], ignore_index=True)
                    updated_sheet = updated_sheet.applymap(
                        lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                    )
                    # Replace NaN with empty strings to ensure JSON compatibility
                    updated_sheet = updated_sheet.fillna("")
                    spreadsheet1 = client.open('HRSA64_TA_Request')
                    worksheet1 = spreadsheet1.worksheet('Main')
                    worksheet1.update([updated_sheet.columns.values.tolist()] + updated_sheet.values.tolist())
                    
                    # Clear cache to refresh data
                    st.cache_data.clear()
                    
                    # Send email notifications to all coordinators
                    coordinator_emails = [coord_email for coord_email, user in USERS.items() if "Coordinator" in user]
                    #coordinator_emails = ["jw2104@georgetown.edu"]

                    subject = f"New TA Request Submitted: {new_ticket_id}"
                    for coord_email in coordinator_emails:
                        coordinator_name = USERS[coord_email]["Coordinator"]["name"]
                        personalized_body = f"""
                        Hi {coordinator_name},

                        A new Technical Assistance request has been submitted:

                        Ticket ID: {new_ticket_id}
                        Jurisdiction: {location}
                        Organization: {organization}
                        Name: {name}
                        Description: {ta_description}
                        Priority: {priority_status}
                        Attachments: {drive_links or 'None'}

                        Please review and assign this request via the GU-TAP System: https://hrsagutap.streamlit.app/.
                        Please contact gutap@georgetown.edu for any questions or concerns.

                        Best,
                        GU-TAP System
                        """
                        try:
                            send_email_mailjet(
                                to_email=coord_email,
                                subject=subject,
                                body=personalized_body,
                            )
                        except Exception as e:
                            st.warning(f"⚠️ Failed to send email to coordinator {coord_email}: {e}")


                    # Send confirmation email to requester
                    confirmation_subject = f"Your TA Request ({new_ticket_id}) has been received"
                    confirmation_body = f"""
                    Hi {name},

                    Thank you for submitting your Technical Assistance request to Georgetown University's Technical Assistance Provider (GU-TAP) team.

                    Here is a summary of your submission:
                    - Ticket ID: {new_ticket_id}
                    - Jurisdiction: {location}
                    - Organization: {organization}
                    - Name: {name}
                    - Title/Position: {title}
                    - Email Address: {email}
                    - Phone Number: {formatted_phone}
                    - Focus Area: {focus_area}
                    - TA Type: {type_TA}
                    - Targeted Due Date: {due_date.strftime("%Y-%m-%d")}
                    - Priority: {priority_status}
                    - Description: {ta_description}

                    A TA Coordinator will review your request and assign a coach to your request within 2-3 business days.
                    Please contact gutap@georgetown.edu for any questions or concerns.

                    Best,
                    GU-TAP System
                    """

                    try:
                        send_email_mailjet(
                            to_email=email,
                            subject=confirmation_subject,
                            body=confirmation_body,
                        )
                    except Exception as e:
                        st.warning(f"⚠️ Failed to send confirmation email to requester: {e}")


                    st.success("✅ Submission successful!")
                    
                    # Clear cache to refresh data
                    st.cache_data.clear()
                    
                    # Clear form fields by using st.rerun() which will reset all form widgets
                    st.session_state.clear()
                    
                    # Wait a moment then redirect to main page
                    time.sleep(3)
                    st.rerun()

                except Exception as e:
                    st.error(f"Error updating Google Sheets: {str(e)}")




    # --- Coordinator or Staff: Require login
    elif st.session_state.role in ["Coordinator", "Assignee/Staff", "Research Assistant"]:
        if not st.session_state.authenticated:
            st.subheader("🔐 Login Required")

            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            login = st.button("Login")

            if login:
                # Normalize email to lowercase for case-insensitive lookup
                email_normalized = email.strip().lower() if email else ""
                user_roles = USERS.get(email_normalized)
                if user_roles and st.session_state.role in user_roles:
                    user = user_roles[st.session_state.role]
                    if user["password"] == password:
                        st.session_state.authenticated = True
                        st.session_state.user_email = email_normalized
                        st.success("Login successful!")
                        st.rerun()
                    else:
                        st.error("Invalid credentials or role mismatch.")
                else:
                    st.error("Invalid credentials or role mismatch.")

        else:
            if st.session_state.role == "Coordinator":
                user_info = USERS.get(st.session_state.user_email)
                coordinator_name = user_info["Coordinator"]["name"]
                # Check if current coordinator is Mabintou (only sees Travel Authorization Review Center)
                is_mabintou_coordinator = st.session_state.user_email == "mo887@georgetown.edu"
                st.markdown(
                    """
                    <div style='
                        display: flex;
                        flex-direction: column;
                        align-items: center;
                        justify-content: center;
                        background: #f8f9fa;
                        padding: 2em 0 1em 0;
                        border-radius: 18px;
                        box-shadow: 0 4px 24px rgba(0,0,0,0.07);
                        margin-bottom: 2em;
                    '>
                        <img src='https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png' width='200' style='margin-bottom: 1em;'/>
                        <h1 style='
                            color: #1a237e;
                            font-family: "Segoe UI", "Arial", sans-serif;
                            font-weight: 700;
                            margin: 0;
                            font-size: 2.2em;
                            text-align: center;
                        '>📬 Coordinator Dashboard</h1>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                #st.header("📬 Coordinator Dashboard")
                # Personalized greeting
                if user_info and "Coordinator" in user_info:
                    st.markdown(f"""
                    <div style='                      
                    background: #f8f9fa;                        
                    border-radius: 12px;                        
                    box-shadow: 0 2px 8px rgba(0,0,0,0.04);                        
                    padding: 1.2em 1em 1em 1em;                        
                    margin-bottom: 1.5em;                        
                    text-align: center;                        
                    font-family: Arial, "Segoe UI", sans-serif;                    
                    '>
                        <span style='                           
                        font-size: 1.15em;
                        font-weight: 700;
                        color: #1a237e;
                        letter-spacing: 0.5px;'>
                            👋 Welcome, {coordinator_name}!
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Dashboard Overview Metrics - visible to all coordinators (no header, just show metrics directly)
                col1, col2, col3 = st.columns(3)
                total_request = df['Ticket ID'].nunique()
                inprogress_request = df[df['Status'] == 'In Progress']['Ticket ID'].nunique()
                completed_request = df[df['Status'] == 'Completed']['Ticket ID'].nunique()

                col1.metric(label="# of Total Requests", value= millify(total_request, precision=2))
                col2.metric(label="# of In-Progress Requests", value= millify(inprogress_request, precision=2))
                col3.metric(label="# of Completed Requests", value= millify(completed_request, precision=2))
                style_metric_cards(border_left_color="#DBF227")

                col1, col2, col3 = st.columns(3)
                # create column span
                today = datetime.today()
                last_week = today - timedelta(days=7)
                last_month = today - timedelta(days=30)
                undone_request = df[df['Status'] == 'Submitted']['Ticket ID'].nunique()
                pastweek_request = df[df['Submit Date'] >= last_week]['Ticket ID'].nunique()
                pastmonth_request = df[df['Submit Date'] >= last_month]['Ticket ID'].nunique()
                col1.metric(label="# of Unassigned Requests", value= millify(undone_request, precision=2))
                col2.metric(label="# of Requests from past week", value= millify(pastweek_request, precision=2))
                col3.metric(label="# of Requests from past month", value= millify(pastmonth_request, precision=2))
                style_metric_cards(border_left_color="#DBF227")
                
                # Hide all expanders for Mabintou except Travel Authorization Review Center
                if not is_mabintou_coordinator:
                    with st.expander("🔎 **MONITOR IN-PROGRESS REQUESTS**"):
                        st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🔎 In-Progress Requests Monitor
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Track all active Technical Assistance requests, view staff assignments, and monitor upcoming deadlines. Use interactive charts and filters to stay on top of your team's workload.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                        # Convert date columns
                        df["Assigned Date"] = pd.to_datetime(df["Assigned Date"], errors="coerce")
                        df["Targeted Due Date"] = pd.to_datetime(df["Targeted Due Date"], errors="coerce")
                        df["Submit Date"] = pd.to_datetime(df["Submit Date"], errors="coerce")

                        inprogress = df[df['Status'] == 'In Progress']
                        next_month = today + timedelta(days=30)
                        request_pastmonth = df[(df["Targeted Due Date"] <= next_month)&(df['Status'] == 'In Progress')]

                        col4, col5 = st.columns(2)
                        # --- Pie 1: In Progress
                        with col4:
                            st.markdown("##### 🟡 In Progress Requests by Coach")
                            if not inprogress.empty:
                                chart_data = inprogress['Assigned Coach'].value_counts().reset_index()
                                chart_data.columns = ['Assigned Coach', 'Count']
                                pie1 = alt.Chart(chart_data).mark_arc(innerRadius=50).encode(
                                    theta=alt.Theta(field="Count", type="quantitative"),
                                    color=alt.Color(field="Assigned Coach", type="nominal"),
                                    tooltip=["Assigned Coach", "Count"]
                                ).properties(width=250, height=250)
                                st.altair_chart(pie1, use_container_width=True)
                            else:
                                st.info("No in-progress requests to show.")

                        # --- Pie 2: Requests in Past Month
                        with col5:
                            st.markdown("##### 📅 Due in 30 Days by Coach")
                            if not request_pastmonth.empty:
                                chart_data = request_pastmonth['Assigned Coach'].value_counts().reset_index()
                                chart_data.columns = ['Assigned Coach', 'Count']
                                pie2 = alt.Chart(chart_data).mark_arc(innerRadius=50).encode(
                                    theta=alt.Theta(field="Count", type="quantitative"),
                                    color=alt.Color(field="Assigned Coach", type="nominal"),
                                    tooltip=["Assigned Coach", "Count"]
                                ).properties(width=250, height=250)
                                st.altair_chart(pie2, use_container_width=True)
                            else:
                                st.info("No requests with coming dues to show.")


                        staff_list = ["Jenevieve Opoku", "Deus Bazira", "Kemisha Denny", "Katherine Robsky", 
                        "Martine Etienne-Mesubi", "Seble Kassaye", "Weijun Yu", "Jiaqin Wu", "Zelalem Temesgen", "Carlos Rodriguez-Diaz",
                        "Kiah Moorehead","Vanessa Da Costa","Trena Mukherjee","Abby Jordan","Megan Highland","Jesus Hernandez Burgos",
                        "Samantha Cinnick","Bryan Shaw","Tara Hixson","Dzifa Awunyo-Akaba","Masill Miranda","Joy Berry","Ashley Clonchmore",
                        "Grace Hazlett","Lauren Wagner"]

                        staff_list_sorted = sorted(staff_list, key=lambda x: x.split()[0])

                        selected_staff = st.selectbox("Select a staff to view their requests", staff_list_sorted, index=None,
                                placeholder="Select option...")

                        today = datetime.today()
                        last_month = today - timedelta(days=30)
                        staff_dff = df[df["Assigned Coach"] == selected_staff].copy()
                        in_progress_count = staff_dff[staff_dff["Status"] == "In Progress"]['Ticket ID'].nunique()
                        due_soon_count = staff_dff[
                            (staff_dff["Status"] == "In Progress") & 
                            (staff_dff["Targeted Due Date"] <= next_month)
                        ]['Ticket ID'].nunique()
                        completed_recently = staff_dff[
                            (staff_dff["Status"] == "Completed") & 
                            (staff_dff["Submit Date"] >= last_month)
                        ]['Ticket ID'].nunique()

                        # Metric 
                        col1, col2, col3 = st.columns(3)
                        col1.metric("🟡 In Progress", in_progress_count)
                        col2.metric("📅 Due in 30 Days", due_soon_count)
                        col3.metric("✅ Completed (Last 30 Days)", completed_recently)

                        # Detailed Table
                        st.markdown("##### 📋 Detailed Request List")

                        # Status filter
                        status_options = ["In Progress", "Completed"]
                        selected_status = st.multiselect(
                            "Filter by request status",
                            options=status_options,
                            default=["In Progress"]
                        )

                        # Apply filters: staff and status
                        staff_dfff = staff_dff[staff_dff["Status"].isin(selected_status)].copy()

                        display_cols = [
                            "Ticket ID", "Jurisdiction", "Organization", "Name", "Focus Area", "TA Type",
                            "Targeted Due Date", "Priority", "Status", "TA Description","Document"
                        ]

                        # Sort by due date
                        staff_dfff = staff_dfff.sort_values(by="Targeted Due Date")

                        # Format dates for display
                        staff_dfff["Targeted Due Date"] = staff_dfff["Targeted Due Date"].dt.strftime("%Y-%m-%d")

                        st.dataframe(staff_dfff[display_cols].reset_index(drop=True))

                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                    with st.expander("📝 **ASSIGN TA REQUESTS**"):
                        st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                📝 TA Request Assignment Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Review all unassigned Technical Assistance requests and assign them to the appropriate staff member. Use the table and filters below to prioritize and manage new requests efficiently.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                        st.markdown("#### 📋 Unassigned Requests")

                        # Filter submitted requests
                        submitted_requests = df[df["Status"] == "Submitted"].copy()

                        if submitted_requests.empty:
                            st.info("No submitted requests at the moment.")
                        else:
                            # Define custom priority order
                            priority_order = {"Critical": 1, "High": 2, "Normal": 3, "Low": 4}

                            # Create a temporary column for sort priority
                            submitted_requests["PriorityOrder"] = submitted_requests["Priority"].map(priority_order)

                            # Convert date columns if needed
                            submitted_requests["Submit Date"] = pd.to_datetime(submitted_requests["Submit Date"], errors='coerce')
                            submitted_requests["Targeted Due Date"] = pd.to_datetime(submitted_requests["Targeted Due Date"], errors='coerce')

                            # Format dates to "YYYY-MM-DD" for display
                            submitted_requests["Submit Date"] = submitted_requests["Submit Date"].dt.strftime("%Y-%m-%d")
                            submitted_requests["Targeted Due Date"] = submitted_requests["Targeted Due Date"].dt.strftime("%Y-%m-%d")

                            # Sort by custom priority, then submit date, then due date
                            submitted_requests_sorted = submitted_requests.sort_values(
                                by=["PriorityOrder", "Submit Date", "Targeted Due Date"],
                                ascending=[True, True, True]
                            )

                            # Display clean table (exclude PriorityOrder column)
                            st.dataframe(submitted_requests_sorted[[
                                "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                                "Focus Area", "TA Type", "Submit Date", "Targeted Due Date", "Priority", "TA Description","Document"
                            ]].reset_index(drop=True))

                            # Select request by index 
                            request_indices = submitted_requests_sorted.index.tolist()
                            selected_request_index = st.selectbox(
                                "Select a request to assign",
                                options=request_indices,
                                format_func=lambda idx: f"{submitted_requests_sorted.at[idx, 'Ticket ID']} | {submitted_requests_sorted.at[idx, 'Name']} | {submitted_requests_sorted.at[idx, 'Jurisdiction']}",
                            )

                            # Select coach
                            selected_coach = st.selectbox(
                                "Assign a coach",
                                options=staff_list_sorted,
                                index=None,
                                placeholder="Select option..."
                            )

                            # Assign button
                            if st.button("✅ Assign Coach and Start TA"):
                                try:
                                    updated_df = df.copy()
                                    # Update the selected row
                                    updated_df.loc[selected_request_index, "Assigned Coach"] = selected_coach
                                    updated_df.loc[selected_request_index, "Assigned Coordinator"] = coordinator_name
                                    updated_df.loc[selected_request_index, "Status"] = "In Progress"
                                    updated_df.loc[selected_request_index, "Assigned Date"] = datetime.today().strftime("%Y-%m-%d")

                                    updated_df = updated_df.applymap(
                                        lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (pd.Timestamp, datetime)) and not pd.isna(x) else x
                                    )
                                    updated_df = updated_df.fillna("") 
                                    spreadsheet1 = client.open('HRSA64_TA_Request')
                                    worksheet1 = spreadsheet1.worksheet('Main')

                                    # Push to Google Sheet
                                    worksheet1.update([updated_df.columns.values.tolist()] + updated_df.values.tolist())

                                    # Clear cache to refresh data
                                    st.cache_data.clear()
                                    
                                    st.success(f"Coach {selected_coach} assigned! Status updated to 'In Progress'.")

                                    # Send email to staff   
                                    # Find staff email by name
                                    staff_email = None
                                    for email, roles in USERS.items():
                                        if "Assignee/Staff" in roles and roles["Assignee/Staff"]["name"] == selected_coach:
                                            staff_email = email
                                            break

                                    if staff_email:
                                        staff_subject = f"You have been assigned a new TA request: {updated_df.loc[selected_request_index, 'Ticket ID']}"
                                        staff_body = f"""
                                        Hi {selected_coach},

                                        You have been assigned as the coach for the following Technical Assistance request:

                                        Ticket ID: {updated_df.loc[selected_request_index, 'Ticket ID']}
                                        Jurisdiction: {updated_df.loc[selected_request_index, 'Jurisdiction']}
                                        Organization: {updated_df.loc[selected_request_index, 'Organization']}
                                        Name: {updated_df.loc[selected_request_index, 'Name']}
                                        Description: {updated_df.loc[selected_request_index, 'TA Description']}
                                        Priority: {updated_df.loc[selected_request_index, 'Priority']}
                                        Targeted Due Date: {updated_df.loc[selected_request_index, 'Targeted Due Date']}
                                        Attachments: {updated_df.loc[selected_request_index, 'Document'] or 'None'}

                                        Please view and manage this request via the GU-TAP System: https://hrsagutap.streamlit.app/.
                                        Please contact gutap@georgetown.edu for any questions or concerns.

                                        Best,
                                        GU-TAP System
                                        """
                                        try:
                                            send_email_mailjet(
                                                to_email=staff_email,
                                                subject=staff_subject,
                                                body=staff_body,
                                            )
                                        except Exception as e:
                                            st.warning(f"⚠️ Failed to send assignment email to staff {selected_coach}: {e}")

                                    time.sleep(2)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")

                            # --- Submit button styling (CSS injection)
                            st.markdown("""
                                <style>
                                .stButton > button {
                                    width: 100%;
                                    background-color: #cdb4db;
                                    color: black;
                                    font-weight: 600;
                                    border-radius: 8px;
                                    padding: 0.6em;
                                    margin-top: 1em;
                                }
                                </style>
                            """, unsafe_allow_html=True)

                    # --- Transfer TA Requests (Coordinator only)
                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                    with st.expander("🔄 **TRANSFER TA REQUESTS**"):
                        st.markdown("""
                            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                                <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                    🔄 TA Request Transfer Center
                                </div>
                                <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                    Reassign in-progress requests to different coaches when needed. Track transfer history and maintain clear communication throughout the process.
                                </div>
                            </div>
                        """, unsafe_allow_html=True)

                        inprogress_for_transfer = df[df['Status'] == 'In Progress'].copy()

                        if inprogress_for_transfer.empty:
                            st.info("No in-progress requests available to transfer.")
                        else:
                            # Selection of ticket to transfer
                            transfer_indices = inprogress_for_transfer.index.tolist()
                            selected_transfer_index = st.selectbox(
                                "Select a request to transfer",
                                options=transfer_indices,
                                format_func=lambda idx: f"{df.at[idx, 'Ticket ID']} | {df.at[idx, 'Jurisdiction']} (Current: {df.at[idx, 'Assigned Coach']})",
                            )

                            current_coach = df.at[selected_transfer_index, 'Assigned Coach']
                            # Exclude current coach from options
                            available_new_coaches = [c for c in staff_list_sorted if c != current_coach]
                            new_coach = st.selectbox(
                                "Transfer to coach",
                                options=available_new_coaches,
                                index=None,
                                placeholder="Select option..."
                            )

                            reason_transfer = st.text_area("Reason for transfer (optional)", placeholder="Enter text", height=100)

                            st.markdown("""
                                <style>
                                .stButton > button {
                                    width: 100%;
                                    background-color: #cdb4db;
                                    color: black;
                                    font-weight: 600;
                                    border-radius: 8px;
                                    padding: 0.6em;
                                    margin-top: 1em;
                                }
                                </style>
                            """, unsafe_allow_html=True)

                            if st.button("✅ Confirm Transfer"):
                                if not new_coach:
                                    st.warning("Please select a new coach to transfer to.")
                                else:
                                    try:
                                        updated_df = df.copy()
                                        old_coach = current_coach

                                        # Update transfer metadata
                                        updated_df.loc[selected_transfer_index, "Last Transfer From"] = old_coach or ""
                                        updated_df.loc[selected_transfer_index, "Last Transfer To"] = new_coach
                                        updated_df.loc[selected_transfer_index, "Last Transfer Date"] = datetime.today().strftime("%Y-%m-%d")
                                        updated_df.loc[selected_transfer_index, "Last Transfer By"] = coordinator_name
                                        updated_df.loc[selected_transfer_index, "Reason for Transfer"] = reason_transfer or ""

                                        # Update active assignment
                                        updated_df.loc[selected_transfer_index, "Assigned Coach"] = new_coach

                                        # Append transfer history
                                        history_entry = f"{datetime.today().strftime('%Y-%m-%d')} | By: {coordinator_name} | From: {old_coach or 'N/A'} -> To: {new_coach}" + (f" | Reason: {reason_transfer.strip()}" if reason_transfer else "")
                                        existing_history = str(updated_df.loc[selected_transfer_index, "Transfer History"]).strip()
                                        if existing_history and existing_history.lower() != "nan":
                                            updated_df.loc[selected_transfer_index, "Transfer History"] = existing_history + "\n" + history_entry
                                        else:
                                            updated_df.loc[selected_transfer_index, "Transfer History"] = history_entry

                                        # Stringify dates for sheet
                                        updated_df = updated_df.applymap(
                                            lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (pd.Timestamp, datetime)) and not pd.isna(x) else x
                                        )
                                        updated_df = updated_df.fillna("")

                                        spreadsheet1 = client.open('HRSA64_TA_Request')
                                        worksheet1 = spreadsheet1.worksheet('Main')
                                        worksheet1.update([updated_df.columns.values.tolist()] + updated_df.values.tolist())

                                        st.cache_data.clear()

                                        st.success(f"Request {updated_df.loc[selected_transfer_index, 'Ticket ID']} transferred from {old_coach or 'N/A'} to {new_coach}.")

                                        # Email notification to new assignee
                                        staff_email = None
                                        for _email, roles in USERS.items():
                                            if "Assignee/Staff" in roles and roles["Assignee/Staff"]["name"] == new_coach:
                                                staff_email = _email
                                                break

                                        if staff_email:
                                            staff_subject = f"You have been transferred a TA request: {updated_df.loc[selected_transfer_index, 'Ticket ID']}"
                                            staff_body = f"""
                                            Hi {new_coach},

                                            You have been assigned (via transfer) as the coach for the following Technical Assistance request:

                                            Ticket ID: {updated_df.loc[selected_transfer_index, 'Ticket ID']}
                                            Jurisdiction: {updated_df.loc[selected_transfer_index, 'Jurisdiction']}
                                            Organization: {updated_df.loc[selected_transfer_index, 'Organization']}
                                            Previous Coach: {old_coach or 'N/A'}
                                            Description: {updated_df.loc[selected_transfer_index, 'TA Description']}
                                            Priority: {updated_df.loc[selected_transfer_index, 'Priority']}
                                            Targeted Due Date: {updated_df.loc[selected_transfer_index, 'Targeted Due Date']}
                                            Attachments: {updated_df.loc[selected_transfer_index, 'Document'] or 'None'}

                                            Decision by: {coordinator_name}
                                            {('Reason: ' + reason_transfer) if reason_transfer else ''}

                                            Please view and manage this request via the GU-TAP System: https://hrsagutap.streamlit.app/.
                                            Please contact gutap@georgetown.edu for any questions or concerns.

                                            Best,
                                            GU-TAP System
                                            """
                                            try:
                                                send_email_mailjet(
                                                    to_email=staff_email,
                                                    subject=staff_subject,
                                                    body=staff_body,
                                                )
                                            except Exception as e:
                                                st.warning(f"⚠️ Failed to send transfer email to staff {new_coach}: {e}")

                                        # Email notification to previous coach (if exists)
                                        if old_coach and old_coach != new_coach:
                                            previous_coach_email = None
                                            for _email, roles in USERS.items():
                                                if "Assignee/Staff" in roles and roles["Assignee/Staff"]["name"] == old_coach:
                                                    previous_coach_email = _email
                                                    break

                                            if previous_coach_email:
                                                previous_subject = f"TA Request Transferred: {updated_df.loc[selected_transfer_index, 'Ticket ID']}"
                                                previous_body = f"""
                                                Hi {old_coach},

                                                The following Technical Assistance request has been transferred from you to another coach:

                                                Ticket ID: {updated_df.loc[selected_transfer_index, 'Ticket ID']}
                                                Jurisdiction: {updated_df.loc[selected_transfer_index, 'Jurisdiction']}
                                                Organization: {updated_df.loc[selected_transfer_index, 'Organization']}
                                                New Coach: {new_coach}
                                                Description: {updated_df.loc[selected_transfer_index, 'TA Description']}
                                                Priority: {updated_df.loc[selected_transfer_index, 'Priority']}
                                                Targeted Due Date: {updated_df.loc[selected_transfer_index, 'Targeted Due Date']}

                                                Decision by: {coordinator_name}
                                                {('Reason: ' + reason_transfer) if reason_transfer else ''}

                                                You no longer need to work on this request. Please contact gutap@georgetown.edu for any questions or concerns.

                                                Best,
                                                GU-TAP System
                                                """
                                                try:
                                                    send_email_mailjet(
                                                        to_email=previous_coach_email,
                                                        subject=previous_subject,
                                                        body=previous_body,
                                                    )
                                                except Exception as e:
                                                    st.warning(f"⚠️ Failed to send transfer notification to previous coach {old_coach}: {e}")

                                        time.sleep(2)
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error updating Google Sheets: {str(e)}")

                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                    with st.expander("👍 **DETAILS OF IN-PROGRESS & COMPLETED REQUESTS**"):
                        st.markdown("""
                            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                                <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                    👍 Request Management Center
                                </div>
                                <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                    Leave comments and updates for in-progress TA requests, and review the status and details of completed requests. Track progress and maintain clear communication.
                                </div>
                            </div>
                        """, unsafe_allow_html=True)

                        st.markdown("#### 🚧 In-progress Requests")


                        # Filter "In Progress" requests
                        in_progress_df = df[df["Status"] == "In Progress"].copy()

                        if in_progress_df.empty:
                            st.info("No requests currently in progress.")
                        else:
                            # Convert date columns
                            in_progress_df["Assigned Date"] = pd.to_datetime(in_progress_df["Assigned Date"], errors="coerce")
                            in_progress_df["Targeted Due Date"] = pd.to_datetime(in_progress_df["Targeted Due Date"], errors="coerce")
                            in_progress_df['Expected Duration (Days)'] = (in_progress_df["Targeted Due Date"]-in_progress_df["Assigned Date"]).dt.days

                            # Format dates
                            in_progress_df["Assigned Date"] = in_progress_df["Assigned Date"].dt.strftime("%Y-%m-%d")
                            in_progress_df["Targeted Due Date"] = in_progress_df["Targeted Due Date"].dt.strftime("%Y-%m-%d")


                            # --- Filters
                            st.markdown("##### 🔍 Filter Options")

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                priority_filter = st.multiselect(
                                    "Filter by Priority",
                                    options=in_progress_df["Priority"].unique(),
                                    default=in_progress_df["Priority"].unique(), key='in1'
                                )

                            with col2:
                                ta_type_filter = st.multiselect(
                                    "Filter by TA Type",
                                    options=in_progress_df["TA Type"].unique(),
                                    default=in_progress_df["TA Type"].unique(), key='in2'
                                )

                            with col3:
                                focus_area_filter = st.multiselect(
                                    "Filter by Focus Area",
                                    options=in_progress_df["Focus Area"].unique(),
                                    default=in_progress_df["Focus Area"].unique(), key='in3'
                                )


                            # Apply filters
                            filtered_df = in_progress_df[
                                (in_progress_df["Priority"].isin(priority_filter)) &
                                (in_progress_df["TA Type"].isin(ta_type_filter)) &
                                (in_progress_df["Focus Area"].isin(focus_area_filter))
                            ]

                            # Display filtered table
                            st.dataframe(filtered_df[[
                                "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                                "Focus Area", "TA Type", "Assigned Date", "Targeted Due Date","Expected Duration (Days)","Priority",
                                "Assigned Coach", "TA Description","Document","Coordinator Comment History"
                            ]].sort_values(by="Expected Duration (Days)").reset_index(drop=True))

                            # Select request by index (row number in submitted_requests)
                            request_indices1 = filtered_df.index.tolist()
                            selected_request_index1 = st.selectbox(
                                "Select a request to comment",
                                options=request_indices1,
                                format_func=lambda idx: f"{filtered_df.at[idx, 'Ticket ID']} | {filtered_df.at[idx, 'Name']} | {filtered_df.at[idx, 'Jurisdiction']}",
                            )

                            # Input + submit comment
                            comment_input = st.text_area("Comments", placeholder="Enter comments", height=150, key='comm')

                            if st.button("✅ Submit Comments"):
                                try:
                                    # Find the actual row index in the original df (map back using ID or index)
                                    selected_row_global_index = filtered_df.loc[selected_request_index1].name

                                    # Copy and update df
                                    updated_df = df.copy()
                                    # Keep latest comment in main field
                                    updated_df.loc[selected_row_global_index, "Coordinator Comment"] = comment_input
                                    # Append to history with timestamp and author
                                    ts = datetime.today().strftime("%Y-%m-%d %H:%M")
                                    author = coordinator_name
                                    entry = f"{ts} | {author}: {comment_input}" if comment_input else ""
                                    if entry:
                                        existing = str(updated_df.loc[selected_row_global_index, "Coordinator Comment History"]).strip()
                                        if existing and existing.lower() != "nan":
                                            updated_df.loc[selected_row_global_index, "Coordinator Comment History"] = existing + "\n" + entry
                                        else:
                                            updated_df.loc[selected_row_global_index, "Coordinator Comment History"] = entry
                                    updated_df = updated_df.applymap(
                                        lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (pd.Timestamp, datetime)) and not pd.isna(x) else x
                                    )
                                    updated_df = updated_df.fillna("") 


                                    # Push to Google Sheets
                                    spreadsheet1 = client.open('HRSA64_TA_Request')
                                    worksheet1 = spreadsheet1.worksheet('Main')
                                    worksheet1.update([updated_df.columns.values.tolist()] + updated_df.values.tolist())

                                    # Clear cache to refresh data
                                    st.cache_data.clear()
                                    
                                    st.success("💬 Comment saved successfully!.")
                                    time.sleep(2)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")

                        st.markdown("#### ✅ Completed Requests")


                        # Filter "Completed" requests
                        complete_df = df[df["Status"] == "Completed"].copy()

                        if complete_df.empty:
                            st.info("No requests currently completed.")
                        else:
                            # Convert date columns
                            complete_df["Assigned Date"] = pd.to_datetime(complete_df["Assigned Date"], errors="coerce")
                            complete_df["Close Date"] = pd.to_datetime(complete_df["Close Date"], errors="coerce")
                            complete_df["Targeted Due Date"] = pd.to_datetime(complete_df["Targeted Due Date"], errors="coerce")
                            complete_df['Expected Duration (Days)'] = (complete_df["Targeted Due Date"]-complete_df["Assigned Date"]).dt.days
                            complete_df['Actual Duration (Days)'] = (complete_df["Close Date"]-complete_df["Assigned Date"]).dt.days

                            # Format dates
                            complete_df["Assigned Date"] = complete_df["Assigned Date"].dt.strftime("%Y-%m-%d")
                            complete_df["Targeted Due Date"] = complete_df["Targeted Due Date"].dt.strftime("%Y-%m-%d")
                            complete_df["Close Date"] = complete_df["Close Date"].dt.strftime("%Y-%m-%d")

                            # --- Filters
                            st.markdown("##### 🔍 Filter Options")

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                priority_filter1 = st.multiselect(
                                    "Filter by Priority",
                                    options=complete_df["Priority"].unique(),
                                    default=complete_df["Priority"].unique(), key='com1'
                                )

                            with col2:
                                ta_type_filter1 = st.multiselect(
                                    "Filter by TA Type",
                                    options=complete_df["TA Type"].unique(),
                                    default=complete_df["TA Type"].unique(), key='com2'
                                )

                            with col3:
                                focus_area_filter1 = st.multiselect(
                                    "Filter by Focus Area",
                                    options=complete_df["Focus Area"].unique(),
                                    default=complete_df["Focus Area"].unique(), key='com3'
                                )

                            # Apply filters
                            filtered_df1 = complete_df[
                                (complete_df["Priority"].isin(priority_filter1)) &
                                (complete_df["TA Type"].isin(ta_type_filter1)) &
                                (complete_df["Focus Area"].isin(focus_area_filter1))
                            ]

                            # Display filtered table
                            st.dataframe(filtered_df1[[
                                "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                                "Focus Area", "TA Type", "Priority", "Assigned Coach", "TA Description","Document","Assigned Date",
                                "Targeted Due Date", "Close Date", "Expected Duration (Days)",
                                'Actual Duration (Days)', "Coordinator Comment History", "Staff Comment History", "Transfer History"
                            ]].reset_index(drop=True))

                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                    with st.expander("🗒️ **CHECK & SUBMIT INTERACTION LOG**"):
                        st.markdown("""
                            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                                <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                    🗒️ Interaction Management Center
                                </div>
                                <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                    Review your previous interactions and submit new ones. Track all your communications with jurisdictions and TA requests.
                                </div>
                            </div>
                        """, unsafe_allow_html=True)

                        # Upper section: Previous Interactions
                        st.markdown("""
                            <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em;'>
                                <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                    📊 Your Previous Interactions
                                </h3>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        # Get interaction data properly
                        df_int_coord = df_int[df_int["Submitted By"] == coordinator_name].copy()
                        if not df_int_coord.empty:
                            # Remove columns we don't want to display
                            display_cols = [col for col in df_int_coord.columns if col not in ['Submitted By', 'Submission Date']]
                            df_int_coord_display = df_int_coord[display_cols].copy()
                            
                            # Sort by Date of Interaction (most recent first)
                            df_int_coord_display["Date of Interaction"] = pd.to_datetime(df_int_coord_display["Date of Interaction"], errors="coerce")
                            df_int_coord_display = df_int_coord_display.sort_values("Date of Interaction", ascending=True)
                            df_int_coord_display["Date of Interaction"] = df_int_coord_display["Date of Interaction"].dt.strftime("%Y-%m-%d")
                            
                            # Add summary stats
                            total_interactions = len(df_int_coord_display)
                            recent_interactions = len(df_int_coord_display[df_int_coord_display["Date of Interaction"] >= (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")])
                            
                            st.markdown(f"""
                                <div style='background: #e3f2fd; border-radius: 10px; padding: 1em; margin-bottom: 1em; text-align: center;'>
                                    <div style='display: flex; justify-content: space-around;'>
                                        <div>
                                            <div style='font-size: 1.5em; font-weight: bold; color: #1976d2;'>{total_interactions}</div>
                                            <div style='font-size: 0.9em; color: #666;'>Total Interactions</div>
                                        </div>
                                        <div>
                                            <div style='font-size: 1.5em; font-weight: bold; color: #388e3c;'>{recent_interactions}</div>
                                            <div style='font-size: 0.9em; color: #666;'>Last 30 Days</div>
                                        </div>
                                    </div>
                                </div>
                            """, unsafe_allow_html=True)

                            st.dataframe(df_int_coord_display.reset_index(drop=True), use_container_width=True)
                            
                        else:
                            st.markdown("""
                                <div style='background: #fff3e0; border-radius: 15px; padding: 2em; text-align: center; border: 2px dashed #ff9800;'>
                                    <div style='font-size: 3em; margin-bottom: 0.5em;'>📝</div>
                                    <h4 style='color: #e65100; margin-bottom: 0.5em;'>No Previous Interactions</h4>
                                    <p style='color: #666; margin: 0;'>You haven't logged any interactions yet. Start by submitting your first interaction below!</p>
                                </div>
                            """, unsafe_allow_html=True)

                        # Middle section: View Interactions by Ticket ID
                        st.markdown("""
                            <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em; margin-top: 2em;'>
                                <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                    🔍 View Interactions by Ticket ID
                                </h3>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        # Get ticket IDs assigned to this coordinator
                        assigned_tickets = df[df["Assigned Coordinator"] == coordinator_name]["Ticket ID"].dropna().astype(str).unique().tolist()
                        assigned_tickets_sorted = sorted(assigned_tickets)
                        
                        if assigned_tickets_sorted:
                            selected_ticket_view = st.selectbox(
                                "Select a Ticket ID to view all interactions",
                                options=[""] + assigned_tickets_sorted,
                                index=0,
                                key='view_interactions_ticket_coord',
                                help="Select a ticket ID from your assigned requests to view all interactions for that ticket"
                            )
                            
                            if selected_ticket_view:
                                # Get all interactions for this ticket ID (regardless of who submitted)
                                # Handle NaN values properly
                                df_ticket_int = df_int[
                                    (df_int["Ticket ID"].notna()) & 
                                    (df_int["Ticket ID"].astype(str) == selected_ticket_view)
                                ].copy()
                                
                                if not df_ticket_int.empty:
                                    # Remove columns we don't want to display
                                    display_cols_ticket = [col for col in df_ticket_int.columns if col not in ['Submission Date']]
                                    df_ticket_int_display = df_ticket_int[display_cols_ticket].copy()
                                    
                                    # Sort by Date of Interaction (most recent first)
                                    df_ticket_int_display["Date of Interaction"] = pd.to_datetime(df_ticket_int_display["Date of Interaction"], errors="coerce")
                                    df_ticket_int_display = df_ticket_int_display.sort_values("Date of Interaction", ascending=True)
                                    df_ticket_int_display["Date of Interaction"] = df_ticket_int_display["Date of Interaction"].dt.strftime("%Y-%m-%d")
                                    
                                    st.markdown(f"**All interactions for Ticket ID: {selected_ticket_view}**")
                                    st.dataframe(df_ticket_int_display.reset_index(drop=True), use_container_width=True)
                                else:
                                    st.info(f"No interactions found for Ticket ID: {selected_ticket_view}")
                        else:
                            st.info("No assigned ticket IDs available to view interactions.")

                        # Lower section: Submit New Interaction
                        st.markdown("""
                            <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 1em;'>
                                <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                    ✍️ Submit New Interaction
                                </h3>
                            </div>
                        """, unsafe_allow_html=True)
                        lis_ticket = ["No Ticket ID"] + sorted([tid for tid in df["Ticket ID"].dropna().astype(str).unique().tolist()])

                        # Interaction Log form
                        col1, col2 = st.columns(2)
                        with col1:
                            ticket_id_int = st.selectbox("Ticket ID *", lis_ticket, index=None,
                                placeholder="Select option...", key='interaction_coord')
                        with col2:
                            date_int = st.date_input("Date of Interaction *", value=datetime.today().date())

                        # If No Ticket ID, ask for Jurisdiction
                        jurisdiction_for_no_ticket = None
                        if ticket_id_int == "No Ticket ID":
                            jurisdiction_for_no_ticket = st.selectbox(
                                "Jurisdiction *",
                                lis_location,
                                index=None,
                                placeholder="Select option...",
                                key='juris_interaction_coord'
                            )

                        list_interaction = [
                            "Email", "Phone Call", "In-Person Meeting", "Online Meeting", "Other"
                        ]

                        type_interaction = st.selectbox(
                            "Type of Interaction *",
                            list_interaction,
                            index=None,
                            placeholder="Select option..."
                        )

                        # If "Other" is selected, show a text input for custom value
                        if type_interaction == "Other":
                            type_interaction_other = st.text_input("Please specify the Type of Interaction *")
                            if type_interaction_other:
                                type_interaction = type_interaction_other 
                        
                        interaction_description = st.text_area("Short Summary *", placeholder='Enter text', height=150, key='interaction_description_coord') 

                        document_int = st.file_uploader(
                            "Upload any files or attachments that are relevant to this interaction.", 
                            accept_multiple_files=True
                        )

                        # Submit button
                        st.markdown("""
                            <style>
                            .stButton > button {
                                width: 100%;
                                background-color: #cdb4db;
                                color: black;
                                font-family: Arial, "Segoe UI", sans-serif;
                                font-weight: 600;
                                border-radius: 8px;
                                padding: 0.6em;
                                margin-top: 1em;
                            }
                            </style>
                        """, unsafe_allow_html=True)

                        # Submit logic
                        if st.button("Submit",key='interaction_submit_coord'):
                            errors = []
                            drive_links_int = ""  # Initialize here
                            # Required field checks
                            if not ticket_id_int: errors.append("Ticket ID is required.")
                            if ticket_id_int == "No Ticket ID" and not jurisdiction_for_no_ticket:
                                errors.append("Jurisdiction is required when Ticket ID is not provided.")
                            if not date_int: errors.append("Date of interaction is required.")
                            if not type_interaction: errors.append("Type of interaction is required.")
                            if not interaction_description: errors.append("Short summary is required.")

                            # Show warnings or success
                            if errors:
                                for error in errors:
                                    st.warning(error)
                            else:
                                # Only upload files if all validation passes
                                if document_int:
                                    try:
                                        folder_id_int = "19-Sm8W151tg1zyDN0Nh14DUvOVUieqq7" 
                                        links_int = []
                                        upload_count = 0
                                        for file in document_int:
                                            # Rename file as: GU0001_filename.pdf
                                            renamed_filename = f"{ticket_id_int}_{file.name}"
                                            link = upload_file_to_drive(
                                                file=file,
                                                filename=renamed_filename,
                                                folder_id=folder_id_int,
                                                creds_dict=st.secrets["gcp_service_account"]
                                            )
                                            links_int.append(link)
                                            upload_count += 1
                                            st.success(f"✅ Successfully uploaded: {file.name}")
                                        drive_links_int = ", ".join(links_int)
                                        if upload_count > 0:
                                            st.success(f"✅ All {upload_count} file(s) uploaded successfully to Google Drive!")    
                                    except Exception as e:
                                        st.error(f"❌ Error uploading file(s) to Google Drive: {str(e)}")

                                new_row_int = {
                                    'Ticket ID': ticket_id_int,
                                    "Date of Interaction": date_int.strftime("%Y-%m-%d"),
                                    "Type of Interaction": type_interaction,
                                    "Short Summary": interaction_description,
                                    "Document": drive_links_int,
                                    "Jurisdiction": (lambda: (
                                        # If a Ticket ID is provided, extract jurisdiction from Main sheet
                                        str(df.loc[df["Ticket ID"].astype(str) == str(ticket_id_int), "Jurisdiction"].iloc[0])
                                        if ticket_id_int != "No Ticket ID" and not df.loc[df["Ticket ID"].astype(str) == str(ticket_id_int), "Jurisdiction"].empty
                                        else (jurisdiction_for_no_ticket or "")
                                    ))(),
                                    "Submitted By": coordinator_name,
                                    "Submission Date": datetime.today().strftime("%Y-%m-%d %H:%M")
                                }
                                new_data_int = pd.DataFrame([new_row_int])

                                try:
                                    # Append new data to Google Sheet
                                    updated_sheet1 = pd.concat([df_int, new_data_int], ignore_index=True)
                                    updated_sheet1= updated_sheet1.applymap(
                                        lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                                    )
                                    # Replace NaN with empty strings to ensure JSON compatibility
                                    updated_sheet1 = updated_sheet1.fillna("")
                                    spreadsheet2 = client.open('HRSA64_TA_Request')
                                    worksheet2 = spreadsheet2.worksheet('Interaction')
                                    worksheet2.update([updated_sheet1.columns.values.tolist()] + updated_sheet1.values.tolist())

                                    # Clear cache to refresh data
                                    st.cache_data.clear()
                                    
                                    st.success("✅ Submission successful!")
                                    time.sleep(2)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")

                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                    with st.expander("📦 **SUBMIT DELIVERY FORM**"):
                        st.markdown("""
                            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                                <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                    📦 Delivery Management Center
                                </div>
                                <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                    Record new deliveries including reports, dashboards, and data. Upload files and provide comprehensive summaries of completed work.
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                        lis_ticket1 = df["Ticket ID"].unique().tolist()

                        # Interaction Log form
                        col1, col2 = st.columns(2)
                        with col1:
                            ticket_id_del = st.selectbox("Ticket ID *",lis_ticket1, index=None,
                                placeholder="Select option...",key='delivery')
                        with col2:
                            date_del = st.date_input("Date of Delivery *",value=datetime.today().date())

                        list_delivery = [
                            "Report", "Email Reply", "Dashboard", "New Data Points", "Peer Learning Facilitation", "TA Meeting", "Other"
                        ]

                        type_delivery = st.selectbox(
                            "Type of Delivery *",
                            list_delivery,
                            index=None,
                            placeholder="Select option..."
                        )

                        # If "Other" is selected, show a text input for custom value
                        if type_delivery == "Other":
                            type_delivery_other = st.text_input("Please specify the Type of Delivery *")
                            if type_delivery_other:
                                type_delivery = type_delivery_other 
                        delivery_description = st.text_area("Short Summary *", placeholder='Enter text', height=150,key='delivery_description') 
                        document_del = st.file_uploader(
                            "Upload any files or attachments that are relevant to this delivery.",accept_multiple_files=True
                        )

                        # Submit button
                        st.markdown("""
                            <style>
                            .stButton > button {
                                width: 100%;
                                background-color: #cdb4db;
                                color: black;
                                font-family: Arial, "Segoe UI", sans-serif;
                                font-weight: 600;
                                border-radius: 8px;
                                padding: 0.6em;
                                margin-top: 1em;
                            }
                            </style>
                        """, unsafe_allow_html=True)

                        # Submit logic
                        if st.button("Submit",key='delivery_submit'):
                            errors = []
                            drive_links_del = ""
                            # Required field checks
                            if not ticket_id_del: errors.append("Ticket ID is required.")
                            if not date_del: errors.append("Date of delivery is required.")
                            if not type_delivery: errors.append("Type of delivery is required.")
                            if not delivery_description: errors.append("Short summary is required.")

                            # Show warnings or success
                            if errors:
                                for error in errors:
                                    st.warning(error)
                            else:
                                # Only upload files if all validation passes
                                if document_del:
                                    try:
                                        folder_id_del = "1gXfWxys2cxd67YDk8zKPmG_mLGID4qL2" 
                                        links_del = []
                                        upload_count = 0
                                        for file in document_del:
                                            # Rename file as: GU0001_filename.pdf
                                            renamed_filename = f"{ticket_id_del}_{file.name}"
                                            link = upload_file_to_drive(
                                                file=file,
                                                filename=renamed_filename,
                                                folder_id=folder_id_del,
                                                creds_dict=st.secrets["gcp_service_account"]
                                            )
                                            links_del.append(link)
                                            upload_count += 1
                                            st.success(f"✅ Successfully uploaded: {file.name}")
                                        drive_links_del = ", ".join(links_del)
                                        if upload_count > 0:
                                            st.success(f"✅ All {upload_count} file(s) uploaded successfully to Google Drive!")    
                                    except Exception as e:
                                        st.error(f"❌ Error uploading file(s) to Google Drive: {str(e)}")

                                new_row_del = {
                                    'Ticket ID': ticket_id_del,
                                    "Date of Delivery": date_del.strftime("%Y-%m-%d"),  # Convert to string
                                    "Type of Delivery": type_delivery,
                                    "Short Summary": delivery_description,
                                    "Document": drive_links_del,
                                    "Submitted By": coordinator_name,
                                    "Submission Date": datetime.today().strftime("%Y-%m-%d %H:%M")
                                }
                                new_data_del = pd.DataFrame([new_row_del])

                                try:
                                    # Append new data to Google Sheet
                                    updated_sheet2 = pd.concat([df_del, new_data_del], ignore_index=True)
                                    updated_sheet2= updated_sheet2.applymap(
                                        lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                                    )
                                    # Replace NaN with empty strings to ensure JSON compatibility
                                    updated_sheet2 = updated_sheet2.fillna("")
                                    spreadsheet3 = client.open('HRSA64_TA_Request')
                                    worksheet3 = spreadsheet3.worksheet('Delivery')
                                    worksheet3.update([updated_sheet2.columns.values.tolist()] + updated_sheet2.values.tolist())

                                    # Clear cache to refresh data
                                    st.cache_data.clear()
                                    
                                    st.success("✅ Submission successful!")
                                    time.sleep(2)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")


                # Travel Authorization Review Center - visible to Jen, Kemisha, Lauren, Jiaqin, and Mabintou
                # (Close the if not is_mabintou_coordinator block here)
                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                with st.expander("✈️ **REVIEW & APPROVE TRAVEL AUTHORIZATION FORMS**"):
                    # Check access control
                    current_coordinator_email = st.session_state.user_email
                    is_kemisha = current_coordinator_email == "kd802@georgetown.edu"
                    is_mabintou = current_coordinator_email == "mo887@georgetown.edu"
                    is_jen = current_coordinator_email == "Jenevieve.Opoku@georgetown.edu"
                    is_lauren = current_coordinator_email == "lm1353@georgetown.edu"
                    is_jiaqin = current_coordinator_email == "jw2104@georgetown.edu"
                    
                    can_view_travel_review = is_kemisha or is_mabintou or is_jen or is_lauren or is_jiaqin
                    
                    if not can_view_travel_review:
                        st.info("This section is not available to you.")
                    else:
                        st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                ✈️ Travel Authorization Review Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Review and approve pending travel authorization forms. View PDFs, add signatures, and approve travel requests.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                    
                        # Load travel sheet data
                        try:
                            df_travel_review = load_travel_sheet()
                            
                            # Access control: Only Jen, Kemisha, Lauren, Jiaqin, and Mabintou can view this section
                            if not (is_kemisha or is_jen or is_lauren or is_jiaqin or is_mabintou):
                                st.info("This section is only available for Kemisha Denny, Jenevieve Opoku, Lauren Mathae, Jiaqin Wu, and Mabintou Ouattara.")
                            else:
                                # Determine status column based on coordinator
                                # Note: Approval routing is dynamic based on traveler:
                                # - Kemisha's requests → Mabintou + Jen
                                # - Mabintou's requests → Lauren + Kemisha
                                # - Others → Mabintou + Kemisha (or alternatives)
                                if is_kemisha:
                                    status_col = 'Kemisha Approval Status'
                                    approval_date_col = 'Kemisha Approval Date'
                                    signature_col = 'Kemisha Signature'
                                    note_col = 'Kemisha Note'
                                    coordinator_display_name = "Kemisha Denny"
                                elif is_jen:
                                    status_col = 'Jen Approval Status'
                                    approval_date_col = 'Jen Approval Date'
                                    signature_col = 'Jen Signature'
                                    note_col = 'Jen Note'
                                    coordinator_display_name = "Jenevieve Opoku"
                                elif is_lauren:
                                    status_col = 'Lauren Approval Status'
                                    approval_date_col = 'Lauren Approval Date'
                                    signature_col = 'Lauren Signature'
                                    note_col = 'Lauren Note'
                                    coordinator_display_name = "Lauren Mathae"
                                elif is_jiaqin:
                                    # Jiaqin can view but needs to determine which column based on form routing
                                    # For now, default to Kemisha column, but will be filtered by actual routing
                                    status_col = 'Kemisha Approval Status'
                                    approval_date_col = 'Kemisha Approval Date'
                                    signature_col = 'Kemisha Signature'
                                    note_col = 'Kemisha Note'
                                    coordinator_display_name = "Jiaqin Wu"
                                else:
                                    # Fallback (shouldn't reach here due to access control)
                                    status_col = 'Mabintou Approval Status'
                                    approval_date_col = 'Mabintou Approval Date'
                                    signature_col = 'Mabintou Signature'
                                    note_col = 'Mabintou Note'
                                    coordinator_display_name = "Mabintou Ouattara"
                            
                            # Filter for forms pending this coordinator's approval
                            # First, determine which forms are routed to this coordinator based on traveler
                            pending_forms = df_travel_review.copy()
                            
                            # Filter forms that are routed to this coordinator
                            def is_routed_to_coordinator(row):
                                """Check if form is routed to current coordinator"""
                                traveler_name_check = str(row.get('Name', '')).lower()
                                traveler_email_check = str(row.get('Email', '')).lower()
                                
                                is_kemisha_traveler = (traveler_email_check == 'kd802@georgetown.edu' or 
                                                      'kemisha' in traveler_name_check)
                                is_mabintou_traveler = (traveler_email_check == 'mo887@georgetown.edu' or 
                                                       'mabintou' in traveler_name_check)
                                
                                # Determine routing based on traveler
                                if is_kemisha_traveler:
                                    # Kemisha's requests → Mabintou + Jen
                                    if is_mabintou:
                                        return status_col == 'Mabintou Approval Status'
                                    elif is_jen:
                                        return status_col == 'Jen Approval Status'
                                    else:
                                        return False
                                elif is_mabintou_traveler:
                                    # Mabintou's requests → Lauren + Kemisha
                                    if is_lauren:
                                        return status_col == 'Lauren Approval Status'
                                    elif is_kemisha:
                                        return status_col == 'Kemisha Approval Status'
                                    else:
                                        return False
                                else:
                                    # Others → Mabintou + Kemisha (or alternatives)
                                    if is_mabintou:
                                        return status_col == 'Mabintou Approval Status'
                                    elif is_kemisha:
                                        return status_col == 'Kemisha Approval Status'
                                    elif is_jen:
                                        # Jen might be alternative for Kemisha
                                        return status_col == 'Jen Approval Status'
                                    elif is_lauren:
                                        # Lauren might be alternative for Mabintou
                                        return status_col == 'Lauren Approval Status'
                                    else:
                                        return False
                            
                            # Filter by routing
                            if len(pending_forms) > 0:
                                routed_mask = pending_forms.apply(is_routed_to_coordinator, axis=1)
                                pending_forms = pending_forms[routed_mask].copy()
                            
                            # Then filter by status
                            if status_col in pending_forms.columns and len(pending_forms) > 0:
                                pending_forms = pending_forms[
                                    (pending_forms[status_col].astype(str).str.lower() == 'pending') |
                                    (pending_forms[status_col].isna()) |
                                    (pending_forms[status_col].astype(str) == '') |
                                    (pending_forms[status_col].astype(str) == 'nan')
                                ].copy()
                                
                                # Filter out forms where the other approver has rejected
                                # Determine the other approver's status column based on routing
                                def get_other_approver_status_col(row):
                                    traveler_name_check = str(row.get('Name', '')).lower()
                                    traveler_email_check = str(row.get('Email', '')).lower()
                                    
                                    is_kemisha_traveler = (traveler_email_check == 'kd802@georgetown.edu' or 
                                                          'kemisha' in traveler_name_check)
                                    is_mabintou_traveler = (traveler_email_check == 'mo887@georgetown.edu' or 
                                                           'mabintou' in traveler_name_check)
                                    
                                    if is_kemisha_traveler:
                                        # Kemisha's requests → Mabintou + Jen
                                        if status_col == 'Mabintou Approval Status':
                                            return 'Jen Approval Status'
                                        elif status_col == 'Jen Approval Status':
                                            return 'Mabintou Approval Status'
                                    elif is_mabintou_traveler:
                                        # Mabintou's requests → Lauren + Kemisha
                                        if status_col == 'Lauren Approval Status':
                                            return 'Kemisha Approval Status'
                                        elif status_col == 'Kemisha Approval Status':
                                            return 'Lauren Approval Status'
                                    else:
                                        # Others → Mabintou + Kemisha (or alternatives)
                                        if status_col == 'Mabintou Approval Status':
                                            return 'Kemisha Approval Status'
                                        elif status_col == 'Kemisha Approval Status':
                                            return 'Mabintou Approval Status'
                                        elif status_col == 'Jen Approval Status':
                                            return 'Mabintou Approval Status'
                                        elif status_col == 'Lauren Approval Status':
                                            return 'Mabintou Approval Status'
                                    return None
                                
                                # Filter out forms where other approver has rejected
                                if len(pending_forms) > 0:
                                    def not_rejected_by_other(row):
                                        other_status_col = get_other_approver_status_col(row)
                                        if other_status_col and other_status_col in pending_forms.columns:
                                            other_status = str(row.get(other_status_col, '')).lower()
                                            return other_status != 'reject'
                                        return True
                                    
                                    pending_forms = pending_forms[pending_forms.apply(not_rejected_by_other, axis=1)].copy()
                                
                            elif status_col not in pending_forms.columns:
                                # If column doesn't exist, no forms pending
                                pending_forms = pd.DataFrame()
                            
                            if pending_forms.empty:
                                st.info("✅ No travel forms pending your approval at this time.")
                            else:
                                st.markdown(f"#### 📋 Forms Pending Your Approval ({coordinator_display_name})")
                                
                                # Display pending forms - show relevant approval status columns
                                display_cols = ['Name', 'Destination', 'Departure Date', 'Return Date', 'Purpose of Travel', 
                                              'Submission Date', 'PDF Link']
                                # Add approval status columns that exist
                                approval_status_cols = ['Kemisha Approval Status', 'Mabintou Approval Status', 
                                                       'Jen Approval Status', 'Lauren Approval Status']
                                for col in approval_status_cols:
                                    if col in pending_forms.columns:
                                        display_cols.append(col)
                                available_cols = [col for col in display_cols if col in pending_forms.columns]
                                
                                pending_display = pending_forms[available_cols].copy()
                                st.dataframe(pending_display.reset_index(drop=True), use_container_width=True)
                                
                                # Select form to review
                                form_indices = pending_forms.index.tolist()
                                if form_indices:
                                    selected_form_idx = st.selectbox(
                                        "Select a travel form to review and approve",
                                        options=form_indices,
                                        format_func=lambda idx: f"{pending_forms.at[idx, 'Name']} | {pending_forms.at[idx, 'Destination']} | {pending_forms.at[idx, 'Departure Date']}",
                                        key='travel_review_select'
                                    )
                                    
                                    selected_form = pending_forms.loc[selected_form_idx]
                                    
                                    st.markdown("---")
                                    st.markdown("#### 📄 Form Details")
                                    
                                    col_info1, col_info2 = st.columns(2)
                                    with col_info1:
                                        st.markdown(f"**Traveler:** {selected_form.get('Name', 'N/A')}")
                                        st.markdown(f"**Destination:** {selected_form.get('Destination', 'N/A')}")
                                        st.markdown(f"**Departure Date:** {selected_form.get('Departure Date', 'N/A')}")
                                        st.markdown(f"**Return Date:** {selected_form.get('Return Date', 'N/A')}")
                                    
                                    with col_info2:
                                        st.markdown(f"**Purpose:** {selected_form.get('Purpose of Travel', 'N/A')}")
                                        st.markdown(f"**Attendees:** {selected_form.get('Attendees', 'N/A')}")
                                        st.markdown(f"**Deliverables:** {selected_form.get('Deliverables', 'N/A')}")
                                        st.markdown(f"**Submitted:** {selected_form.get('Submission Date', 'N/A')}")
                                    
                                    # Show approval status
                                    kemisha_status = selected_form.get('Kemisha Approval Status', 'pending')
                                    mabintou_status = selected_form.get('Mabintou Approval Status', 'pending')
                                    # Handle NaN values
                                    if pd.isna(kemisha_status) or str(kemisha_status).lower() == 'nan':
                                        kemisha_status = 'pending'
                                    if pd.isna(mabintou_status) or str(mabintou_status).lower() == 'nan':
                                        mabintou_status = 'pending'
                                    st.markdown(f"**Kemisha Status:** {kemisha_status}")
                                    st.markdown(f"**Mabintou Status:** {mabintou_status}")
                                    
                                    # Display PDF link
                                    pdf_link = selected_form.get('PDF Link', '')
                                    if pdf_link:
                                        st.markdown(f"**PDF Link:** [View PDF]({pdf_link})")
                                    
                                    # Support files
                                    support_files = selected_form.get('Support Files', '')
                                    if support_files and str(support_files).strip():
                                        st.markdown(f"**Support Files:** {support_files}")
                                    
                                    st.markdown("---")
                                    st.markdown("#### ✍️ Decision Section")
                                    
                                    
                                    # Approval decision selection
                                    approval_decision = st.radio(
                                        "**Select Your Decision:**",
                                        ["✅ Approve", "❌ Reject"],
                                        key="travel_approval_decision",
                                        horizontal=True,
                                        help="Select your decision to approve or reject this travel authorization form"
                                    )
                                    
                                    approval_date = st.date_input(
                                        "**Approval Date:**",
                                        value=datetime.now().date(),
                                        key="travel_approval_date"
                                    )
                                    
                                    st.markdown("<br>", unsafe_allow_html=True)
                                    
                                    # Conditional display based on decision
                                    if approval_decision == "✅ Approve":
                                        # Show signature section for approval
                                        st.markdown("""
                                            <div style='background: #e8f5e9; border-left: 4px solid #4caf50; padding: 1em; border-radius: 5px; margin-bottom: 1em;'>
                                                <strong style='color: #2e7d32;'>Approval Selected</strong>
                                            </div>
                                        """, unsafe_allow_html=True)
                                        
                                        coordinator_signature_text = st.text_input(
                                            "**Type your full name to sign:**",
                                            key="travel_coordinator_signature",
                                            placeholder="Type your full name",
                                            help="Your typed name will be converted to a signature-style image"
                                        )
                                        
                                        # Show signature preview
                                        if coordinator_signature_text:
                                            try:
                                                preview_img = generate_signature_image(coordinator_signature_text, width=600, height=120, scale_factor=2)
                                                if preview_img:
                                                    if preview_img.mode != 'RGB':
                                                        rgb_preview = PILImage.new('RGB', preview_img.size, (255, 255, 255))
                                                        if preview_img.mode == 'RGBA':
                                                            rgb_preview.paste(preview_img, mask=preview_img.split()[3])
                                                        else:
                                                            rgb_preview.paste(preview_img)
                                                        preview_img = rgb_preview
                                                    preview_display = preview_img.resize((400, int(400 * preview_img.size[1] / preview_img.size[0])))
                                                    st.image(preview_display, caption="Signature Preview", width=400)
                                            except Exception as e:
                                                pass
                                        
                                        # Approve button
                                        approve_button_clicked = st.button("✅ Sign and Approve", key="travel_approve_button", type="primary", use_container_width=True)
                                    else:  # Reject selected
                                        # Show rejection reason section
                                        st.markdown("""
                                            <div style='background: #ffebee; border-left: 4px solid #f44336; padding: 1em; border-radius: 5px; margin-bottom: 1em;'>
                                                <strong style='color: #c62828;'>Rejection Selected</strong>
                                            </div>
                                        """, unsafe_allow_html=True)
                                        
                                        reject_note = st.text_area(
                                            "**Reason for rejection:** *",
                                            key="travel_reject_note",
                                            height=150,
                                            placeholder="Please provide a detailed reason for rejection (required)",
                                            help="This reason will be sent to the traveler via email"
                                        )
                                        
                                        coordinator_signature_text = ""  # Not needed for rejection
                                        approve_button_clicked = False
                                    
                                    # Handle approval/rejection based on decision
                                    if approval_decision == "✅ Approve" and approve_button_clicked:
                                        if not coordinator_signature_text or not coordinator_signature_text.strip():
                                            st.warning("⚠️ Please enter your signature (full name) to approve.")
                                        else:
                                                try:
                                                    # Update the travel form status
                                                    updated_df_travel = df_travel_review.copy()
                                                    updated_df_travel.loc[selected_form_idx, status_col] = 'approve'
                                                    updated_df_travel.loc[selected_form_idx, approval_date_col] = approval_date.strftime('%Y-%m-%d')
                                                    updated_df_travel.loc[selected_form_idx, signature_col] = coordinator_signature_text
                                                    
                                                    updated_df_travel = updated_df_travel.fillna("")
                                                    spreadsheet_travel = client.open('HRSA64_TA_Request')
                                                    try:
                                                        worksheet_travel = spreadsheet_travel.worksheet('Travel')
                                                    except:
                                                        worksheet_travel = spreadsheet_travel.add_worksheet(title='Travel', rows=1000, cols=20)
                                                    
                                                    worksheet_travel.update([updated_df_travel.columns.values.tolist()] + updated_df_travel.values.tolist())
                                                    
                                                    # Determine which status columns to check based on form routing
                                                    # Check traveler to determine routing
                                                    traveler_name_check = selected_form.get('Name', '').lower()
                                                    traveler_email_check = selected_form.get('Email', '').lower()
                                                    
                                                    is_kemisha_traveler_check = (traveler_email_check == 'kd802@georgetown.edu' or 
                                                                                'kemisha' in traveler_name_check)
                                                    is_mabintou_traveler_check = (traveler_email_check == 'mo887@georgetown.edu' or 
                                                                                 'mabintou' in traveler_name_check)
                                                    
                                                    # Determine the two approver status columns for this form
                                                    if is_kemisha_traveler_check:
                                                        approver1_status_col_check = 'Mabintou Approval Status'
                                                        approver2_status_col_check = 'Jen Approval Status'
                                                        approver1_sig_col = 'Mabintou Signature'
                                                        approver2_sig_col = 'Jen Signature'
                                                        approver1_date_col = 'Mabintou Approval Date'
                                                        approver2_date_col = 'Jen Approval Date'
                                                        approver1_name_final = "Mabintou Ouattara"
                                                        approver2_name_final = "Jenevieve Opoku"
                                                    elif is_mabintou_traveler_check:
                                                        approver1_status_col_check = 'Lauren Approval Status'
                                                        approver2_status_col_check = 'Kemisha Approval Status'
                                                        approver1_sig_col = 'Lauren Signature'
                                                        approver2_sig_col = 'Kemisha Signature'
                                                        approver1_date_col = 'Lauren Approval Date'
                                                        approver2_date_col = 'Kemisha Approval Date'
                                                        approver1_name_final = "Lauren Mathae"
                                                        approver2_name_final = "Kemisha Denny"
                                                    else:
                                                        # Default routing: Mabintou + Kemisha (or alternatives)
                                                        # Check for alternatives based on out of office status
                                                        out_of_office_check = {
                                                            'kemisha': False,  # Set to True if Kemisha is out
                                                            'mabintou': False  # Set to True if Mabintou is out
                                                        }
                                                        
                                                        if out_of_office_check.get('mabintou', False):
                                                            # Mabintou is out, use Lauren
                                                            approver1_status_col_check = 'Lauren Approval Status'
                                                            approver1_sig_col = 'Lauren Signature'
                                                            approver1_date_col = 'Lauren Approval Date'
                                                            approver1_name_final = "Lauren Mathae"
                                                        else:
                                                            approver1_status_col_check = 'Mabintou Approval Status'
                                                            approver1_sig_col = 'Mabintou Signature'
                                                            approver1_date_col = 'Mabintou Approval Date'
                                                            approver1_name_final = "Mabintou Ouattara"
                                                        
                                                        if out_of_office_check.get('kemisha', False):
                                                            # Kemisha is out, use Jen
                                                            approver2_status_col_check = 'Jen Approval Status'
                                                            approver2_sig_col = 'Jen Signature'
                                                            approver2_date_col = 'Jen Approval Date'
                                                            approver2_name_final = "Jenevieve Opoku"
                                                        else:
                                                            approver2_status_col_check = 'Kemisha Approval Status'
                                                            approver2_sig_col = 'Kemisha Signature'
                                                            approver2_date_col = 'Kemisha Approval Date'
                                                            approver2_name_final = "Kemisha Denny"
                                                    
                                                    # Ensure columns exist
                                                    for col in [approver1_status_col_check, approver2_status_col_check]:
                                                        if col not in updated_df_travel.columns:
                                                            updated_df_travel[col] = ''
                                                    
                                                    # Check if both have approved
                                                    approver1_status_new = updated_df_travel.loc[selected_form_idx, approver1_status_col_check] if approver1_status_col_check in updated_df_travel.columns else ''
                                                    approver2_status_new = updated_df_travel.loc[selected_form_idx, approver2_status_col_check] if approver2_status_col_check in updated_df_travel.columns else ''
                                                    
                                                    # Handle NaN values
                                                    if pd.isna(approver1_status_new) or str(approver1_status_new).lower() == 'nan':
                                                        approver1_status_new = ''
                                                    if pd.isna(approver2_status_new) or str(approver2_status_new).lower() == 'nan':
                                                        approver2_status_new = ''
                                                    
                                                    if str(approver1_status_new).lower() == 'approve' and str(approver2_status_new).lower() == 'approve':
                                                        # Both approved - generate final PDF with both signatures and send to traveler
                                                        try:
                                                            # Get both signatures using dynamic columns
                                                            approver1_sig = updated_df_travel.loc[selected_form_idx, approver1_sig_col] if approver1_sig_col in updated_df_travel.columns else ''
                                                            approver2_sig = updated_df_travel.loc[selected_form_idx, approver2_sig_col] if approver2_sig_col in updated_df_travel.columns else ''
                                                            
                                                            # Get both approval dates
                                                            approver1_date = updated_df_travel.loc[selected_form_idx, approver1_date_col] if approver1_date_col in updated_df_travel.columns else ''
                                                            approver2_date = updated_df_travel.loc[selected_form_idx, approver2_date_col] if approver2_date_col in updated_df_travel.columns else ''
                                                            
                                                            # Map to PDF format (Mabintou goes to Program Assistant, others to Lead)
                                                            # For Kemisha's requests: Mabintou (Program Assistant) + Jen (Lead)
                                                            # For Mabintou's requests: Lauren (Program Assistant) + Kemisha (Lead)
                                                            # For others: Mabintou (Program Assistant) + Kemisha (Lead)
                                                            if is_kemisha_traveler_check:
                                                                mabintou_sig = approver1_sig  # Mabintou is approver1
                                                                kemisha_sig = approver2_sig   # Jen is approver2, but goes to Lead position
                                                                mabintou_date = approver1_date
                                                                kemisha_date = approver2_date
                                                                # Note: Jen's signature goes to Lead position, Mabintou to Program Assistant
                                                                # We need to adjust this mapping
                                                                mabintou_sig_pdf = approver1_sig
                                                                kemisha_sig_pdf = approver2_sig  # Jen's signature
                                                            elif is_mabintou_traveler_check:
                                                                mabintou_sig = approver1_sig  # Lauren is approver1, goes to Program Assistant
                                                                kemisha_sig = approver2_sig   # Kemisha is approver2, goes to Lead
                                                                mabintou_date = approver1_date
                                                                kemisha_date = approver2_date
                                                                mabintou_sig_pdf = approver1_sig  # Lauren's signature
                                                                kemisha_sig_pdf = approver2_sig  # Kemisha's signature
                                                            else:
                                                                mabintou_sig = approver1_sig
                                                                kemisha_sig = approver2_sig
                                                                mabintou_date = approver1_date
                                                                kemisha_date = approver2_date
                                                                mabintou_sig_pdf = approver1_sig
                                                                kemisha_sig_pdf = approver2_sig
                                                            
                                                            # Helper function to safely parse JSON from sheet
                                                            def safe_json_loads(value, default=[], data_type='auto'):
                                                                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                                                                    return default
                                                                try:
                                                                    parsed = json.loads(str(value))
                                                                    if isinstance(parsed, list):
                                                                        if data_type == 'bool':
                                                                            # Convert to boolean: True for 1, True, "true", "True", False otherwise
                                                                            return [bool(x) if isinstance(x, bool) else (True if str(x).lower() in ['true', '1', 1] else False) for x in parsed]
                                                                        elif data_type == 'int':
                                                                            # Convert to integers
                                                                            return [int(float(x)) if isinstance(x, (int, float)) or (isinstance(x, str) and x.replace('.','').replace('-','').isdigit()) else (int(x) if isinstance(x, int) else 0) for x in parsed]
                                                                        elif data_type == 'float':
                                                                            # Convert to floats
                                                                            return [float(x) if isinstance(x, (int, float)) or (isinstance(x, str) and x.replace('.','').replace('-','').isdigit()) else 0.0 for x in parsed]
                                                                        else:
                                                                            # Auto-detect: preserve strings (especially dates), convert numbers
                                                                            result = []
                                                                            for x in parsed:
                                                                                if isinstance(x, bool):
                                                                                    result.append(x)
                                                                                elif isinstance(x, (int, float)):
                                                                                    result.append(float(x))
                                                                                elif isinstance(x, str):
                                                                                    # Preserve date strings (contain '/') and other non-numeric strings
                                                                                    if '/' in x or '-' in x or len(x) > 10:
                                                                                        # Likely a date or non-numeric string, preserve as-is
                                                                                        result.append(x)
                                                                                    elif x.replace('.','').replace('-','').isdigit():
                                                                                        # Numeric string, convert to float
                                                                                        result.append(float(x))
                                                                                    else:
                                                                                        # Other string, preserve as-is
                                                                                        result.append(x)
                                                                                else:
                                                                                    result.append(x)
                                                                            return result
                                                                    return default
                                                                except:
                                                                    return default
                                                            
                                                            def safe_get(value, default=''):
                                                                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                                                                    return default
                                                                return str(value)
                                                            
                                                            def safe_get_numeric(value, default=0):
                                                                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                                                                    return default
                                                                try:
                                                                    return float(value) if value else default
                                                                except:
                                                                    return default
                                                            
                                                            # Retrieve all fields from Google Sheet
                                                            form_data_for_pdf = {
                                                                # Basic traveler information
                                                                'name': safe_get(selected_form.get('Name', '')),
                                                                'email': safe_get(selected_form.get('Email', '')),
                                                                'destination': safe_get(selected_form.get('Destination', '')),
                                                                'departure_date': safe_get(selected_form.get('Departure Date', '')),
                                                                'return_date': safe_get(selected_form.get('Return Date', '')),
                                                                'purpose_of_travel': safe_get(selected_form.get('Purpose of Travel', '')),
                                                                'objective': safe_get(selected_form.get('Objective', '')),
                                                                'attendees': safe_get(selected_form.get('Attendees', '')),
                                                                'deliverables': safe_get(selected_form.get('Deliverables', '')),
                                                                'support_files': safe_get(selected_form.get('Support Files', '')),
                                                                # Traveler address information
                                                                'address1': safe_get(selected_form.get('Address1', '')),
                                                                'address2': safe_get(selected_form.get('Address2', '')),
                                                                'city': safe_get(selected_form.get('City', '')),
                                                                'state': safe_get(selected_form.get('State', '')),
                                                                'zip': safe_get(selected_form.get('Zip', '')),
                                                                'organization': safe_get(selected_form.get('Organization', 'Georgetown University')),
                                                                'signature': safe_get(selected_form.get('Signature', '')),
                                                                'signature_date': safe_get(selected_form.get('Signature Date', '')),
                                                                # Mileage data - dates as strings, amounts as floats
                                                                'mileage_dates': safe_json_loads(selected_form.get('Mileage Dates', '[]'), data_type='auto'),
                                                                'mileage_amounts': safe_json_loads(selected_form.get('Mileage Amounts', '[]'), data_type='float'),
                                                                'total_mileage': safe_get_numeric(selected_form.get('Total Mileage', 0)),
                                                                # Expense data - dates as strings, amounts as floats
                                                                'expense_dates': safe_json_loads(selected_form.get('Expense Dates', '[]'), data_type='auto'),
                                                                'airfare': safe_json_loads(selected_form.get('Airfare', '[]'), data_type='float'),
                                                                'ground_transport': safe_json_loads(selected_form.get('Ground Transport', '[]'), data_type='float'),
                                                                'parking': safe_json_loads(selected_form.get('Parking', '[]'), data_type='float'),
                                                                'lodging': safe_json_loads(selected_form.get('Lodging', '[]'), data_type='float'),
                                                                'baggage': safe_json_loads(selected_form.get('Baggage', '[]'), data_type='float'),
                                                                'misc': safe_json_loads(selected_form.get('Misc', '[]'), data_type='float'),
                                                                'misc2': safe_json_loads(selected_form.get('Misc2', '[]'), data_type='float'),
                                                                'misc_desc1': safe_get(selected_form.get('Misc Desc1', '')),
                                                                'misc_desc2': safe_get(selected_form.get('Misc Desc2', '')),
                                                                # Expense totals
                                                                'total_airfare': safe_get_numeric(selected_form.get('Total Airfare', 0)),
                                                                'total_ground_transport': safe_get_numeric(selected_form.get('Total Ground Transport', 0)),
                                                                'total_parking': safe_get_numeric(selected_form.get('Total Parking', 0)),
                                                                'total_lodging': safe_get_numeric(selected_form.get('Total Lodging', 0)),
                                                                'total_baggage': safe_get_numeric(selected_form.get('Total Baggage', 0)),
                                                                'total_misc': safe_get_numeric(selected_form.get('Total Misc', 0)),
                                                                # Per diem data - dates as strings, amounts as integers, checks as booleans
                                                                'per_diem_dates': safe_json_loads(selected_form.get('Per Diem Dates', '[]'), data_type='auto'),
                                                                'per_diem_amounts': safe_json_loads(selected_form.get('Per Diem Amounts', '[]'), data_type='int'),
                                                                'breakfast_checks': safe_json_loads(selected_form.get('Breakfast Checks', '[]'), data_type='bool'),
                                                                'lunch_checks': safe_json_loads(selected_form.get('Lunch Checks', '[]'), data_type='bool'),
                                                                'dinner_checks': safe_json_loads(selected_form.get('Dinner Checks', '[]'), data_type='bool'),
                                                                'total_per_diem': safe_get_numeric(selected_form.get('Total Per Diem', 0)),
                                                                'total_amount_due': safe_get_numeric(selected_form.get('Total Amount Due', 0)),
                                                                # Coordinator signatures and dates
                                                                # Note: mabintou_signature goes to Program Assistant row, kemisha_signature goes to Lead row
                                                                # For Kemisha's requests: Mabintou (Program Assistant) + Jen (Lead)
                                                                # For Mabintou's requests: Lauren (Program Assistant) + Kemisha (Lead)
                                                                # For others: Mabintou (Program Assistant) + Kemisha (Lead)
                                                                'mabintou_signature': mabintou_sig_pdf,  # Goes to Program Assistant row
                                                                'kemisha_signature': kemisha_sig_pdf,    # Goes to Lead Technical Assistance Provider row
                                                                'mabintou_approval_date': safe_get(mabintou_date),
                                                                'kemisha_approval_date': safe_get(kemisha_date),
                                                            }
                                                            
                                                            # Regenerate PDF with coordinator signatures
                                                            try:
                                                                wb, ws = load_excel_template()
                                                            except:
                                                                ws = None
                                                            
                                                            # Generate new PDF with signatures
                                                            final_pdf_buffer = create_pdf(form_data_for_pdf, ws)
                                                            final_pdf_filename = f"Travel_Authorization_Form_Approved_{selected_form.get('Name','')}_{selected_form.get('Departure Date','')}.pdf"
                                                            
                                                            # Upload new PDF to Google Drive
                                                            folder_id_travel_pdf = "1_O_L-jPR7bldiryRNB3WxbAaG8VqvmCt"
                                                            pdf_file_obj = io.BytesIO(final_pdf_buffer.getvalue())
                                                            pdf_file_obj.name = final_pdf_filename
                                                            pdf_file_obj.type = 'application/pdf'
                                                            
                                                            final_pdf_link = upload_file_to_drive(
                                                                file=pdf_file_obj,
                                                                filename=final_pdf_filename,
                                                                folder_id=folder_id_travel_pdf,
                                                                creds_dict=st.secrets["gcp_service_account"]
                                                            )
                                                            
                                                            # Update PDF link in sheet
                                                            updated_df_travel.loc[selected_form_idx, 'PDF Link'] = final_pdf_link
                                                            updated_df_travel = updated_df_travel.fillna("")
                                                            worksheet_travel.update([updated_df_travel.columns.values.tolist()] + updated_df_travel.values.tolist())
                                                            
                                                            traveler_email = selected_form.get('Email', '')
                                                            traveler_name = selected_form.get('Name', 'Unknown')
                                                            
                                                            if traveler_email and traveler_email.strip():
                                                                final_approval_subject = f"Travel Authorization Form Approved - {selected_form.get('Destination', '')}"
                                                                final_approval_body = f"""
Dear {traveler_name},

Your travel authorization form has been fully approved by both coordinators!

Travel Details:
- Destination: {selected_form.get('Destination', 'N/A')}
- Departure Date: {selected_form.get('Departure Date', 'N/A')}
- Return Date: {selected_form.get('Return Date', 'N/A')}

Approved by:
- {approver1_name_final}: {mabintou_date}
- {approver2_name_final}: {kemisha_date}

PDF Link: {final_pdf_link}

Your travel authorization form is now fully approved and ready for use.

Best regards,
GU-TAP System
                                                                """
                                                                
                                                                try:
                                                                    send_email_mailjet(
                                                                        to_email=traveler_email,
                                                                        subject=final_approval_subject,
                                                                        body=final_approval_body.strip()
                                                                    )
                                                                    st.success(f"✅ Travel form fully approved! Notification sent to {traveler_email}")
                                                                except Exception as e:
                                                                    st.warning(f"⚠️ Form approved but failed to send email: {str(e)}")
                                                            else:
                                                                st.success("✅ Travel form fully approved!")
                                                        except Exception as e:
                                                            st.warning(f"⚠️ Error processing final approval: {str(e)}")
                                                    else:
                                                        st.success(f"✅ Your approval has been recorded. Waiting for the other coordinator's approval.")
                                                    
                                                    st.cache_data.clear()
                                                    time.sleep(2)
                                                    st.rerun()
                                                    
                                                except Exception as e:
                                                    st.error(f"❌ Error approving travel form: {str(e)}")
                                    
                                    # Handle rejection
                                    if approval_decision == "❌ Reject":
                                        reject_button_clicked = st.button("❌ Reject Travel Form", key="travel_reject_button", type="primary", use_container_width=True)
                                        
                                        if reject_button_clicked:
                                            reject_note = st.session_state.get('travel_reject_note', '')
                                            if not reject_note or not reject_note.strip():
                                                st.warning("⚠️ Please provide a reason for rejection.")
                                            else:
                                                try:
                                                    updated_df_travel = df_travel_review.copy()
                                                    updated_df_travel.loc[selected_form_idx, status_col] = 'reject'
                                                    updated_df_travel.loc[selected_form_idx, approval_date_col] = approval_date.strftime('%Y-%m-%d')
                                                    updated_df_travel.loc[selected_form_idx, note_col] = reject_note
                                                    
                                                    # Determine the other approver's status column and clear it
                                                    # Check traveler to determine routing
                                                    traveler_name_check = selected_form.get('Name', '').lower()
                                                    traveler_email_check = selected_form.get('Email', '').lower()
                                                    
                                                    is_kemisha_traveler_check = (traveler_email_check == 'kd802@georgetown.edu' or 
                                                                                'kemisha' in traveler_name_check)
                                                    is_mabintou_traveler_check = (traveler_email_check == 'mo887@georgetown.edu' or 
                                                                                 'mabintou' in traveler_name_check)
                                                    
                                                    # Determine the other approver's status column based on routing
                                                    other_status_col = None
                                                    if is_kemisha_traveler_check:
                                                        # Kemisha's requests → Mabintou + Jen
                                                        if status_col == 'Mabintou Approval Status':
                                                            other_status_col = 'Jen Approval Status'
                                                        elif status_col == 'Jen Approval Status':
                                                            other_status_col = 'Mabintou Approval Status'
                                                    elif is_mabintou_traveler_check:
                                                        # Mabintou's requests → Lauren + Kemisha
                                                        if status_col == 'Lauren Approval Status':
                                                            other_status_col = 'Kemisha Approval Status'
                                                        elif status_col == 'Kemisha Approval Status':
                                                            other_status_col = 'Lauren Approval Status'
                                                    else:
                                                        # Others → Mabintou + Kemisha (or alternatives)
                                                        if status_col == 'Mabintou Approval Status':
                                                            other_status_col = 'Kemisha Approval Status'
                                                        elif status_col == 'Kemisha Approval Status':
                                                            other_status_col = 'Mabintou Approval Status'
                                                        elif status_col == 'Jen Approval Status':
                                                            # If Jen is alternative, clear Mabintou or Kemisha
                                                            other_status_col = 'Mabintou Approval Status'
                                                        elif status_col == 'Lauren Approval Status':
                                                            # If Lauren is alternative, clear Mabintou or Kemisha
                                                            other_status_col = 'Mabintou Approval Status'
                                                    
                                                    # Clear the other approver's status (set to blank)
                                                    if other_status_col and other_status_col in updated_df_travel.columns:
                                                        updated_df_travel.loc[selected_form_idx, other_status_col] = ''
                                                        # Also clear related columns
                                                        other_date_col = other_status_col.replace('Status', 'Date')
                                                        other_sig_col = other_status_col.replace('Approval Status', 'Signature')
                                                        if other_date_col in updated_df_travel.columns:
                                                            updated_df_travel.loc[selected_form_idx, other_date_col] = ''
                                                        if other_sig_col in updated_df_travel.columns:
                                                            updated_df_travel.loc[selected_form_idx, other_sig_col] = ''
                                                    
                                                    updated_df_travel = updated_df_travel.fillna("")
                                                    spreadsheet_travel = client.open('HRSA64_TA_Request')
                                                    try:
                                                        worksheet_travel = spreadsheet_travel.worksheet('Travel')
                                                    except:
                                                        worksheet_travel = spreadsheet_travel.add_worksheet(title='Travel', rows=1000, cols=20)
                                                    
                                                    worksheet_travel.update([updated_df_travel.columns.values.tolist()] + updated_df_travel.values.tolist())
                                                    
                                                    # Send rejection email to traveler immediately
                                                    traveler_email = selected_form.get('Email', '')
                                                    traveler_name = selected_form.get('Name', 'Unknown')
                                                    
                                                    if traveler_email and traveler_email.strip():
                                                        rejection_subject = f"Travel Authorization Form Rejected - {selected_form.get('Destination', '')}"
                                                        rejection_body = f"""
Dear {traveler_name},

Your travel authorization form has been rejected.

Travel Details:
- Destination: {selected_form.get('Destination', 'N/A')}
- Departure Date: {selected_form.get('Departure Date', 'N/A')}
- Return Date: {selected_form.get('Return Date', 'N/A')}

Rejected by: {coordinator_display_name}
Rejection Date: {datetime.now().strftime('%Y-%m-%d')}
Reason: {reject_note}

Please review the reason and resubmit if needed.

Best regards,
GU-TAP System
                                                        """
                                                        
                                                        try:
                                                            send_email_mailjet(
                                                                to_email=traveler_email,
                                                                subject=rejection_subject,
                                                                body=rejection_body.strip()
                                                            )
                                                            st.success(f"❌ Travel form rejected. Rejection notification sent to {traveler_email}")
                                                        except Exception as e:
                                                            st.warning(f"⚠️ Form rejected but failed to send email: {str(e)}")
                                                    else:
                                                        st.success("❌ Travel form rejected.")
                                                    
                                                    st.cache_data.clear()
                                                    time.sleep(2)
                                                    st.rerun()
                                                    
                                                except Exception as e:
                                                    st.error(f"❌ Error rejecting travel form: {str(e)}")
                            
                            # Show approved forms section
                            st.markdown("---")
                            st.markdown("#### ✅ Fully Approved Forms")
                            
                            # Check for fully approved forms based on dynamic routing
                            # A form is fully approved when both approvers (determined by routing) have approved
                            def is_fully_approved(row):
                                """Check if form is fully approved based on its routing"""
                                traveler_name_check = str(row.get('Name', '')).lower()
                                traveler_email_check = str(row.get('Email', '')).lower()
                                
                                is_kemisha_traveler = (traveler_email_check == 'kd802@georgetown.edu' or 
                                                      'kemisha' in traveler_name_check)
                                is_mabintou_traveler = (traveler_email_check == 'mo887@georgetown.edu' or 
                                                       'mabintou' in traveler_name_check)
                                
                                if is_kemisha_traveler:
                                    # Kemisha's requests → Mabintou + Jen must both approve
                                    mabintou_status = str(row.get('Mabintou Approval Status', '')).lower()
                                    jen_status = str(row.get('Jen Approval Status', '')).lower()
                                    return mabintou_status == 'approve' and jen_status == 'approve'
                                elif is_mabintou_traveler:
                                    # Mabintou's requests → Lauren + Kemisha must both approve
                                    lauren_status = str(row.get('Lauren Approval Status', '')).lower()
                                    kemisha_status = str(row.get('Kemisha Approval Status', '')).lower()
                                    return lauren_status == 'approve' and kemisha_status == 'approve'
                                else:
                                    # Others → Mabintou + Kemisha (or alternatives) must both approve
                                    mabintou_status = str(row.get('Mabintou Approval Status', '')).lower()
                                    kemisha_status = str(row.get('Kemisha Approval Status', '')).lower()
                                    jen_status = str(row.get('Jen Approval Status', '')).lower()
                                    lauren_status = str(row.get('Lauren Approval Status', '')).lower()
                                    
                                    # Check primary approvers
                                    if mabintou_status == 'approve' and kemisha_status == 'approve':
                                        return True
                                    # Check alternatives
                                    if lauren_status == 'approve' and kemisha_status == 'approve':
                                        return True
                                    if mabintou_status == 'approve' and jen_status == 'approve':
                                        return True
                                    return False
                            
                            if len(df_travel_review) > 0:
                                approved_mask = df_travel_review.apply(is_fully_approved, axis=1)
                                fully_approved = df_travel_review[approved_mask].copy()
                            else:
                                fully_approved = pd.DataFrame()
                            
                            if not fully_approved.empty:
                                approved_display_cols = ['Name', 'Destination', 'Departure Date', 'Return Date', 'PDF Link']
                                # Add all approval status and date columns that exist
                                approval_cols = ['Kemisha Approval Status', 'Mabintou Approval Status', 'Jen Approval Status', 
                                               'Lauren Approval Status', 'Kemisha Approval Date', 'Mabintou Approval Date',
                                               'Jen Approval Date', 'Lauren Approval Date']
                                for col in approval_cols:
                                    if col in fully_approved.columns:
                                        approved_display_cols.append(col)
                                available_approved_cols = [col for col in approved_display_cols if col in fully_approved.columns]
                                st.dataframe(fully_approved[available_approved_cols].reset_index(drop=True), use_container_width=True)
                            else:
                                st.info("No fully approved forms yet.")
                    
                        except Exception as e:
                            st.error(f"Error loading travel forms: {str(e)}")

                # Hide Check Interaction & Delivery Patterns for Mabintou
                if not is_mabintou_coordinator:
                    st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                    with st.expander("📦 **CHECK INTERACTION & DELIVERY PATTERNS**"):
                        st.markdown("""
                            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                                <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                    📦 Activity Analytics Center
                                </div>
                                <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                    Visualize and analyze communication and delivery patterns for all Technical Assistance requests. Use charts and filters to spot trends and monitor engagement.
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                        # Fetch data from Google Sheets     
                        try:
                            df = load_main_sheet()  # Use cached function
                        except Exception as e:
                            st.error(f"Error fetching data from Google Sheets: {str(e)}")

                        try:
                            df_int = load_interaction_sheet()  # Use cached function
                        except Exception as e:
                            st.error(f"Error fetching data from Google Sheets: {str(e)}")

                        try:
                            df_del = load_delivery_sheet()  # Use cached function
                        except Exception as e:
                            st.error(f"Error fetching data from Google Sheets: {str(e)}")
                        
                        num_interaction = df_int.shape[0]
                        num_delivery = df_del.shape[0]

                        col1, col2 = st.columns(2)
                        col1.metric("🟡 # of Interactions", num_interaction)
                        col2.metric("🟡 # of Deliveries", num_delivery)

                        # Group by Ticket ID and Type of Interaction, count occurrences
                        interaction_counts = df_int.groupby(['Ticket ID', 'Type of Interaction']).size().reset_index(name='Count')
                        delivery_counts = df_del.groupby(['Ticket ID', 'Type of Delivery']).size().reset_index(name='Count')


                        st.markdown("##### 🟡 Top 10 with most Interactions by Interaction Type")
                        if not interaction_counts.empty:
                            pie1 = alt.Chart(interaction_counts).mark_bar().encode(
                                y=alt.Y('Ticket ID:N', sort='-x', title='Ticket ID'),
                                x=alt.X('Count:Q', title='Number of Interactions'),
                                color=alt.Color('Type of Interaction:N', title='Interaction Type'),
                                tooltip=['Ticket ID', 'Type of Interaction', 'Count']
                            ).properties(
                                width=600,
                                height=400
                            )
                            st.altair_chart(pie1, use_container_width=True)
                        else:
                            st.info("No any interaction to show.")

                        st.markdown("##### 🟡 Top 10 with most Deliveries by Delivery Type")
                        if not delivery_counts.empty:
                            pie1 = alt.Chart(delivery_counts).mark_bar().encode(
                                y=alt.Y('Ticket ID:N', sort='-x', title='Ticket ID'),
                                x=alt.X('Count:Q', title='Number of Deliveries'),
                                color=alt.Color('Type of Delivery:N', title='Delivery Type'),
                                tooltip=['Ticket ID', 'Type of Delivery', 'Count']
                            ).properties(
                                width=600,
                                height=400
                            )
                            st.altair_chart(pie1, use_container_width=True)
                        else:
                            st.info("No any delivery to show.")


                        unique_id = sorted(set(df_int["Ticket ID"].unique().tolist() + df_del["Ticket ID"].unique().tolist()))
                        selected_ticket_id = st.selectbox("Select a Ticket ID", unique_id, placeholder="Select option...")
                        # Filter records
                        interactions_for_ticket = df_int[df_int["Ticket ID"] == selected_ticket_id]
                        deliveries_for_ticket = df_del[df_del["Ticket ID"] == selected_ticket_id]
                        num_interaction_ticket = interactions_for_ticket.shape[0]
                        num_delivery_ticket = deliveries_for_ticket.shape[0]
                        st.markdown(f"##### 🟡 Ticket ID: {selected_ticket_id}")
                        st.markdown(f"🟡 # of Interactions: {num_interaction_ticket if num_interaction_ticket > 0 else 0}")
                        if num_interaction_ticket > 0:
                            st.dataframe(interactions_for_ticket)
                        else:
                            st.info("No interaction records found for this Ticket ID.")
                        st.markdown(f"🟡 # of Deliveries: {num_delivery_ticket if num_delivery_ticket > 0 else 0}")
                        if num_delivery_ticket > 0:
                            st.dataframe(deliveries_for_ticket)
                        else:
                            st.info("No delivery records found for this Ticket ID.")
                  


                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

            elif st.session_state.role == "Assignee/Staff":
                # Add staff content here
                user_info = USERS.get(st.session_state.user_email)
                user_email = st.session_state.user_email
                staff_name = user_info["Assignee/Staff"]["name"] if user_info and "Assignee/Staff" in user_info else None
                st.markdown(
                    """
                    <div style='
                        display: flex;
                        flex-direction: column;
                        align-items: center;
                        justify-content: center;
                        background: #f8f9fa;
                        padding: 2em 0 1em 0;
                        border-radius: 18px;
                        box-shadow: 0 4px 24px rgba(0,0,0,0.07);
                        margin-bottom: 2em;
                    '>
                        <img src='https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png' width='200' style='margin-bottom: 1em;'/>
                        <h1 style='
                            color: #1a237e;
                            font-family: "Segoe UI", "Arial", sans-serif;
                            font-weight: 700;
                            margin: 0;
                            font-size: 2.2em;
                            text-align: center;
                        '>👷 Staff Dashboard</h1>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                #st.header("👷 Staff Dashboard")
                # Personalized greeting
                if staff_name:
                    st.markdown(f"""
                    <div style='                      
                    background: #f8f9fa;                        
                    border-radius: 12px;                        
                    box-shadow: 0 2px 8px rgba(0,0,0,0.04);                        
                    padding: 1.2em 1em 1em 1em;                        
                    margin-bottom: 1.5em;                        
                    text-align: center;                        
                    font-family: Arial, "Segoe UI", sans-serif;                    
                    '>
                        <span style='                           
                        font-size: 1.15em;
                        font-weight: 700;
                        color: #1a237e;
                        letter-spacing: 0.5px;'>
                            👋 Welcome, {staff_name}!
                        </span>
                    </div>
                    """, unsafe_allow_html=True)


                # Filter requests assigned to current staff and In Progress
                staff_df = df[(df["Assigned Coach"] == staff_name) & (df["Status"] == "In Progress")].copy()
                com_df = df[(df["Assigned Coach"] == staff_name) & (df["Status"] == "Completed")].copy()


                # Ensure date columns are datetime
                staff_df["Targeted Due Date"] = pd.to_datetime(staff_df["Targeted Due Date"], errors="coerce")
                staff_df["Assigned Date"] = pd.to_datetime(staff_df["Assigned Date"], errors="coerce")

                # --- Top Summary Cards
                col1, col2 = st.columns(2)
                col3, col4 = st.columns(2)
                # 1. Total In Progress
                total_in_progress = staff_df['Ticket ID'].nunique()
                total_complete = com_df['Ticket ID'].nunique()

                # 2. Newly Assigned: within last 3 days
                recent_cutoff = datetime.today() - timedelta(days=3)
                newly_assigned = staff_df[staff_df["Assigned Date"] >= recent_cutoff]['Ticket ID'].nunique()

                # 3. Due within 1 month
                due_soon_cutoff = datetime.today() + timedelta(days=30)
                due_soon = staff_df[staff_df["Targeted Due Date"] <= due_soon_cutoff]['Ticket ID'].nunique()

                col1.metric("🟡 In Progress", total_in_progress)
                col2.metric("✅ Completed", total_complete)
                col3.metric("🆕 Newly Assigned (Last 3 days)", newly_assigned)
                col4.metric("📅 Due Within 1 Month", due_soon)

                style_metric_cards(border_left_color="#DBF227")


                # --- Section 2: Filter, Sort, Comment
                with st.expander("🚧 **IN-PROGRESS REQUESTS**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🚧 In-Progress Requests Management
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Manage your assigned requests, add comments, and track progress. Filter and sort your active TA requests efficiently.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Filter "In Progress" requests
                    if staff_df.empty:
                        st.info("No requests currently in progress.")
                    else:
                        # Convert date columns
                        staff_df["Assigned Date"] = pd.to_datetime(staff_df["Assigned Date"], errors="coerce")
                        staff_df["Targeted Due Date"] = pd.to_datetime(staff_df["Targeted Due Date"], errors="coerce")
                        staff_df['Expected Duration (Days)'] = (staff_df["Targeted Due Date"]-staff_df["Assigned Date"]).dt.days

                        # Format dates
                        staff_df["Assigned Date"] = staff_df["Assigned Date"].dt.strftime("%Y-%m-%d")
                        staff_df["Targeted Due Date"] = staff_df["Targeted Due Date"].dt.strftime("%Y-%m-%d")


                        # --- Filters
                        st.markdown("##### 🔍 Filter Options")

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            priority_filter = st.multiselect(
                                "Filter by Priority",
                                options=staff_df["Priority"].unique(),
                                default=staff_df["Priority"].unique(), key='sta1'
                            )

                        with col2:
                            ta_type_filter = st.multiselect(
                                "Filter by TA Type",
                                options=staff_df["TA Type"].unique(),
                                default=staff_df["TA Type"].unique(), key='sta2'
                            )

                        with col3:
                            focus_area_filter = st.multiselect(
                                "Filter by Focus Area",
                                options=staff_df["Focus Area"].unique(),
                                default=staff_df["Focus Area"].unique(), key='sta3'
                            )

                        # Apply filters
                        filtered_df2 = staff_df[
                            (staff_df["Priority"].isin(priority_filter)) &
                            (staff_df["TA Type"].isin(ta_type_filter)) &
                            (staff_df["Focus Area"].isin(focus_area_filter))
                        ]

                        # Display filtered table
                        st.dataframe(filtered_df2[[
                            "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                            "Focus Area", "TA Type", "Assigned Date", "Targeted Due Date","Expected Duration (Days)","Priority", "Assigned Coach", "TA Description",
                            "Document","Coordinator Comment History", "Staff Comment History", "Transfer History"
                        ]].sort_values(by="Expected Duration (Days)").reset_index(drop=True))

                        # Select request by index (row number in submitted_requests)
                        request_indices2 = filtered_df2.index.tolist()
                        selected_request_index1 = st.selectbox(
                            "Select a request to comment",
                            options=request_indices2,
                            format_func=lambda idx: f"{filtered_df2.at[idx, 'Ticket ID']} | {filtered_df2.at[idx, 'Name']} | {filtered_df2.at[idx, 'Jurisdiction']}",
                        )

                        # Input comment
                        comment_text = st.text_area("Staff Comment", placeholder="Enter comments", height=150, key='commm')

                        # Submit
                        if st.button("✅ Submit Comments"):
                            try:
                                # Get the index of the selected row in the full df
                                global_index = filtered_df2.loc[selected_request_index1].name

                                # Copy df and update
                                updated_df = df.copy()
                                # Keep latest comment in main field
                                updated_df.loc[global_index, "Staff Comment"] = comment_text
                                # Append to history with timestamp and author
                                ts = datetime.today().strftime("%Y-%m-%d %H:%M")
                                author = staff_name or "Staff"
                                entry = f"{ts} | {author}: {comment_text}" if comment_text else ""
                                if entry:
                                    existing = str(updated_df.loc[global_index, "Staff Comment History"]).strip()
                                    if existing and existing.lower() != "nan":
                                        updated_df.loc[global_index, "Staff Comment History"] = existing + "\n" + entry
                                    else:
                                        updated_df.loc[global_index, "Staff Comment History"] = entry
                                updated_df = updated_df.applymap(
                                    lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (pd.Timestamp, datetime)) and not pd.isna(x) else x
                                )
                                updated_df = updated_df.fillna("") 

                                # Push to Google Sheets
                                spreadsheet1 = client.open('HRSA64_TA_Request')
                                worksheet1 = spreadsheet1.worksheet('Main')
                                worksheet1.update([updated_df.columns.values.tolist()] + updated_df.values.tolist())

                                st.cache_data.clear()

                                st.success("💬 Comment saved successfully!.")
                                time.sleep(2)
                                st.rerun()

                            except Exception as e:
                                st.error(f"Error saving comment: {str(e)}")
                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                with st.expander("🗒️ **CHECK & SUBMIT INTERACTION LOG**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🗒️ Interaction Management Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Review your previous interactions and submit new ones. Track all your communications with jurisdictions and TA requests.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Upper section: Previous Interactions
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                📊 Your Previous Interactions
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Get interaction data properly
                    df_int_staff = df_int[df_int["Submitted By"] == staff_name].copy()
                    if not df_int_staff.empty:
                        # Remove columns we don't want to display
                        display_cols = [col for col in df_int_staff.columns if col not in ['Submitted By', 'Submission Date']]
                        df_int_staff_display = df_int_staff[display_cols].copy()
                        
                        # Sort by Date of Interaction (most recent first)
                        df_int_staff_display["Date of Interaction"] = pd.to_datetime(df_int_staff_display["Date of Interaction"], errors="coerce")
                        df_int_staff_display = df_int_staff_display.sort_values("Date of Interaction", ascending=True)
                        df_int_staff_display["Date of Interaction"] = df_int_staff_display["Date of Interaction"].dt.strftime("%Y-%m-%d")
                        
                        # Add summary stats
                        total_interactions = len(df_int_staff_display)
                        recent_interactions = len(df_int_staff_display[df_int_staff_display["Date of Interaction"] >= (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")])
                        
                        st.markdown(f"""
                            <div style='background: #e3f2fd; border-radius: 10px; padding: 1em; margin-top: 1em; text-align: center;'>
                                <div style='display: flex; justify-content: space-around;'>
                                    <div>
                                        <div style='font-size: 1.5em; font-weight: bold; color: #1976d2;'>{total_interactions}</div>
                                        <div style='font-size: 0.9em; color: #666;'>Total Interactions</div>
                                    </div>
                                    <div>
                                        <div style='font-size: 1.5em; font-weight: bold; color: #388e3c;'>{recent_interactions}</div>
                                        <div style='font-size: 0.9em; color: #666;'>Last 30 Days</div>
                                    </div>
                                </div>
                            </div>
                        """, unsafe_allow_html=True)

                        st.dataframe(df_int_staff_display.reset_index(drop=True), use_container_width=True)
                        

                    else:
                        st.markdown("""
                            <div style='background: #fff3e0; border-radius: 15px; padding: 2em; text-align: center; border: 2px dashed #ff9800;'>
                                <div style='font-size: 3em; margin-bottom: 0.5em;'>📝</div>
                                <h4 style='color: #e65100; margin-bottom: 0.5em;'>No Previous Interactions</h4>
                                <p style='color: #666; margin: 0;'>You haven't logged any interactions yet. Start by submitting your first interaction below!</p>
                            </div>
                        """, unsafe_allow_html=True)

                    # Middle section: View Interactions by Ticket ID
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em; margin-top: 2em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                🔍 View Interactions by Ticket ID
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Get ticket IDs assigned to this staff member
                    assigned_tickets_staff = df[df["Assigned Coach"] == staff_name]["Ticket ID"].dropna().astype(str).unique().tolist()
                    assigned_tickets_staff_sorted = sorted(assigned_tickets_staff)
                    
                    if assigned_tickets_staff_sorted:
                        selected_ticket_view_staff = st.selectbox(
                            "Select a Ticket ID to view all interactions",
                            options=[""] + assigned_tickets_staff_sorted,
                            index=0,
                            key='view_interactions_ticket_staff',
                            help="Select a ticket ID from your assigned requests to view all interactions for that ticket"
                        )
                        
                        if selected_ticket_view_staff:
                            # Get all interactions for this ticket ID (regardless of who submitted)
                            # Handle NaN values properly
                            df_ticket_int_staff = df_int[
                                (df_int["Ticket ID"].notna()) & 
                                (df_int["Ticket ID"].astype(str) == selected_ticket_view_staff)
                            ].copy()
                            
                            if not df_ticket_int_staff.empty:
                                # Remove columns we don't want to display
                                display_cols_ticket_staff = [col for col in df_ticket_int_staff.columns if col not in ['Submission Date']]
                                df_ticket_int_staff_display = df_ticket_int_staff[display_cols_ticket_staff].copy()
                                
                                # Sort by Date of Interaction (most recent first)
                                df_ticket_int_staff_display["Date of Interaction"] = pd.to_datetime(df_ticket_int_staff_display["Date of Interaction"], errors="coerce")
                                df_ticket_int_staff_display = df_ticket_int_staff_display.sort_values("Date of Interaction", ascending=True)
                                df_ticket_int_staff_display["Date of Interaction"] = df_ticket_int_staff_display["Date of Interaction"].dt.strftime("%Y-%m-%d")
                                
                                st.markdown(f"**All interactions for Ticket ID: {selected_ticket_view_staff}**")
                                st.dataframe(df_ticket_int_staff_display.reset_index(drop=True), use_container_width=True)
                            else:
                                st.info(f"No interactions found for Ticket ID: {selected_ticket_view_staff}")
                    else:
                        st.info("No assigned ticket IDs available to view interactions.")

                    # Lower section: Submit New Interaction
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 1em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                ✍️ Submit New Interaction
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    lis_ticket = ["No Ticket ID"] + sorted([tid for tid in df["Ticket ID"].dropna().astype(str).unique().tolist()])

                    # Interaction Log form
                    col1, col2 = st.columns(2)
                    with col1:
                        ticket_id_int = st.selectbox("Ticket ID *", lis_ticket, index=None,
                            placeholder="Select option...", key='interaction1')
                    with col2:
                        date_int = st.date_input("Date of Interaction *", value=datetime.today().date())
                    
                    # If No Ticket ID, ask for Jurisdiction
                    jurisdiction_for_no_ticket1 = None
                    if ticket_id_int == "No Ticket ID":
                        jurisdiction_for_no_ticket1 = st.selectbox(
                            "Jurisdiction *",
                            lis_location,
                            index=None,
                            placeholder="Select option...",
                            key='juris_interaction1'
                        )

                    list_interaction = [
                        "Email", "Phone Call", "In-Person Meeting", "Online Meeting", "Other"
                    ]

                    type_interaction = st.selectbox(
                        "Type of Interaction *",
                        list_interaction,
                        index=None,
                        placeholder="Select option..."
                    )

                    # If "Other" is selected, show a text input for custom value
                    if type_interaction == "Other":
                        type_interaction_other = st.text_input("Please specify the Type of Interaction *")
                        if type_interaction_other:
                            type_interaction = type_interaction_other 
                    
                    interaction_description = st.text_area("Short Summary *", placeholder='Enter text', height=150, key='interaction_description1') 

                    document_int = st.file_uploader(
                        "Upload any files or attachments that are relevant to this interaction.", 
                        accept_multiple_files=True
                    )

                    # Submit button
                    st.markdown("""
                        <style>
                        .stButton > button {
                            width: 100%;
                            background-color: #cdb4db;
                            color: black;
                            font-family: Arial, "Segoe UI", sans-serif;
                            font-weight: 600;
                            border-radius: 8px;
                            padding: 0.6em;
                            margin-top: 1em;
                        }
                        </style>
                    """, unsafe_allow_html=True)

                    try:
                        df = load_main_sheet()  # Use cached function
                    except Exception as e:
                        st.error(f"Error fetching data from Google Sheets: {str(e)}")

                    try:
                        df_int = load_interaction_sheet()  # Use cached function
                    except Exception as e:
                        st.error(f"Error fetching data from Google Sheets: {str(e)}")

                    try:
                        df_del = load_delivery_sheet()  # Use cached function
                    except Exception as e:
                        st.error(f"Error fetching data from Google Sheets: {str(e)}")

                    # Submit logic
                    if st.button("Submit",key='interaction_submit1'):
                        errors = []
                        drive_links_int = ""  # Initialize here
                        # Required field checks
                        if not ticket_id_int: errors.append("Ticket ID is required.")
                        if ticket_id_int == "No Ticket ID" and not jurisdiction_for_no_ticket1:
                            errors.append("Jurisdiction is required when Ticket ID is not provided.")
                        if not date_int: errors.append("Date of interaction is required.")
                        if not type_interaction: errors.append("Type of interaction is required.")
                        if not interaction_description: errors.append("Short summary is required.")

                        # Show warnings or success
                        if errors:
                            for error in errors:
                                st.warning(error)
                        else:
                            # Only upload files if all validation passes
                            if document_int:
                                try:
                                    folder_id_int = "19-Sm8W151tg1zyDN0Nh14DUvOVUieqq7" 
                                    links_int = []
                                    upload_count = 0
                                    for file in document_int:
                                        # Rename file as: GU0001_filename.pdf
                                        renamed_filename = f"{ticket_id_int}_{file.name}"
                                        link = upload_file_to_drive(
                                            file=file,
                                            filename=renamed_filename,
                                            folder_id=folder_id_int,
                                            creds_dict=st.secrets["gcp_service_account"]
                                        )
                                        links_int.append(link)
                                        upload_count += 1
                                        st.success(f"✅ Successfully uploaded: {file.name}")
                                    drive_links_int = ", ".join(links_int)
                                    if upload_count > 0:
                                        st.success(f"✅ All {upload_count} file(s) uploaded successfully to Google Drive!")    
                                except Exception as e:
                                    st.error(f"❌ Error uploading file(s) to Google Drive: {str(e)}")

                            new_row_int = {
                                'Ticket ID': ticket_id_int,
                                "Date of Interaction": date_int.strftime("%Y-%m-%d"),  # Convert to string
                                "Type of Interaction": type_interaction,
                                "Short Summary": interaction_description,
                                "Document": drive_links_int,
                                "Jurisdiction": (lambda: (
                                    str(df.loc[df["Ticket ID"].astype(str) == str(ticket_id_int), "Jurisdiction"].iloc[0])
                                    if ticket_id_int != "No Ticket ID" and not df.loc[df["Ticket ID"].astype(str) == str(ticket_id_int), "Jurisdiction"].empty
                                    else (jurisdiction_for_no_ticket1 or "")
                                ))(),
                                "Submitted By": staff_name,
                                "Submission Date": datetime.today().strftime("%Y-%m-%d %H:%M")
                            }
                            new_data_int = pd.DataFrame([new_row_int])

                            try:
                                # Append new data to Google Sheet
                                updated_sheet2 = pd.concat([df_int, new_data_int], ignore_index=True)
                                updated_sheet2= updated_sheet2.applymap(
                                    lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                                )
                                # Replace NaN with empty strings to ensure JSON compatibility
                                updated_sheet2 = updated_sheet2.fillna("")
                                
                                # Get the worksheet first
                                spreadsheet3 = client.open('HRSA64_TA_Request')
                                worksheet3 = spreadsheet3.worksheet('Interaction')
                                worksheet3.update([updated_sheet2.columns.values.tolist()] + updated_sheet2.values.tolist())

                                # Clear cache to refresh data
                                st.cache_data.clear()
                                
                                st.success("✅ Submission successful!")
                                time.sleep(2)
                                st.rerun()

                            except Exception as e:
                                st.error(f"Error updating Google Sheets: {str(e)}")

                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                with st.expander("👨‍💻 **SUBMIT STUDENT SUPPORT REQUEST FORM**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                👨‍💻 Student Support Request Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Submit new student support requests with time preferences. The system will automatically notify available research assistants.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Start with Anticipated Delivery
                    anticipated_delivery = st.selectbox("Anticipated Delivery *", options=["Meeting notes", "Dashboard", "Peer learning facilitation", "TA meeting", "Other"], index=0) 
                    if anticipated_delivery == "Other":
                        anticipated_delivery_other = st.text_input("Please specify the Anticipated Delivery *")
                        if anticipated_delivery_other:
                            anticipated_delivery = anticipated_delivery_other

                    # Conditional form fields based on delivery type
                    if anticipated_delivery == "Meeting notes":
                        st.markdown("**📅 Meeting Details**")
                        date_support = st.date_input("Date of Meeting *", value=datetime.today().date())
                        col1, col2 = st.columns(2)
                        
                        # Create time options every 15 minutes from 8 AM to 6 PM
                        time_options = []
                        for hour in range(8, 18):  # 8 AM to 5 PM
                            for minute in [0, 15, 30, 45]:
                                time_str = f"{hour:02d}:{minute:02d}"
                                time_options.append(time_str)
                        
                        # Default to current time if within range, otherwise 9 AM
                        current_time_str = datetime.now().strftime("%H:%M")
                        if current_time_str in time_options:
                            default_start_idx = time_options.index(current_time_str)
                        else:
                            default_start_idx = time_options.index("09:00")

                        with col1:
                            start_time_idx = st.selectbox("Start Time *", 
                                                        options=range(len(time_options)), 
                                                        index=default_start_idx,
                                                        format_func=lambda x: time_options[x])
                            start_time = time_options[start_time_idx]
                        
                        with col2:
                            # Calculate default end time (1 hour later)
                            start_hour, start_min = map(int, start_time.split(":"))
                            end_hour = start_hour + 1
                            end_min = start_min
                            
                            # Handle hour overflow and cap at 18:00
                            if end_hour >= 18:
                                end_time = "18:00"
                            else:
                                end_time = f"{end_hour:02d}:{end_min:02d}"
                            
                            # End time options (from start time to 18:00)
                            end_time_options = []
                            for i, time_str in enumerate(time_options):
                                if i > start_time_idx and time_str <= "18:00":
                                    end_time_options.append((i, time_str))
                            
                            if end_time_options:
                                default_end_idx = 0
                                if end_time in [t[1] for t in end_time_options]:
                                    default_end_idx = next(i for i, t in enumerate(end_time_options) if t[1] == end_time)
                                
                                end_time_idx = st.selectbox("End Time *",
                                                            options=range(len(end_time_options)),
                                                            index=default_end_idx,
                                                            format_func=lambda x: end_time_options[x][1])
                                end_time = end_time_options[end_time_idx][1]
                            else:
                                end_time = "18:00"

                        time_support = f"{start_time} - {end_time}"
                        request_description = st.text_area("Meeting Description *", placeholder='Describe the meeting topic, agenda, or specific requirements...', height=150, key='meeting_description')
                        
                    else:
                        st.markdown("**⏰ Project Timeline**")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            time_commitment = st.selectbox("Anticipated Time Commitment *", 
                                                         options=["1-2 hours", "3-4 hours", "5-8 hours", "1-2 days", "3-5 days", "1-2 weeks", "More than 2 weeks"], 
                                                         index=None, 
                                                         placeholder="Select option...")
                        
                        with col2:
                            anticipated_deadline = st.date_input("Anticipated Deadline *", value=datetime.today().date() + timedelta(days=7))
                        
                        request_description = st.text_area("Project Description *", placeholder='Describe the project requirements, deliverables, and any specific details...', height=150, key='project_description')
                        
                        # Set default values for non-meeting requests
                        date_support = None
                        time_support = None 

                    # Preferred RA Selection
                    st.markdown("**👤 Preferred Research Assistant Assignment (Optional)**")
                    ra_list = ["No preference"] + sorted([name for name in STUDENT_SCHEDULE.keys()])
                    preferred_ra = st.selectbox(
                        "Select a preferred Research Assistant to assign (or leave as 'No preference' to notify all available RAs)",
                        options=ra_list,
                        index=0,
                        key='preferred_ra_selection'
                    )

                    # Submit button
                    st.markdown("""
                        <style>
                        .stButton > button {
                            width: 100%;
                            background-color: #cdb4db;
                            color: black;
                            font-family: Arial, "Segoe UI", sans-serif;
                            font-weight: 600;
                            border-radius: 8px;
                            padding: 0.6em;
                            margin-top: 1em;
                        }
                        </style>
                    """, unsafe_allow_html=True)

                    # Submit logic
                    if st.button("Submit",key='support_submit1'):
                        errors = []
                        drive_links_del = ""  # Ensure always defined
                        
                        # Required field checks based on delivery type
                        if not anticipated_delivery: 
                            errors.append("Anticipated delivery is required.")
                        if not request_description: 
                            errors.append("Description is required.")
                        
                        if anticipated_delivery == "Meeting notes":
                            if not date_support: 
                                errors.append("Date of meeting is required.")
                            if not time_support: 
                                errors.append("Time of meeting is required.")
                        else:
                            if not time_commitment: 
                                errors.append("Time commitment is required.")
                            if not anticipated_deadline: 
                                errors.append("Anticipated deadline is required.")

                        # Show warnings or success
                        if errors:
                            for error in errors:
                                st.warning(error)
                        else:
                            # Prepare data for Google Sheets
                            # Check if preferred RA is selected
                            has_preferred_ra = preferred_ra and preferred_ra != "No preference"
                            preferred_ra_name = preferred_ra if has_preferred_ra else ""
                            preferred_ra_email = STUDENT_SCHEDULE[preferred_ra]["email"] if has_preferred_ra else ""
                            
                            new_row_support = {
                                "Date": date_support.strftime("%Y-%m-%d") if date_support else "",
                                "Time request needed": time_support if time_support else "",
                                "Request description": request_description,
                                "Anticipated Deliverable": anticipated_delivery,
                                "TAP Name": staff_name,
                                "TAP email": user_email,
                                "Time Commitment": time_commitment if anticipated_delivery != "Meeting notes" else "",
                                "Anticipated Deadline": anticipated_deadline.strftime("%Y-%m-%d") if anticipated_delivery != "Meeting notes" and anticipated_deadline else "",
                                "Request Type": "Meeting" if anticipated_delivery == "Meeting notes" else "Project",
                                "Student assigned": preferred_ra_name,
                                "Student email": preferred_ra_email,
                                "Request status": "Not Started" if has_preferred_ra else ""
                            }
                            new_data_support = pd.DataFrame([new_row_support])

                            try:
                                # Append new data to Google Sheet
                                updated_sheet3 = pd.concat([df_support, new_data_support], ignore_index=True)
                                updated_sheet3= updated_sheet3.applymap(
                                    lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                                )
                                # Replace NaN with empty strings to ensure JSON compatibility
                                updated_sheet3 = updated_sheet3.fillna("")
                                spreadsheet4 = client.open('HRSA64_TA_Request')
                                worksheet4 = spreadsheet4.worksheet('GA_Support')
                                worksheet4.update([updated_sheet3.columns.values.tolist()] + updated_sheet3.values.tolist())

                                # Clear cache to refresh data
                                st.cache_data.clear()
                                
                                st.success("✅ Submission successful!")
                                
                                # Handle notifications based on preferred RA selection
                                if has_preferred_ra:
                                    # Send direct assignment email to preferred RA
                                    st.markdown("---")
                                    st.markdown(f"**📧 Sending assignment notification to {preferred_ra}...**")
                                    
                                    # Format date for email
                                    if date_support:
                                        date_str_email = date_support.strftime("%Y-%m-%d")
                                    elif anticipated_deadline and anticipated_delivery != "Meeting notes":
                                        date_str_email = anticipated_deadline.strftime("%Y-%m-%d")
                                    else:
                                        date_str_email = ""
                                    
                                    ra_subject = f"You have been assigned a support request - {date_str_email if date_str_email else anticipated_delivery}"
                                    ra_body = f"""
Dear {preferred_ra},

You have been assigned to a support request by {staff_name}.

Request Details:
- Date: {date_str_email if date_str_email else 'N/A'}
- Time: {time_support if time_support else 'N/A'}
- TAP Name: {staff_name}
- TAP Email: {user_email}
- Request Description: {request_description}
- Anticipated Deliverable: {anticipated_delivery}
{f"- Time Commitment: {time_commitment}" if time_commitment else ""}
{f"- Anticipated Deadline: {anticipated_deadline.strftime('%Y-%m-%d')}" if anticipated_deadline and anticipated_delivery != "Meeting notes" else ""}

Status: Not Started

Please log into the GU-TAP System to view the request details and update the status as you progress.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                    """
                                    
                                    try:
                                        ra_notification_sent = send_email_mailjet(
                                            to_email=preferred_ra_email,
                                            subject=ra_subject,
                                            body=ra_body.strip()
                                        )
                                        if ra_notification_sent:
                                            st.success(f"✅ Assignment notification sent to {preferred_ra} ({preferred_ra_email})")
                                        else:
                                            st.warning(f"⚠️ Failed to send assignment notification to {preferred_ra}")
                                    except Exception as e:
                                        st.warning(f"⚠️ Failed to send assignment notification to {preferred_ra}: {e}")
                                    
                                    time.sleep(2)
                                    st.rerun()
                                    
                                else:
                                    # Default behavior: Send notifications based on request type
                                    st.markdown("---")
                                    st.markdown("**📧 Sending notifications to research assistants...**")
                                    
                                    notification_sent = False
                                    if anticipated_delivery == "Meeting notes":
                                        # Send to available students only
                                        notification_sent = send_support_request_notifications(
                                            date_str=date_support.strftime("%Y-%m-%d"),
                                            time_str=time_support,
                                            request_description=request_description,
                                            anticipated_delivery=anticipated_delivery,
                                            tap_name=staff_name,
                                            tap_email=user_email
                                        )
                                    else:
                                        # Send to all students for non-meeting requests
                                        notification_sent = send_project_request_notifications(
                                            request_description=request_description,
                                            anticipated_delivery=anticipated_delivery,
                                            time_commitment=time_commitment,
                                            anticipated_deadline=anticipated_deadline.strftime("%Y-%m-%d"),
                                            tap_name=staff_name,
                                            tap_email=user_email
                                        )
                                    
                                    # Wait a moment to show completion status
                                    time.sleep(1)
                                    
                                    # Show final status and rerun
                                    if notification_sent:
                                        st.success("✅ All notifications sent successfully!")
                                    else:
                                        st.warning("⚠️ Some notifications may have failed. Please check the logs above.")
                                    
                                    time.sleep(2)
                                    st.rerun()

                            except Exception as e:
                                st.error(f"Error updating Google Sheets: {str(e)}")

                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                with st.expander("🧳 **GENERATE DOMESTIC TRAVEL AUTHORIZATION FORM**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🧳 Generate Domestic Travel Authorization Form
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Input your travel information to generate a domestic travel authorization form.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("Fill out the form below to generate your Georgetown domestic travel authorization form.")
                    
                    try:
                        # Try to load Excel template - adjust path if needed
                        try:
                            wb, ws = load_excel_template()
                        except FileNotFoundError:
                            # Try alternative path
                            import os
                            alt_path = os.path.join(os.path.dirname(__file__), '..', 'Georgetown_Travel_Form_Generator', 'Georgetown Domestic Travel Authorization Form.xlsx')
                            if os.path.exists(alt_path):
                                wb = openpyxl.load_workbook(alt_path)
                                ws = wb['Reimbursement Form']
                            else:
                                st.warning("⚠️ Excel template file not found. PDF generation will proceed without template validation.")
                                wb, ws = None, None
                        
                        # General Guidance (UI only) - Using button toggle instead of nested expander
                        if 'show_travel_guidance' not in st.session_state:
                            st.session_state.show_travel_guidance = False
                        
                        if st.button("📋 General Guidance (Click to show/hide)", key="travel_guidance_toggle"):
                            st.session_state.show_travel_guidance = not st.session_state.show_travel_guidance
                        
                        if st.session_state.show_travel_guidance:
                            st.markdown("""
                            <div style='padding: 15px; background-color: #fafafa; border-radius: 5px; margin-top: 10px; border: 1px solid #e0e0e0;'>
                            """, unsafe_allow_html=True)
                            st.markdown("""
                                ### General Information
                                Fill out the fields highlighted in green, as applicable. Form must be submitted at least one month prior to your proposed dates of travel. Please inform ADVANCE leadership if extenuating circumstances will prevent you from meeting this deadline.

                                ### Receipts
                                You must submit receipts as part of your Expense Report in GMS for every item associated with your trip. This signed travel authorization will serve as your receipt for meals and incidentals. Meals are reimbursed at the Federal Per Diem rate for the destination city.

                                ### Mileage
                                In lieu of taxi expenses, you can choose to be reimbursed for the mileage driven from your point of origin to the airport, train station, or bus station. Georgetown University uses the IRS mileage rate.

                                Please attach documentation for the specified mileage in your GMS Expense report (e.g., Google Maps, MapQuest). Round all mileage to the nearest mile.

                                ### Airfare, Transportation, Parking, Lodging, Baggage Fees, Miscellaneous/Other
                                - **Airfare**: Should be booked through Concur and paid by Georgetown University. Include it as a cost in this Travel Authorization Form; your airfare should be included as an expense in your GMS Expense Report, but not as a personal reimbursement. If you are being reimbursed for your air travel, you must submit your itinerary and receipt.
                                - **Ground Transportation**: Covers reasonable expenses for taxis or other modes of transportation to and from airports and/or train and bus stations. Receipts must indicate the point of departure and point of arrival.
                                - **Parking**: If you are being reimbursed for parking, you must submit your receipt(s).
                                - **Lodging**: If lodging is purchased by the traveler, hotel receipts must be submitted. Lodging includes room and tax; it does not include telephone calls, room service, or other incidentals.
                                - **Baggage Fees**: Georgetown University will reimburse for one checked bag per passenger for each leg of trip (if the carrier charges for checked bags). For carriers with a free first bag, no reimbursement for additional bags will be allowed.
                                - **Miscellaneous/Other**: Includes pre‑approved travel expenses not listed in this form.

                                ### Meals and Incidental Expenses (M&IE)
                                Georgetown University will reimburse meals and incidentals at the U.S. Government per diem rates. This allowance covers tips, porter fees, etc.

                                Federal Guidelines stipulate that on the first and last travel day, travelers are only eligible for 75 percent of the total M&IE rate.

                                The cost of any meals provided at meetings and conferences will not be reimbursed by Georgetown University. For meals that have been provided, please place an "x" in the appropriate box on the reimbursement form.
                            """)
                            st.markdown("</div>", unsafe_allow_html=True)
                        
                        # Date inputs outside form so they trigger immediate reruns
                        st.header("Travel Dates")
                        col_date1, col_date2 = st.columns(2)
                        with col_date1:
                            departure_date = st.date_input("Departure Date *", key="travel_departure_date")
                        with col_date2:
                            return_date = st.date_input("Return Date *", key="travel_return_date")
                        
                        # Validate date range
                        if departure_date and return_date and return_date < departure_date:
                            st.error("Return Date must be the same as or after the Departure Date.")
                            st.stop()
                        
                        # Track date changes to auto-populate date fields
                        # Initialize session state for date tracking
                        if 'travel_last_departure' not in st.session_state:
                            st.session_state.travel_last_departure = departure_date
                        if 'travel_last_return' not in st.session_state:
                            st.session_state.travel_last_return = return_date
                        
                        # Check if dates changed
                        dates_changed = (departure_date != st.session_state.travel_last_departure or 
                                       return_date != st.session_state.travel_last_return)
                        
                        # Compute total days and generate full date range
                        if departure_date and return_date and return_date >= departure_date:
                            total_days = (return_date - departure_date).days + 1
                            # Reasonable upper bound to avoid runaway UI
                            total_days = min(total_days, 60)
                        else:
                            total_days = 7
                        default_dates = generate_date_range(departure_date, return_date, max_days=total_days)
                        
                        # Update session state when dates change (this happens on rerun)
                        if dates_changed:
                            st.session_state.travel_last_departure = departure_date
                            st.session_state.travel_last_return = return_date
                            # Update all date fields with new defaults when dates change
                            # Clear previous keys generously then set new defaults
                            for i in range(0, 100):
                                if i < len(default_dates) and default_dates[i]:
                                    st.session_state[f'travel_mileage_date_{i}'] = default_dates[i]
                                    st.session_state[f'travel_expense_date_{i}'] = default_dates[i]
                                    st.session_state[f'travel_per_diem_date_{i}'] = default_dates[i]
                                else:
                                    # Clear if beyond date range
                                    st.session_state[f'travel_mileage_date_{i}'] = ''
                                    st.session_state[f'travel_expense_date_{i}'] = ''
                                    st.session_state[f'travel_per_diem_date_{i}'] = ''
                        else:
                            # Initialize session state on first load if not exists
                            for i in range(total_days):
                                if f'travel_mileage_date_{i}' not in st.session_state:
                                    st.session_state[f'travel_mileage_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
                                if f'travel_expense_date_{i}' not in st.session_state:
                                    st.session_state[f'travel_expense_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
                                if f'travel_per_diem_date_{i}' not in st.session_state:
                                    st.session_state[f'travel_per_diem_date_{i}'] = default_dates[i] if i < len(default_dates) else ''
                        
                        with st.form("travel_form"):
                            st.header("Traveler Information")
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                name = st.text_input("Name *", value=staff_name, key="travel_name")
                                organization = st.text_input("Organization", value="Georgetown University", key="travel_organization")
                                destination = st.text_input("Destination *", key="travel_destination")
                                email = st.text_input("Email Address *", value=user_email, key="travel_email")              
                                
                            
                            with col2:
                                address1 = st.text_input("Address Line 1 *", key="travel_address1")
                                address2 = st.text_input("Address Line 2", key="travel_address2")
                                city = st.text_input("City *", key="travel_city")
                                state = st.text_input("State *", key="travel_state")
                                zip_code = st.text_input("Zip *", key="travel_zip")

                            st.header("Purpose of Travel")
                            col_purpose1, col_purpose2 = st.columns([1, 1])
                            with col_purpose1:
                                purpose_of_travel = st.text_area("Purpose of Travel *", key="travel_purpose_of_travel", height=100)
                                attendees = st.text_area("Attendees *", key="travel_attendees", height=100)
                            with col_purpose2:
                                objective = st.text_area("Objective *", key="travel_objective", height=100)
                                deliverables = st.text_area("Deliverables *", key="travel_deliverables", height=100)
                            support_files = st.file_uploader(
                                "Upload Documents (i.e Agenda, TA Request, etc.)",accept_multiple_files=True, key="travel_document"
                            )
                            
                            st.header("Mileage Expenses")
                            st.markdown("**The Mileage (Per Day) should be rounded to the nearest mile.**")
                            st.markdown("**Mileage rate for 2025: $0.70 per mile**")
                            
                            mileage_dates = []
                            mileage_amounts = []
                            
                            # Render mileage inputs in chunks of 7 days per row
                            for chunk_start in range(0, total_days, 7):
                                chunk_len = min(7, total_days - chunk_start)
                                cols = st.columns(chunk_len)
                                for offset in range(chunk_len):
                                    i = chunk_start + offset
                                    with cols[offset]:
                                        mileage_dates.append(st.text_input(f"Day {i+1}", key=f"travel_mileage_date_{i}", placeholder="MM/DD/YY"))
                                        mileage_amounts.append(number_text_input(f"Miles", key=f"travel_mileage_{i}", value=0.0, placeholder="0"))
                            
                            total_mileage = round(sum([m * 0.70 for m in mileage_amounts if m]),2)
                            
                            st.header("Travel Expenses")
                            expense_dates = []
                            airfare = []
                            ground_transport = []
                            parking = []
                            lodging = []
                            baggage = []
                            misc = []
                            misc2 = []
                            # First pass: render Date, Airfare, Ground, Parking, Lodging, Baggage
                            for chunk_start in range(0, total_days, 7):
                                chunk_len = min(7, total_days - chunk_start)
                                cols = st.columns(chunk_len)
                                for offset in range(chunk_len):
                                    i = chunk_start + offset
                                    with cols[offset]:
                                        expense_dates.append(st.text_input(f"Day {i+1}", key=f"travel_expense_date_{i}", placeholder="MM/DD/YY"))
                                        airfare.append(number_text_input(f"Airfare", key=f"travel_airfare_{i}", value=0.0, placeholder="0.00"))
                                        ground_transport.append(number_text_input(f"Ground Transportation", key=f"travel_ground_{i}", value=0.0, placeholder="0.00"))
                                        parking.append(number_text_input(f"Parking", key=f"travel_parking_{i}", value=0.0, placeholder="0.00"))
                                        lodging.append(number_text_input(f"Lodging", key=f"travel_lodging_{i}", value=0.0, placeholder="0.00"))
                                        baggage.append(number_text_input(f"Baggage Fees", key=f"travel_baggage_{i}", value=0.0, placeholder="0.00"))

                            # Descriptions next (always shown above misc rows, once for the section)
                            misc_desc1 = st.text_input("Miscellaneous/Other Description 1", key="travel_misc_desc1", placeholder="e.g., Registration")
                            # Second pass: render Misc Row 1 and Misc Row 2 amounts
                            for chunk_start in range(0, total_days, 7):
                                chunk_len = min(7, total_days - chunk_start)
                                cols = st.columns(chunk_len)
                                for offset in range(chunk_len):
                                    i = chunk_start + offset
                                    with cols[offset]:
                                        misc.append(number_text_input(f"{misc_desc1} Day {i+1}", key=f"travel_misc_{i}", value=0.0, placeholder="0.00"))

                            misc_desc2 = st.text_input("Miscellaneous/Other Description 2", key="travel_misc_desc2", placeholder="e.g., Supplies")

                            # Second pass: render Misc Row 1 and Misc Row 2 amounts
                            for chunk_start in range(0, total_days, 7):
                                chunk_len = min(7, total_days - chunk_start)
                                cols = st.columns(chunk_len)
                                for offset in range(chunk_len):
                                    i = chunk_start + offset
                                    with cols[offset]:
                                        misc2.append(number_text_input(f"{misc_desc2} Day {i+1}", key=f"travel_misc2_{i}", value=0.0, placeholder="0.00"))
                            
                            
                            st.header("Meals and Incidentals Per Diem")
                            st.markdown("**Please confirm the official GSA per diem rate for your travel destination at https://www.gsa.gov/travel/plan-book/per-diem-rates and select the corresponding rate below.**")
                            # Single per diem selection for all days
                            selected_per_diem = st.selectbox("Per Diem Rate (applies to all days)", options=[68,74,80,86,92], index=2, key="travel_per_diem_base")
                            per_diem_dates = []
                            per_diem_amounts = []
                            breakfast_checks = []
                            lunch_checks = []
                            dinner_checks = []
                            st.markdown("**Check boxes if meals were provided**")
                            # Render per diem inputs in chunks of 7 days per row
                            for chunk_start in range(0, total_days, 7):
                                chunk_len = min(7, total_days - chunk_start)
                                cols = st.columns(chunk_len)
                                for offset in range(chunk_len):
                                    i = chunk_start + offset
                                    with cols[offset]:
                                        per_diem_dates.append(st.text_input(f"Day {i+1}", key=f"travel_per_diem_date_{i}", placeholder="MM/DD/YY"))
                                        per_diem_amounts.append(selected_per_diem) 
                                        breakfast_checks.append(st.checkbox(f"Breakfast", key=f"travel_breakfast_{i}"))
                                        lunch_checks.append(st.checkbox(f"Lunch", key=f"travel_lunch_{i}"))
                                        dinner_checks.append(st.checkbox(f"Dinner", key=f"travel_dinner_{i}"))
                            
                            st.header("Additional Information")
                            
                            # E-Signature section
                            st.subheader("Traveler Signature")
                            col1, col2 = st.columns([2, 1])
                            with col1:
                                signature_text = st.text_input("Type your full name", key="travel_signature_text", 
                                                              help="Your typed name will be automatically converted to a signature-style image")
                                if signature_text:
                                    # Show preview of signature (use lower scale for preview to be faster)
                                    try:
                                        preview_img = generate_signature_image(signature_text, width=600, height=120, scale_factor=2)
                                        if preview_img:
                                            # Ensure it's RGB for display (should already be RGB now)
                                            if preview_img.mode != 'RGB':
                                                rgb_preview = PILImage.new('RGB', preview_img.size, (255, 255, 255))
                                                if preview_img.mode == 'RGBA':
                                                    rgb_preview.paste(preview_img, mask=preview_img.split()[3])
                                                else:
                                                    rgb_preview.paste(preview_img)
                                                preview_img = rgb_preview
                                            # Resize preview for display
                                            preview_display = preview_img.resize((400, int(400 * preview_img.size[1] / preview_img.size[0])))
                                            st.image(preview_display, caption="Signature Preview", width=400)
                                    except Exception as e:
                                        pass
                            with col2:
                                signature_date = st.date_input("Signature Date", value=datetime.now().date(), key="travel_sig_date")
                            
                            signature = signature_text.strip() if signature_text else ""
                            
                            submitted = st.form_submit_button("Generate PDF")
                        
                        if submitted:
                            # Validate required Traveler Information fields
                            missing_fields = []
                            if not name or not name.strip():
                                missing_fields.append("Name")
                            if not address1 or not address1.strip():
                                missing_fields.append("Address Line 1")
                            if not city or not city.strip():
                                missing_fields.append("City")
                            if not state or not state.strip():
                                missing_fields.append("State")
                            if not zip_code or not zip_code.strip():
                                missing_fields.append("Zip")
                            if not destination or not destination.strip():
                                missing_fields.append("Destination")
                            if not email or not email.strip():
                                missing_fields.append("Email Address")
                            if not purpose_of_travel or not purpose_of_travel.strip():
                                missing_fields.append("Purpose of Travel")
                            if not attendees or not attendees.strip():
                                missing_fields.append("Attendees")
                            if not deliverables or not deliverables.strip():
                                missing_fields.append("Deliverables")
                            
                            if missing_fields:
                                st.warning(f"⚠️ Please fill in all required fields: {', '.join(missing_fields)}")
                                st.stop()
                            
                            # Check for any input validation errors (check all number inputs)
                            has_validation_errors = False
                            # Check all input keys that might have errors
                            input_prefixes = ['travel_mileage_', 'travel_airfare_', 'travel_ground_', 'travel_parking_', 'travel_lodging_', 'travel_baggage_', 'travel_misc_', 'travel_misc2_']
                            for key in st.session_state.keys():
                                if key.endswith('_has_error') and st.session_state[key]:
                                    # Check if this is one of our input fields
                                    base_key = key.replace('_has_error', '')
                                    if any(base_key.startswith(prefix) for prefix in input_prefixes):
                                        has_validation_errors = True
                                        break
                            
                            if has_validation_errors:
                                st.warning("⚠️ **Cannot generate PDF: Please fix all invalid input fields above.**")
                                st.stop()
                            # Calculate totals
                            total_airfare = sum(airfare)
                            total_ground_transport = sum(ground_transport)
                            total_parking = sum(parking)
                            total_lodging = sum(lodging)
                            total_baggage = sum(baggage)
                            total_misc = sum(misc) + sum(misc2)  # Include both misc rows in total
                            # Calculate adjusted per diem with meal deductions
                            days_with_dates = [i for i, d in enumerate(per_diem_dates) if d and str(d).strip()]
                            num_days = len(days_with_dates)
                            first_day_idx = days_with_dates[0] if days_with_dates else 0
                            last_day_idx = days_with_dates[-1] if days_with_dates else 0
                            
                            meal_deductions = {
                                68: { 'breakfast': 16, 'lunch': 19, 'dinner': 28, 'incidental': 5, 'first_last': 51.00 },
                                74: { 'breakfast': 18, 'lunch': 20, 'dinner': 31, 'incidental': 5, 'first_last': 55.50 },
                                80: { 'breakfast': 20, 'lunch': 22, 'dinner': 33, 'incidental': 5, 'first_last': 60.00 },
                                86: { 'breakfast': 22, 'lunch': 23, 'dinner': 36, 'incidental': 5, 'first_last': 64.50 },
                                92: { 'breakfast': 23, 'lunch': 26, 'dinner': 38, 'incidental': 5, 'first_last': 69.00 },
                            }
                            adjusted_per_diem_daily = []
                            for i in range(len(per_diem_dates)):
                                if i < len(per_diem_dates) and per_diem_dates[i] and str(per_diem_dates[i]).strip():
                                    base_per_diem = int(per_diem_amounts[i]) if (i < len(per_diem_amounts) and per_diem_amounts[i]) else 80
                                    deducts = meal_deductions.get(base_per_diem, meal_deductions[80])
                                    deduction_total = 0.0
                                    if i < len(breakfast_checks) and breakfast_checks[i]:
                                        deduction_total += deducts['breakfast']
                                    if i < len(lunch_checks) and lunch_checks[i]:
                                        deduction_total += deducts['lunch']
                                    if i < len(dinner_checks) and dinner_checks[i]:
                                        deduction_total += deducts['dinner']
                                    # Base already includes incidentals; do not add +$5 here
                                    pre75_total = max(0.0, float(base_per_diem) - deduction_total)
                                    # Apply 75% for first and last day
                                    if i == first_day_idx or i == last_day_idx:
                                        final_per_diem = round(pre75_total * 0.75, 2)
                                    else:
                                        final_per_diem = round(pre75_total, 2)
                                    
                                    adjusted_per_diem_daily.append(final_per_diem)
                                else:
                                    adjusted_per_diem_daily.append(0.0)
                            
                            total_per_diem = sum(adjusted_per_diem_daily)
                            total_amount_due = (total_mileage + total_airfare + total_ground_transport + 
                                              total_parking + total_lodging + total_baggage + 
                                              total_misc + total_per_diem)
                            
                            # Store support files in session state for later upload
                            if support_files:
                                # Store file bytes and metadata in session state
                                file_data_list = []
                                for file in support_files:
                                    file.seek(0)  # Reset file pointer
                                    file_bytes = file.read()
                                    file.seek(0)  # Reset again for potential reuse
                                    file_data_list.append({
                                        'name': file.name,
                                        'bytes': file_bytes,
                                        'type': file.type
                                    })
                                st.session_state['travel_support_files_data'] = file_data_list
                            else:
                                st.session_state['travel_support_files_data'] = []
                            
                            form_data = {
                                'name': name,
                                'address1': address1,
                                'address2': address2,
                                'city': city,
                                'state': state,
                                'zip': zip_code,
                                'organization': organization,
                                'destination': destination,
                                'departure_date': departure_date.strftime('%m/%d/%Y') if departure_date else '',
                                'return_date': return_date.strftime('%m/%d/%Y') if return_date else '',
                                'email': email,
                                'purpose_of_travel': purpose_of_travel,
                                'objective': objective,
                                'attendees': attendees,
                                'deliverables': deliverables,
                                'support_files': '',  # Will be updated after upload
                                'mileage_dates': mileage_dates,
                                'mileage_amounts': mileage_amounts,
                                'total_mileage': total_mileage,
                                'expense_dates': expense_dates,
                                'airfare': airfare,
                                'ground_transport': ground_transport,
                                'parking': parking,
                                'lodging': lodging,
                                'baggage': baggage,
                                'misc': misc,
                                'misc2': misc2,
                                'misc_desc1': misc_desc1,
                                'misc_desc2': misc_desc2,
                                'total_airfare': total_airfare,
                                'total_ground_transport': total_ground_transport,
                                'total_parking': total_parking,
                                'total_lodging': total_lodging,
                                'total_baggage': total_baggage,
                                'total_misc': total_misc,
                                'per_diem_dates': per_diem_dates,
                                'per_diem_amounts': per_diem_amounts,
                                'breakfast_checks': breakfast_checks,
                                'lunch_checks': lunch_checks,
                                'dinner_checks': dinner_checks,
                                'total_per_diem': total_per_diem,
                                'total_amount_due': total_amount_due,
                                'signature': signature,
                                'signature_date': signature_date.strftime('%m/%d/%Y') if signature_date else ''
                            }
                            
                            # Store for review step
                            st.session_state['travel_review_data'] = form_data
                            st.success("Please review the information below and approve to finalize.")

                        # Review & Approve pane
                        if 'travel_review_data' in st.session_state:
                            review = st.session_state['travel_review_data']
                            st.subheader("Review & Approve")
                            colA, colB = st.columns(2)
                            with colA:
                                st.markdown("**Traveler**")
                                traveler_html = f"""
                                <div style='border:1px solid #e0e0e0;border-radius:8px;padding:12px;background:#fafafa;'>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Name</span><strong>{review.get('name','')}</strong>
                                  </div>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Organization</span><strong>{review.get('organization','')}</strong>
                                  </div>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Destination</span><strong>{review.get('destination','')}</strong>
                                  </div>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Email</span><strong>{review.get('email','')}</strong>
                                  </div>
                                </div>
                                """
                                st.markdown(traveler_html, unsafe_allow_html=True)
                            with colB:
                                st.markdown("**Trip**")
                                trip_html = f"""
                                <div style='border:1px solid #e0e0e0;border-radius:8px;padding:12px;background:#fafafa;'>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Departure Date</span><strong>{review.get('departure_date','')}</strong>
                                  </div>
                                  <div style='display:flex;justify-content:space-between;padding:4px 0;'>
                                    <span style='color:#555;'>Return Date</span><strong>{review.get('return_date','')}</strong>
                                  </div>
                                </div>
                                """
                                st.markdown(trip_html, unsafe_allow_html=True)
                            st.markdown("**Totals**")
                            totals_html = f"""
                            <table style='width:100%;border-collapse:collapse;border:1px solid #eee;'>
                              <thead>
                                <tr style='background:#f5f5f5;'>
                                  <th style='text-align:left;padding:8px;border-bottom:1px solid #eee;'>Category</th>
                                  <th style='text-align:right;padding:8px;border-bottom:1px solid #eee;'>Amount</th>
                                </tr>
                              </thead>
                              <tbody>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Mileage</td><td style='padding:8px;text-align:right;'>${int(review.get('total_mileage',0))}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Airfare</td><td style='padding:8px;text-align:right;'>${review.get('total_airfare',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Ground Transport</td><td style='padding:8px;text-align:right;'>${review.get('total_ground_transport',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Parking</td><td style='padding:8px;text-align:right;'>${review.get('total_parking',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Lodging</td><td style='padding:8px;text-align:right;'>${review.get('total_lodging',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Baggage</td><td style='padding:8px;text-align:right;'>${review.get('total_baggage',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Miscellaneous</td><td style='padding:8px;text-align:right;'>${review.get('total_misc',0):.2f}</td></tr>
                                <tr><td style='padding:8px;border-bottom:1px solid #f0f0f0;'>Per Diem</td><td style='padding:8px;text-align:right;'>${review.get('total_per_diem',0):.2f}</td></tr>
                                <tr style='background:#fff8f8;font-weight:600;'>
                                  <td style='padding:8px;border-top:1px solid #eee;'>Total Amount Due</td>
                                  <td style='padding:8px;text-align:right;border-top:1px solid #eee;'>${review.get('total_amount_due',0):.2f}</td>
                                </tr>
                              </tbody>
                            </table>
                            """
                            st.markdown(totals_html, unsafe_allow_html=True)
                            approved = st.checkbox("I have reviewed and approve this travel form.", key="travel_approve_review")
                            generate_now = st.button("✅ Finalize and Send for Approval", disabled=not approved, key="travel_generate_now", type="primary", use_container_width=True)
                            if generate_now and approved:
                                # Upload support files to Google Drive if provided
                                support_files_links = ""
                                if 'travel_support_files_data' in st.session_state and st.session_state['travel_support_files_data']:
                                    try:
                                        folder_id_travel = "1aDE0N_duNN6w8rLDLX5HHeychyhLIqbo"
                                        links = []
                                        name_for_files = review.get('name', 'Unknown')
                                        destination_for_files = review.get('destination', 'Unknown')
                                        
                                        for file_data in st.session_state['travel_support_files_data']:
                                            # Create unique filename
                                            renamed_filename = f"Travel_{name_for_files.replace(' ', '_')}_{destination_for_files.replace(' ', '_')}_{file_data['name']}"
                                            
                                            # Create a file-like object from bytes with required attributes
                                            file_obj = io.BytesIO(file_data['bytes'])
                                            file_obj.name = file_data['name']
                                            file_obj.type = file_data.get('type', 'application/octet-stream')
                                            
                                            # Upload to Google Drive
                                            link = upload_file_to_drive(
                                                file=file_obj,
                                                filename=renamed_filename,
                                                folder_id=folder_id_travel,
                                                creds_dict=st.secrets["gcp_service_account"]
                                            )
                                            links.append(link)
                                        
                                        support_files_links = ", ".join(links)
                                        st.success("✅ Support files uploaded to Google Drive!")
                                        
                                        # Update review data with file links
                                        review['support_files'] = support_files_links
                                        
                                        # Clear file data from session state to free memory
                                        del st.session_state['travel_support_files_data']
                                        
                                    except Exception as e:
                                        st.warning(f"⚠️ Error uploading support files: {str(e)}")
                                
                                # Save to Google Sheets when PDF is generated
                                try:
                                    df_travel = load_travel_sheet()
                                    
                                    # Create new row for travel sheet
                                    new_travel_row = {
                                        'Name': review.get('name', ''),
                                        'Email': review.get('email', ''),
                                        'Destination': review.get('destination', ''),
                                        'Purpose of Travel': review.get('purpose_of_travel', ''),
                                        'Objective': review.get('objective', ''),
                                        'Attendees': review.get('attendees', ''),
                                        'Departure Date': review.get('departure_date', ''),
                                        'Return Date': review.get('return_date', ''),
                                        'Deliverables': review.get('deliverables', ''),
                                        'Support Files': support_files_links,
                                        'Submission Date': datetime.now().strftime('%Y-%m-%d'),
                                        'PDF Link': '',  # Will be filled when sent for approval
                                        # Traveler information
                                        'Address1': review.get('address1', ''),
                                        'Address2': review.get('address2', ''),
                                        'City': review.get('city', ''),
                                        'State': review.get('state', ''),
                                        'Zip': review.get('zip', ''),
                                        'Organization': review.get('organization', 'Georgetown University'),
                                        'Signature': review.get('signature', ''),
                                        'Signature Date': review.get('signature_date', ''),
                                        # Expense details - stored as JSON strings
                                        'Mileage Dates': json.dumps(review.get('mileage_dates', [])),
                                        'Mileage Amounts': json.dumps(review.get('mileage_amounts', [])),
                                        'Total Mileage': review.get('total_mileage', 0),
                                        'Expense Dates': json.dumps(review.get('expense_dates', [])),
                                        'Airfare': json.dumps(review.get('airfare', [])),
                                        'Ground Transport': json.dumps(review.get('ground_transport', [])),
                                        'Parking': json.dumps(review.get('parking', [])),
                                        'Lodging': json.dumps(review.get('lodging', [])),
                                        'Baggage': json.dumps(review.get('baggage', [])),
                                        'Misc': json.dumps(review.get('misc', [])),
                                        'Misc2': json.dumps(review.get('misc2', [])),
                                        'Misc Desc1': review.get('misc_desc1', ''),
                                        'Misc Desc2': review.get('misc_desc2', ''),
                                        'Total Airfare': review.get('total_airfare', 0),
                                        'Total Ground Transport': review.get('total_ground_transport', 0),
                                        'Total Parking': review.get('total_parking', 0),
                                        'Total Lodging': review.get('total_lodging', 0),
                                        'Total Baggage': review.get('total_baggage', 0),
                                        'Total Misc': review.get('total_misc', 0),
                                        'Per Diem Dates': json.dumps(review.get('per_diem_dates', [])),
                                        'Per Diem Amounts': json.dumps(review.get('per_diem_amounts', [])),
                                        'Breakfast Checks': json.dumps(review.get('breakfast_checks', [])),
                                        'Lunch Checks': json.dumps(review.get('lunch_checks', [])),
                                        'Dinner Checks': json.dumps(review.get('dinner_checks', [])),
                                        'Total Per Diem': review.get('total_per_diem', 0),
                                        'Total Amount Due': review.get('total_amount_due', 0),
                                        # Approval fields (all possible approvers)
                                        'Kemisha Approval Status': '',
                                        'Mabintou Approval Status': '',
                                        'Jen Approval Status': '',
                                        'Lauren Approval Status': '',
                                        'Kemisha Approval Date': '',
                                        'Mabintou Approval Date': '',
                                        'Jen Approval Date': '',
                                        'Lauren Approval Date': '',
                                        'Kemisha Signature': '',
                                        'Mabintou Signature': '',
                                        'Jen Signature': '',
                                        'Lauren Signature': '',
                                        'Kemisha Note': '',
                                        'Mabintou Note': '',
                                        'Jen Note': '',
                                        'Lauren Note': '',
                                    }
                                    
                                    new_travel_data = pd.DataFrame([new_travel_row])
                                    
                                    # Append new data to existing travel sheet
                                    updated_travel_sheet = pd.concat([df_travel, new_travel_data], ignore_index=True)
                                    updated_travel_sheet = updated_travel_sheet.fillna("")
                                    
                                    # Update Google Sheet
                                    spreadsheet_travel = client.open('HRSA64_TA_Request')
                                    try:
                                        worksheet_travel = spreadsheet_travel.worksheet('Travel')
                                    except:
                                        # Create worksheet if it doesn't exist
                                        worksheet_travel = spreadsheet_travel.add_worksheet(title='Travel', rows=1000, cols=20)
                                    
                                    worksheet_travel.update([updated_travel_sheet.columns.values.tolist()] + updated_travel_sheet.values.tolist())
                                    
                                    # Clear cache to refresh data
                                    st.cache_data.clear()
                                    
                                    st.success("✅ Travel form data saved to Google Sheets!")
                                    
                                except Exception as e:
                                    st.warning(f"⚠️ Error saving to Google Sheets: {str(e)}")
                                
                                pdf_buffer = create_pdf(review, ws)
                                pdf_filename = f"Travel_Authorization_Form_{review.get('name','')}_{review.get('departure_date','')}_{review.get('return_date','')}.pdf"
                                
                                st.success("✅ PDF generated successfully!")
                                
                                # Upload PDF to Google Drive and send for approval
                                try:
                                    folder_id_travel_pdf = "1_O_L-jPR7bldiryRNB3WxbAaG8VqvmCt"
                                    pdf_file_obj = io.BytesIO(pdf_buffer.getvalue())
                                    pdf_file_obj.name = pdf_filename
                                    pdf_file_obj.type = 'application/pdf'
                                    
                                    pdf_link = upload_file_to_drive(
                                        file=pdf_file_obj,
                                        filename=pdf_filename,
                                        folder_id=folder_id_travel_pdf,
                                        creds_dict=st.secrets["gcp_service_account"]
                                    )
                                    
                                    st.success("✅ PDF uploaded to Google Drive!")
                                    
                                    # Update Google Sheet with PDF link and status
                                    # Reload the sheet to get the latest data
                                    st.cache_data.clear()
                                    df_travel = load_travel_sheet()
                                    
                                    if df_travel.empty:
                                        st.error("❌ No travel forms found in the sheet. Please submit the form first.")
                                        st.stop()
                                    
                                    # Find the row that matches this submission (by name and submission date)
                                    submission_date = datetime.now().strftime('%Y-%m-%d')
                                    traveler_name = review.get('name', '')
                                    
                                    # Try to find matching row
                                    row_idx = None
                                    if 'Name' in df_travel.columns and 'Submission Date' in df_travel.columns:
                                        matching_rows = df_travel[
                                            (df_travel['Name'].astype(str) == str(traveler_name)) &
                                            (df_travel['Submission Date'].astype(str).str.contains(submission_date, na=False))
                                        ]
                                        
                                        if not matching_rows.empty:
                                            # Use the most recent matching row (last one)
                                            row_idx = matching_rows.index[-1]
                                    
                                    # Fallback to last row if no match found
                                    if row_idx is None:
                                        row_idx = len(df_travel) - 1
                                    
                                    if row_idx >= 0 and row_idx < len(df_travel):
                                        updated_df_travel = df_travel.copy()
                                        
                                        # Determine approval routing based on traveler
                                        traveler_email = review.get('email', '').lower()
                                        traveler_name_lower = traveler_name.lower()
                                        
                                        # Check if traveler is Kemisha or Mabintou
                                        is_kemisha_traveler = (traveler_email == 'kd802@georgetown.edu' or 
                                                              'kemisha' in traveler_name_lower)
                                        is_mabintou_traveler = (traveler_email == 'mo887@georgetown.edu' or 
                                                               'mabintou' in traveler_name_lower)
                                        
                                        # Determine approvers based on routing rules:
                                        # - Kemisha's requests → Mabintou + Jen
                                        # - Mabintou's requests → Lauren + Kemisha
                                        # - Others → Mabintou + Kemisha (or alternatives if out)
                                        
                                        if is_kemisha_traveler:
                                            # Kemisha's requests go to Mabintou and Jen
                                            approver1_email = "mo887@georgetown.edu"
                                            approver1_name = "Mabintou Ouattara"
                                            approver1_status_col = 'Mabintou Approval Status'
                                            approver2_email = "Jenevieve.Opoku@georgetown.edu"
                                            approver2_name = "Jenevieve Opoku"
                                            approver2_status_col = 'Jen Approval Status'
                                        elif is_mabintou_traveler:
                                            # Mabintou's requests go to Lauren and Kemisha
                                            approver1_email = "lm1353@georgetown.edu"
                                            approver1_name = "Lauren Mathae"
                                            approver1_status_col = 'Lauren Approval Status'
                                            approver2_email = "kd802@georgetown.edu"
                                            approver2_name = "Kemisha Denny"
                                            approver2_status_col = 'Kemisha Approval Status'
                                        else:
                                            # Default: Mabintou + Kemisha (with alternatives if out)
                                            # Check if Kemisha or Mabintou are out and use alternatives
                                            # If Kemisha is out → Jen is alternative for lead
                                            # If Mabintou is out → Lauren is alternative
                                            
                                            # Out of Office configuration (can be updated as needed)
                                            # Set to True if the person is out of office
                                            out_of_office = {
                                                'kemisha': False,  # Set to True if Kemisha is out
                                                'mabintou': False  # Set to True if Mabintou is out
                                            }
                                            
                                            # Determine approvers with alternatives
                                            if out_of_office.get('mabintou', False):
                                                # Mabintou is out, use Lauren as alternative
                                                approver1_email = "lm1353@georgetown.edu"
                                                approver1_name = "Lauren Mathae"
                                                approver1_status_col = 'Lauren Approval Status'
                                            else:
                                                # Mabintou is available
                                                approver1_email = "mo887@georgetown.edu"
                                                approver1_name = "Mabintou Ouattara"
                                                approver1_status_col = 'Mabintou Approval Status'
                                            
                                            if out_of_office.get('kemisha', False):
                                                # Kemisha is out, use Jen as alternative for lead
                                                approver2_email = "Jenevieve.Opoku@georgetown.edu"
                                                approver2_name = "Jenevieve Opoku"
                                                approver2_status_col = 'Jen Approval Status'
                                            else:
                                                # Kemisha is available
                                                approver2_email = "kd802@georgetown.edu"
                                                approver2_name = "Kemisha Denny"
                                                approver2_status_col = 'Kemisha Approval Status'
                                        
                                        # Ensure required columns exist
                                        required_cols = ['PDF Link', approver1_status_col, approver2_status_col]
                                        for col in required_cols:
                                            if col not in updated_df_travel.columns:
                                                updated_df_travel[col] = ''
                                        
                                        updated_df_travel.loc[row_idx, 'PDF Link'] = pdf_link
                                        updated_df_travel.loc[row_idx, approver1_status_col] = 'pending'
                                        updated_df_travel.loc[row_idx, approver2_status_col] = 'pending'
                                        
                                        updated_df_travel = updated_df_travel.fillna("")
                                        spreadsheet_travel = client.open('HRSA64_TA_Request')
                                        try:
                                            worksheet_travel = spreadsheet_travel.worksheet('Travel')
                                        except:
                                            worksheet_travel = spreadsheet_travel.add_worksheet(title='Travel', rows=1000, cols=20)
                                        
                                        worksheet_travel.update([updated_df_travel.columns.values.tolist()] + updated_df_travel.values.tolist())
                                        
                                        st.success("✅ Travel form status updated in Google Sheets!")
                                        
                                        # Send email notification to both approvers (determined dynamically above)
                                        traveler_name = review.get('name', 'Unknown')
                                        destination = review.get('destination', 'Unknown')
                                        departure_date = review.get('departure_date', 'Unknown')
                                        return_date = review.get('return_date', 'Unknown')
                                        total_amount = review.get('total_amount_due', 0)
                                        
                                        # Helper function to send email to an approver
                                        def send_approval_email(approver_email, approver_name):
                                            email_subject = f"Travel Authorization Form Pending Approval - {traveler_name}"
                                            email_body = f"""
Dear {approver_name},

A new travel authorization form has been submitted and is pending your approval.

Travel Details:
- Traveler: {traveler_name}
- Destination: {destination}
- Departure Date: {departure_date}
- Return Date: {return_date}
- Total Amount Due: ${total_amount:.2f}

PDF Link: {pdf_link}

Please review and approve this travel authorization form via the GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                            """
                                            try:
                                                send_email_mailjet(
                                                    to_email=approver_email,
                                                    subject=email_subject,
                                                    body=email_body.strip()
                                                )
                                                return True, f"✅ Email sent successfully to {approver_name} ({approver_email})"
                                            except Exception as e:
                                                return False, f"⚠️ Failed to send email to {approver_name}: {str(e)}"
                                        
                                        # Send emails to both approvers
                                        email_success_count = 0
                                        email_messages = []
                                        
                                        # Send to approver 1
                                        success1, msg1 = send_approval_email(approver1_email, approver1_name)
                                        if success1:
                                            email_success_count += 1
                                        email_messages.append(msg1)
                                        
                                        # Send to approver 2
                                        success2, msg2 = send_approval_email(approver2_email, approver2_name)
                                        if success2:
                                            email_success_count += 1
                                        email_messages.append(msg2)
                                        
                                        # Display email results
                                        for msg in email_messages:
                                            if msg.startswith("✅"):
                                                st.success(msg)
                                            else:
                                                st.warning(msg)
                                        
                                        if email_success_count == 2:
                                            st.success("🎉 Travel form sent for approval! Both coordinators have been notified.")
                                        elif email_success_count == 1:
                                            st.warning("⚠️ PDF uploaded and one email sent, but one email failed. Please check the messages above.")
                                        else:
                                            st.error("❌ PDF uploaded but failed to send emails to coordinators. Please contact support.")
                                        
                                        # Clear session state
                                        if 'travel_pdf_buffer' in st.session_state:
                                            del st.session_state['travel_pdf_buffer']
                                        if 'travel_pdf_filename' in st.session_state:
                                            del st.session_state['travel_pdf_filename']
                                        if 'travel_review_for_approval' in st.session_state:
                                            del st.session_state['travel_review_for_approval']
                                        if 'travel_review_data' in st.session_state:
                                            del st.session_state['travel_review_data']
                                        if 'travel_submission_date' in st.session_state:
                                            del st.session_state['travel_submission_date']
                                        
                                        st.cache_data.clear()
                                        time.sleep(3)
                                        st.rerun()
                                    else:
                                        st.error("❌ Could not find the travel form entry to update. Please try again.")
                                        st.stop()
                                        
                                except Exception as e:
                                    st.error(f"❌ Error sending for approval: {str(e)}")
                                    st.exception(e)
                    
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        st.exception(e)


                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)


                with st.expander("📦 **CHECK & SUBMIT DELIVERY LOG**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                📦 Delivery Management Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Review your previous deliveries and submit new ones. Track all your completed work including reports, dashboards, and data.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Upper section: Previous Deliveries
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                📊 Your Previous Deliveries
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Get delivery data properly
                    df_del_staff = df_del[df_del["Submitted By"] == staff_name].copy()
                    if not df_del_staff.empty:
                        # Remove columns we don't want to display
                        display_cols = [col for col in df_del_staff.columns if col not in ['Submitted By', 'Submission Date']]
                        df_del_staff_display = df_del_staff[display_cols].copy()
                        
                        # Sort by Date of Delivery (most recent first)
                        df_del_staff_display["Date of Delivery"] = pd.to_datetime(df_del_staff_display["Date of Delivery"], errors="coerce")
                        df_del_staff_display = df_del_staff_display.sort_values("Date of Delivery", ascending=True)
                        df_del_staff_display["Date of Delivery"] = df_del_staff_display["Date of Delivery"].dt.strftime("%Y-%m-%d")
                        
                        # Add summary stats
                        total_deliveries = len(df_del_staff_display)
                        recent_deliveries = len(df_del_staff_display[df_del_staff_display["Date of Delivery"] >= (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")])
                        
                        st.markdown(f"""
                            <div style='background: #e3f2fd; border-radius: 10px; padding: 1em; margin-top: 1em; text-align: center;'>
                                <div style='display: flex; justify-content: space-around;'>
                                    <div>
                                        <div style='font-size: 1.5em; font-weight: bold; color: #1976d2;'>{total_deliveries}</div>
                                        <div style='font-size: 0.9em; color: #666;'>Total Deliveries</div>
                                    </div>
                                    <div>
                                        <div style='font-size: 1.5em; font-weight: bold; color: #388e3c;'>{recent_deliveries}</div>
                                        <div style='font-size: 0.9em; color: #666;'>Last 30 Days</div>
                                    </div>
                                </div>
                            </div>
                        """, unsafe_allow_html=True)

                        st.dataframe(df_del_staff_display.reset_index(drop=True), use_container_width=True)
                        

                    else:
                        st.markdown("""
                            <div style='background: #fff3e0; border-radius: 15px; padding: 2em; text-align: center; border: 2px dashed #ff9800;'>
                                <div style='font-size: 3em; margin-bottom: 0.5em;'>📦</div>
                                <h4 style='color: #e65100; margin-bottom: 0.5em;'>No Previous Deliveries</h4>
                                <p style='color: #666; margin: 0;'>You haven't logged any deliveries yet. Start by submitting your first delivery below!</p>
                            </div>
                        """, unsafe_allow_html=True)

                    # Middle section: View Deliveries by Ticket ID
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 2em; margin-top: 2em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                🔍 View Deliveries by Ticket ID
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Get ticket IDs assigned to this staff member
                    assigned_tickets_del = df[df["Assigned Coach"] == staff_name]["Ticket ID"].dropna().astype(str).unique().tolist()
                    assigned_tickets_del_sorted = sorted(assigned_tickets_del)
                    
                    if assigned_tickets_del_sorted:
                        selected_ticket_view_del = st.selectbox(
                            "Select a Ticket ID to view all deliveries",
                            options=[""] + assigned_tickets_del_sorted,
                            index=0,
                            key='view_deliveries_ticket_staff',
                            help="Select a ticket ID from your assigned requests to view all deliveries for that ticket"
                        )
                        
                        if selected_ticket_view_del:
                            # Get all deliveries for this ticket ID (regardless of who submitted)
                            # Handle NaN values properly
                            df_ticket_del = df_del[
                                (df_del["Ticket ID"].notna()) & 
                                (df_del["Ticket ID"].astype(str) == selected_ticket_view_del)
                            ].copy()
                            
                            if not df_ticket_del.empty:
                                # Remove columns we don't want to display
                                display_cols_ticket_del = [col for col in df_ticket_del.columns if col not in ['Submission Date']]
                                df_ticket_del_display = df_ticket_del[display_cols_ticket_del].copy()
                                
                                # Sort by Date of Delivery (most recent first)
                                df_ticket_del_display["Date of Delivery"] = pd.to_datetime(df_ticket_del_display["Date of Delivery"], errors="coerce")
                                df_ticket_del_display = df_ticket_del_display.sort_values("Date of Delivery", ascending=True)
                                df_ticket_del_display["Date of Delivery"] = df_ticket_del_display["Date of Delivery"].dt.strftime("%Y-%m-%d")
                                
                                st.markdown(f"**All deliveries for Ticket ID: {selected_ticket_view_del}**")
                                st.dataframe(df_ticket_del_display.reset_index(drop=True), use_container_width=True)
                            else:
                                st.info(f"No deliveries found for Ticket ID: {selected_ticket_view_del}")
                    else:
                        st.info("No assigned ticket IDs available to view deliveries.")

                    # Lower section: Submit New Delivery
                    st.markdown("""
                        <div style='background: #f8f9fa; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); padding: 1.5em; margin-bottom: 1em;'>
                            <h3 style='color: #1a237e; font-family: "Segoe UI", sans-serif; font-weight: 700; margin-bottom: 1em; text-align: center;'>
                                ✍️ Submit New Delivery
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    lis_ticket1 = df["Ticket ID"].unique().tolist()

                    # Delivery Log form
                    col1, col2 = st.columns(2)
                    with col1:
                        ticket_id_del = st.selectbox("Ticket ID *",lis_ticket1, index=None,
                            placeholder="Select option...",key='delivery1')
                    with col2:
                        date_del = st.date_input("Date of Delivery *",value=datetime.today().date())

                    list_delivery = [
                        "Report", "Email Reply", "Dashboard", "New Data Points","Peer Learning Facilitation", "TA Meeting", "Other"
                    ]

                    type_delivery = st.selectbox(
                        "Type of Delivery *",
                        list_delivery,
                        index=None,
                        placeholder="Select option..."
                    )

                    # If "Other" is selected, show a text input for custom value
                    if type_delivery == "Other":
                        type_delivery_other = st.text_input("Please specify the Type of Delivery *")
                        if type_delivery_other:
                            type_delivery = type_delivery_other 
                    delivery_description = st.text_area("Short Summary *", placeholder='Enter text', height=150,key='delivery_description1') 
                    document_del = st.file_uploader(
                        "Upload any files or attachments that are relevant to this delivery.",accept_multiple_files=True
                    )

                    # Submit button
                    st.markdown("""
                        <style>
                        .stButton > button {
                            width: 100%;
                            background-color: #cdb4db;
                            color: black;
                            font-family: Arial, "Segoe UI", sans-serif;
                            font-weight: 600;
                            border-radius: 8px;
                            padding: 0.6em;
                            margin-top: 1em;
                        }
                        </style>
                    """, unsafe_allow_html=True)

                    # Submit logic
                    if st.button("Submit",key='delivery_submit1'):
                        errors = []
                        drive_links_del = ""  # Ensure always defined
                        # Required field checks
                        if not ticket_id_del: errors.append("Ticket ID is required.")
                        if not date_del: errors.append("Date of delivery is required.")
                        if not type_delivery: errors.append("Type of delivery is required.")
                        if not delivery_description: errors.append("Short summary is required.")

                        # Show warnings or success
                        if errors:
                            for error in errors:
                                st.warning(error)
                        else:
                            # Only upload files if all validation passes
                            if document_del:
                                try:
                                    folder_id_del = "1gXfWxys2cxd67YDk8zKPmG_mLGID4qL2" 
                                    links_del = []
                                    upload_count = 0
                                    for file in document_del:
                                        # Rename file as: GU0001_filename.pdf
                                        renamed_filename = f"{ticket_id_del}_{file.name}"
                                        link = upload_file_to_drive(
                                            file=file,
                                            filename=renamed_filename,
                                            folder_id=folder_id_del,
                                            creds_dict=st.secrets["gcp_service_account"]
                                        )
                                        links_del.append(link)
                                        upload_count += 1
                                        st.success(f"✅ Successfully uploaded: {file.name}")
                                    drive_links_del = ", ".join(links_del)
                                    if upload_count > 0:
                                        st.success(f"✅ All {upload_count} file(s) uploaded successfully to Google Drive!")    
                                except Exception as e:
                                    st.error(f"❌ Error uploading file(s) to Google Drive: {str(e)}")

                            new_row_del = {
                                'Ticket ID': ticket_id_del,
                                "Date of Delivery": date_del.strftime("%Y-%m-%d"),  # Convert to string
                                "Type of Delivery": type_delivery,
                                "Short Summary": delivery_description,
                                "Document": drive_links_del,
                                "Submitted By": staff_name,
                                "Submission Date": datetime.today().strftime("%Y-%m-%d %H:%M")
                            }
                            new_data_del = pd.DataFrame([new_row_del])

                            try:
                                # Append new data to Google Sheet
                                updated_sheet2 = pd.concat([df_del, new_data_del], ignore_index=True)
                                updated_sheet2= updated_sheet2.applymap(
                                    lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, pd.Timestamp)) else x
                                )
                                # Replace NaN with empty strings to ensure JSON compatibility
                                updated_sheet2 = updated_sheet2.fillna("")
                                spreadsheet3 = client.open('HRSA64_TA_Request')
                                worksheet3 = spreadsheet3.worksheet('Delivery')
                                worksheet3.update([updated_sheet2.columns.values.tolist()] + updated_sheet2.values.tolist())

                                # Clear cache to refresh data
                                st.cache_data.clear()
                                
                                st.success("✅ Submission successful!")
                                time.sleep(2)
                                st.rerun()

                            except Exception as e:
                                st.error(f"Error updating Google Sheets: {str(e)}")

                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                # --- Section 1: Mark as Completed
                with st.expander("✅ **MARK REQUESTS AS COMPLETED**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                ✅ Request Completion Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Mark your assigned requests as completed when finished. Review request details and finalize your work efficiently.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                    if staff_df.empty:
                        st.info("No requests currently in progress to mark as completed.")
                    else:
                        # Ensure datetime before using .dt
                        staff_df["Assigned Date"] = pd.to_datetime(staff_df["Assigned Date"], errors="coerce")
                        staff_df["Targeted Due Date"] = pd.to_datetime(staff_df["Targeted Due Date"], errors="coerce")

                        # Format dates
                        staff_df["Assigned Date"] = staff_df["Assigned Date"].dt.strftime("%Y-%m-%d")
                        staff_df["Targeted Due Date"] = staff_df["Targeted Due Date"].dt.strftime("%Y-%m-%d")

                        # Display clean table (exclude PriorityOrder column)
                        st.dataframe(staff_df[[
                            "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                            "Focus Area", "TA Type", "Assigned Date", "Targeted Due Date", "Priority", "TA Description","Document","Coordinator Comment History"
                        ]].reset_index(drop=True))

                        # Select request by index (row number in submitted_requests)
                        request_indices = staff_df.index.tolist()
                        selected_request_index = st.selectbox(
                            "Select a request to marked as completed",
                            options=request_indices,
                            format_func=lambda idx: f"{staff_df.at[idx, 'Ticket ID']} | {staff_df.at[idx, 'Name']} | {staff_df.at[idx, 'Jurisdiction']}",
                        )


                        # Submit completion
                        if st.button("✅ Mark as Completed"):
                            try:
                                # Map back to original df index
                                global_index = staff_df.loc[selected_request_index].name

                                # Copy + update
                                updated_df = df.copy()
                                updated_df.loc[global_index, "Status"] = "Completed"
                                updated_df.loc[global_index, "Close Date"] = datetime.today().strftime("%Y-%m-%d")

                                updated_df = updated_df.applymap(
                                    lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (pd.Timestamp, datetime)) and not pd.isna(x) else x
                                )
                                updated_df = updated_df.fillna("") 
                                spreadsheet1 = client.open('HRSA64_TA_Request')
                                worksheet1 = spreadsheet1.worksheet('Main')

                                # Push to Google Sheet
                                worksheet1.update([updated_df.columns.values.tolist()] + updated_df.values.tolist())

                                # Clear cache to refresh data
                                st.cache_data.clear()
                                
                                st.success("✅ Request marked as completed.")
                                time.sleep(2)
                                st.rerun()

                            except Exception as e:
                                st.error(f"Error updating Google Sheets: {str(e)}")

                    # --- Submit button styling (CSS injection)
                    st.markdown("""
                        <style>
                        .stButton > button {
                            width: 100%;
                            background-color: #cdb4db;
                            color: black;
                            font-weight: 600;
                            border-radius: 8px;
                            padding: 0.6em;
                            margin-top: 1em;
                        }
                        </style>
                    """, unsafe_allow_html=True)

                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

                # --- Section: View Completed Requests (Staff)
                with st.expander("✅ **COMPLETED REQUESTS**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                ✅ Completed Requests
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                View your completed TA requests.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    staff_completed_df = com_df.copy()
                    if staff_completed_df.empty:
                        st.info("You have no completed requests yet.")
                    else:
                        staff_completed_df["Assigned Date"] = pd.to_datetime(staff_completed_df["Assigned Date"], errors="coerce")
                        staff_completed_df["Targeted Due Date"] = pd.to_datetime(staff_completed_df["Targeted Due Date"], errors="coerce")
                        staff_completed_df["Close Date"] = pd.to_datetime(staff_completed_df["Close Date"], errors="coerce")

                        staff_completed_df["Assigned Date"] = staff_completed_df["Assigned Date"].dt.strftime("%Y-%m-%d")
                        staff_completed_df["Targeted Due Date"] = staff_completed_df["Targeted Due Date"].dt.strftime("%Y-%m-%d")
                        staff_completed_df["Close Date"] = staff_completed_df["Close Date"].dt.strftime("%Y-%m-%d")

                        st.dataframe(staff_completed_df[[
                            "Ticket ID","Jurisdiction", "Organization", "Name", "Title/Position", "Email Address", "Phone Number",
                            "Focus Area", "TA Type", "Priority", "Assigned Coach", "TA Description","Document",
                            "Assigned Date", "Targeted Due Date", "Close Date",
                            "Coordinator Comment History", "Staff Comment History", "Transfer History"
                        ]].reset_index(drop=True))

                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)

            elif st.session_state.role == "Research Assistant":
                # Add staff content here
                user_info = USERS.get(st.session_state.user_email)
                user_email = st.session_state.user_email
                ga_support_name = user_info["Research Assistant"]["name"] if user_info and "Research Assistant" in user_info else None
                st.markdown(
                    """
                    <div style='
                        display: flex;
                        flex-direction: column;
                        align-items: center;
                        justify-content: center;
                        background: #f8f9fa;
                        padding: 2em 0 1em 0;
                        border-radius: 18px;
                        box-shadow: 0 4px 24px rgba(0,0,0,0.07);
                        margin-bottom: 2em;
                    '>
                        <img src='https://raw.githubusercontent.com/JiaqinWu/HRSA64_Dash/main/Georgetown_logo_blueRGB.png' width='200' style='margin-bottom: 1em;'/>
                        <h1 style='
                            color: #1a237e;
                            font-family: "Segoe UI", "Arial", sans-serif;
                            font-weight: 700;
                            margin: 0;
                            font-size: 2.2em;
                            text-align: center;
                        '>🧑‍🎓 Research Assistant Dashboard</h1>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                # Personalized greeting
                if ga_support_name:
                    st.markdown(f"""
                    <div style='                      
                    background: #f8f9fa;                        
                    border-radius: 12px;                        
                    box-shadow: 0 2px 8px rgba(0,0,0,0.04);                        
                    padding: 1.2em 1em 1em 1em;                        
                    margin-bottom: 1.5em;                        
                    text-align: center;                        
                    font-family: Arial, "Segoe UI", sans-serif;                    
                    '>
                        <span style='                           
                        font-size: 1.15em;
                        font-weight: 700;
                        color: #1a237e;
                        letter-spacing: 0.5px;'>
                            👋 Welcome, {ga_support_name}!
                        </span>
                    </div>
                    """, unsafe_allow_html=True)


                # Filter requests assigned to current staff and In Progress
                ga_in_progress_df = df_support[(df_support["Student assigned"] == ga_support_name) & (df_support["Request status"].isin(["Not Started","In Progress"]))].copy()
                ga_completed_df = df_support[(df_support["Student assigned"] == ga_support_name) & (df_support["Request status"] == "Completed")].copy()


                # Ensure date columns are datetime
                ga_in_progress_df["Date"] = pd.to_datetime(ga_in_progress_df["Date"], errors="coerce")

                # --- Top Summary Cards
                col1, col2 = st.columns(2)
                col3, col4 = st.columns(2)
                # 1. Total In Progress
                total_in_progress = ga_in_progress_df.shape[0]
                total_complete = ga_completed_df.shape[0]

                # 2. Newly Assigned: within last 3 days
                recent_cutoff = datetime.today() + timedelta(days=3)
                newly_assigned = ga_in_progress_df[ga_in_progress_df["Date"] <= recent_cutoff].shape[0]

                # 3. Due within 2 weeks
                due_soon_cutoff = datetime.today() + timedelta(days=14)
                due_soon = ga_in_progress_df[ga_in_progress_df["Date"] <= due_soon_cutoff].shape[0]

                col1.metric("🟡 In Progress", total_in_progress)
                col2.metric("✅ Completed", total_complete)
                col3.metric("🆕 Coming in 3 Days", newly_assigned)
                col4.metric("📅 Coming in 2 Weeks", due_soon)

                style_metric_cards(border_left_color="#DBF227")

                # --- Section: View and Manage Submitted Support Requests
                with st.expander("📋 **VIEW & MANAGE SUPPORT REQUESTS**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                📋 Support Request Management Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                View all submitted support requests and update their status. Assign yourself to requests and track your progress efficiently.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Get all support requests
                    all_support_requests = df_support.copy()
                    
                    if all_support_requests.empty:
                        st.info("No support requests have been submitted yet.")
                    else:
                        # Convert date column
                        all_support_requests["Date"] = pd.to_datetime(all_support_requests["Date"], errors="coerce")
                        
                        # Format date for display
                        all_support_requests["Date"] = all_support_requests["Date"].dt.strftime("%Y-%m-%d")
                        
                        # Display all requests
                        st.markdown("#### 📝 All Submitted Support Requests")
                        st.dataframe(all_support_requests[[
                            "Date", "Time request needed", "Request description", "Anticipated Deliverable", 
                            "TAP Name", "TAP email", "Student assigned", "Student email", "Request status"
                        ]].sort_values(by="Date", ascending=True).reset_index(drop=True))

                        # Filter for unassigned requests (where Student assigned is empty or NaN)
                        unassigned_requests = all_support_requests[
                            (all_support_requests["Student assigned"].isna()) | 
                            (all_support_requests["Student assigned"] == "") |
                            (all_support_requests["Student assigned"] == "nan")
                        ].copy()

                        if not unassigned_requests.empty:
                            st.markdown("#### 🆕 Unassigned Requests")
                            
                            # Select request to assign
                            request_indices = unassigned_requests.index.tolist()
                            selected_request_idx = st.selectbox(
                                "Select a request to assign to yourself",
                                options=request_indices,
                                format_func=lambda idx: f"{unassigned_requests.at[idx, 'Date']} | {unassigned_requests.at[idx, 'TAP Name']} | {unassigned_requests.at[idx, 'Time request needed']}",
                                key='unassigned_requests'
                            )

                            if st.button("✅ Assign to Me (Not Started)", key='assign_not_started'):
                                try:
                                    updated_df_support = df_support.copy()
                                    updated_df_support.loc[selected_request_idx, "Student assigned"] = ga_support_name
                                    updated_df_support.loc[selected_request_idx, "Student email"] = st.session_state.user_email
                                    updated_df_support.loc[selected_request_idx, "Request status"] = "Not Started"

                                    # Update Google Sheet
                                    updated_df_support = updated_df_support.fillna("")
                                    spreadsheet_support = client.open('HRSA64_TA_Request')
                                    worksheet_support = spreadsheet_support.worksheet('GA_Support')
                                    worksheet_support.update([updated_df_support.columns.values.tolist()] + updated_df_support.values.tolist())

                                    st.cache_data.clear()
                                    st.success(f"Request assigned to you with status 'Not Started'!")
                                    
                                    # Send notification email to TAP
                                    tap_email = updated_df_support.loc[selected_request_idx, "TAP email"]
                                    tap_name = updated_df_support.loc[selected_request_idx, "TAP Name"]
                                    request_date = updated_df_support.loc[selected_request_idx, "Date"]
                                    request_time = updated_df_support.loc[selected_request_idx, "Time request needed"]
                                    request_description = updated_df_support.loc[selected_request_idx, "Request description"]
                                    anticipated_deliverable = updated_df_support.loc[selected_request_idx, "Anticipated Deliverable"]
                                    
                                    if tap_email and tap_email.strip():
                                        tap_subject = f"Support Request Assigned - {request_date} at {request_time}"
                                        tap_body = f"""
Dear {tap_name},

Your support request has been assigned to a research assistant.

Request Details:
- Date: {request_date}
- Time: {request_time}
- Request Description: {request_description}
- Anticipated Deliverable: {anticipated_deliverable}

Assigned Research Assistant:
- Name: {ga_support_name}
- Email: {st.session_state.user_email}
- Status: Not Started

The research assistant will begin working on your request and update the status as they progress. You can track the progress through the GU-TAP System.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                        """
                                        
                                        try:
                                            send_email_mailjet(
                                                to_email=tap_email,
                                                subject=tap_subject,
                                                body=tap_body.strip()
                                            )
                                            st.success(f"📧 Notification sent to TAP: {tap_name} ({tap_email})")
                                        except Exception as e:
                                            st.warning(f"⚠️ Failed to send notification to TAP {tap_name}: {e}")
                                    
                                    time.sleep(3)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")
                        else:
                            st.info("No unassigned requests at the moment.")

                # --- Section: Manage My Assigned Requests
                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                with st.expander("🚧 **MY ASSIGNED REQUESTS**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🚧 My Assigned Requests Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                Manage your assigned support requests. Update status and track your progress efficiently. Mark requests as completed when finished.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Get my assigned requests
                    my_requests = df_support[
                        (df_support["Student assigned"] == ga_support_name) & 
                        (df_support["Request status"].isin(["Not Started", "In Progress"]))
                    ].copy()

                    if my_requests.empty:
                        st.info("You have no assigned requests at the moment.")
                    else:
                        # Convert date column
                        my_requests["Date"] = pd.to_datetime(my_requests["Date"], errors="coerce")
                        my_requests["Date"] = my_requests["Date"].dt.strftime("%Y-%m-%d")

                        st.markdown("#### 📋 My Active Requests")
                        st.dataframe(my_requests[[
                            "Request Type","Date", "Time request needed", "Time Commitment", "Anticipated Deadline", "Request description", "Anticipated Deliverable", 
                            "TAP Name", "TAP email", "Request status"
                        ]].sort_values(by="Date", ascending=True).reset_index(drop=True))

                        # Select request to update status
                        my_request_indices = my_requests.index.tolist()
                        selected_my_request_idx = st.selectbox(
                            "Select a request to update status",
                            options=my_request_indices,
                            format_func=lambda idx: f"{my_requests.at[idx, 'Date']} | {my_requests.at[idx, 'TAP Name']} | Status: {my_requests.at[idx, 'Request status']}",
                            key='my_requests'
                        )

                        # Status update options
                        current_status = my_requests.at[selected_my_request_idx, 'Request status']
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            if current_status == "Not Started":
                                if st.button("🚀 Mark as In Progress", key='mark_in_progress'):
                                    try:
                                        updated_df_support = df_support.copy()
                                        updated_df_support.loc[selected_my_request_idx, "Request status"] = "In Progress"
                                        updated_df_support.loc[selected_my_request_idx, "Student email"] = st.session_state.user_email

                                        # Update Google Sheet
                                        updated_df_support = updated_df_support.fillna("")
                                        spreadsheet_support = client.open('HRSA64_TA_Request')
                                        worksheet_support = spreadsheet_support.worksheet('GA_Support')
                                        worksheet_support.update([updated_df_support.columns.values.tolist()] + updated_df_support.values.tolist())

                                        st.cache_data.clear()
                                        st.success("Request marked as 'In Progress'!")
                                        
                                        # Send status update email to TAP
                                        tap_email = updated_df_support.loc[selected_my_request_idx, "TAP email"]
                                        tap_name = updated_df_support.loc[selected_my_request_idx, "TAP Name"]
                                        request_date = updated_df_support.loc[selected_my_request_idx, "Date"]
                                        request_time = updated_df_support.loc[selected_my_request_idx, "Time request needed"]
                                        
                                        if tap_email and tap_email.strip():
                                            tap_subject = f"Support Request Status Update - In Progress"
                                            tap_body = f"""
Dear {tap_name},

Your support request status has been updated.

Request Details:
- Date: {request_date}
- Time: {request_time}
- Status: In Progress

Research Assistant: {ga_support_name}
Email: {st.session_state.user_email}

The research assistant has started working on your request and will continue to update the status as they make progress.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                            """
                                            
                                            try:
                                                send_email_mailjet(
                                                    to_email=tap_email,
                                                    subject=tap_subject,
                                                    body=tap_body.strip()
                                                )
                                                st.success(f"📧 Status update sent to TAP: {tap_name}")
                                            except Exception as e:
                                                st.warning(f"⚠️ Failed to send status update to TAP {tap_name}: {e}")
                                        
                                        time.sleep(3)
                                        st.rerun()

                                    except Exception as e:
                                        st.error(f"Error updating Google Sheets: {str(e)}")

                        with col2:
                            if st.button("✅ Mark as Completed", key='mark_completed'):
                                try:
                                    updated_df_support = df_support.copy()
                                    updated_df_support.loc[selected_my_request_idx, "Request status"] = "Completed"
                                    updated_df_support.loc[selected_my_request_idx, "Student email"] = st.session_state.user_email

                                    # Update Google Sheet
                                    updated_df_support = updated_df_support.fillna("")
                                    spreadsheet_support = client.open('HRSA64_TA_Request')
                                    worksheet_support = spreadsheet_support.worksheet('GA_Support')
                                    worksheet_support.update([updated_df_support.columns.values.tolist()] + updated_df_support.values.tolist())

                                    st.cache_data.clear()
                                    st.success("Request marked as 'Completed'!")
                                    
                                    # Send completion email to TAP
                                    tap_email = updated_df_support.loc[selected_my_request_idx, "TAP email"]
                                    tap_name = updated_df_support.loc[selected_my_request_idx, "TAP Name"]
                                    request_date = updated_df_support.loc[selected_my_request_idx, "Date"]
                                    request_time = updated_df_support.loc[selected_my_request_idx, "Time request needed"]
                                    anticipated_deliverable = updated_df_support.loc[selected_my_request_idx, "Anticipated Deliverable"]
                                    
                                    if tap_email and tap_email.strip():
                                        tap_subject = f"Support Request Completed - {request_date} at {request_time}"
                                        tap_body = f"""
Dear {tap_name},

Your support request has been completed!

Request Details:
- Date: {request_date}
- Time: {request_time}
- Status: Completed
- Anticipated Deliverable: {anticipated_deliverable}

Completed by:
- Research Assistant: {ga_support_name}
- Email: {st.session_state.user_email}

Thank you for using the GU-TAP System. Please contact the research assistant directly if you have any questions about the completed work.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                        """
                                        
                                        try:
                                            send_email_mailjet(
                                                to_email=tap_email,
                                                subject=tap_subject,
                                                body=tap_body.strip()
                                            )
                                            st.success(f"📧 Completion notification sent to TAP: {tap_name}")
                                        except Exception as e:
                                            st.warning(f"⚠️ Failed to send completion notification to TAP {tap_name}: {e}")
                                    
                                    time.sleep(3)
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"Error updating Google Sheets: {str(e)}")

                # --- Section: Re-assign Support
                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                with st.expander("🔄 **RE-ASSIGN SUPPORT**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                🔄 Re-assign Support Center
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                If you are unable to complete an assigned support request, you can re-assign it. The request will become unassigned and notifications will be sent to all available students.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Get my assigned requests (Not Started or In Progress)
                    my_assigned_for_reassign = df_support[
                        (df_support["Student assigned"] == ga_support_name) & 
                        (df_support["Request status"].isin(["Not Started", "In Progress"]))
                    ].copy()

                    if my_assigned_for_reassign.empty:
                        st.info("You have no assigned requests that can be re-assigned.")
                    else:
                        # Convert date column
                        my_assigned_for_reassign["Date"] = pd.to_datetime(my_assigned_for_reassign["Date"], errors="coerce")
                        my_assigned_for_reassign["Date"] = my_assigned_for_reassign["Date"].dt.strftime("%Y-%m-%d")

                        st.markdown("#### 📋 My Assigned Requests (Can be Re-assigned)")
                        st.dataframe(my_assigned_for_reassign[[
                            "Request Type","Date", "Time request needed", "Time Commitment", "Anticipated Deadline", "Request description", "Anticipated Deliverable", 
                            "TAP Name", "TAP email", "Request status"
                        ]].sort_values(by="Date", ascending=True).reset_index(drop=True))

                        # Select request to re-assign
                        reassign_indices = my_assigned_for_reassign.index.tolist()
                        selected_reassign_idx = st.selectbox(
                            "Select a request to re-assign",
                            options=reassign_indices,
                            format_func=lambda idx: f"{my_assigned_for_reassign.at[idx, 'Date']} | {my_assigned_for_reassign.at[idx, 'TAP Name']} | Status: {my_assigned_for_reassign.at[idx, 'Request status']}",
                            key='reassign_requests'
                        )

                        st.warning("⚠️ **Warning:** Re-assigning this request will remove your assignment and send notifications to all available students. This action cannot be undone.")

                        if st.button("🔄 Re-assign Support Request", key='reassign_support'):
                            try:
                                updated_df_support = df_support.copy()
                                
                                # Get request details before clearing assignment
                                request_type = updated_df_support.loc[selected_reassign_idx, "Request Type"]
                                request_date = updated_df_support.loc[selected_reassign_idx, "Date"]
                                request_time = updated_df_support.loc[selected_reassign_idx, "Time request needed"]
                                request_time_commitment = updated_df_support.loc[selected_reassign_idx, "Time Commitment"]
                                request_anticipated_deadline = updated_df_support.loc[selected_reassign_idx, "Anticipated Deadline"]
                                request_description = updated_df_support.loc[selected_reassign_idx, "Request description"]
                                anticipated_delivery = updated_df_support.loc[selected_reassign_idx, "Anticipated Deliverable"]
                                tap_name = updated_df_support.loc[selected_reassign_idx, "TAP Name"]
                                tap_email = updated_df_support.loc[selected_reassign_idx, "TAP email"]
                                
                                # Clear assignment fields
                                updated_df_support.loc[selected_reassign_idx, "Student assigned"] = ""
                                updated_df_support.loc[selected_reassign_idx, "Student email"] = ""
                                # Reset status to empty/unassigned state since no one is assigned
                                updated_df_support.loc[selected_reassign_idx, "Request status"] = ""
                                
                                # Update Google Sheet
                                updated_df_support = updated_df_support.fillna("")
                                spreadsheet_support = client.open('HRSA64_TA_Request')
                                worksheet_support = spreadsheet_support.worksheet('GA_Support')
                                worksheet_support.update([updated_df_support.columns.values.tolist()] + updated_df_support.values.tolist())

                                st.cache_data.clear()
                                st.success("✅ Request has been re-assigned! The supporter field has been cleared.")
                                
                                # Send notifications to ALL Research Assistants
                                st.markdown("**📧 Sending notifications to all Research Assistants...**")
                                
                                # Format date for display
                                if isinstance(request_date, str):
                                    try:
                                        date_obj = pd.to_datetime(request_date)
                                        date_str_display = date_obj.strftime("%Y-%m-%d")
                                    except:
                                        date_str_display = request_date
                                else:
                                    date_str_display = pd.to_datetime(request_date).strftime("%Y-%m-%d")
                                
                                # Send notifications to ALL Research Assistants
                                subject = f"Support Request Available for Re-assignment - {date_str_display}"
                                student_list = list(STUDENT_SCHEDULE.items())
                                success_count = 0
                                total_count = len(student_list)
                                
                                for i, (student_name, student_info) in enumerate(student_list, 1):
                                    body = f"""
Dear {student_name},

A support request has been re-assigned and is now available for assignment.

Request Details:
- Date: {date_str_display if request_date else 'N/A'}
- Time: {request_time if request_time else 'N/A'}
- TAP Name: {tap_name}
- TAP Email: {tap_email}
- Request Description: {request_description}
- Anticipated Deliverable: {anticipated_delivery}
{f"- Time Commitment: {request_time_commitment}" if request_time_commitment and str(request_time_commitment).strip() != "nan" else ""}
{f"- Anticipated Deadline: {request_anticipated_deadline}" if request_anticipated_deadline and str(request_anticipated_deadline).strip() != "nan" else ""}

This request was previously assigned but has been made available again. If you are interested in taking this request, please log into the GU-TAP System and assign it to yourself.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                    """
                                    
                                    try:
                                        status = send_email_mailjet(
                                            to_email=student_info['email'],
                                            subject=subject,
                                            body=body.strip()
                                        )
                                        if status:
                                            success_count += 1
                                            st.success(f"📧 ({i}/{total_count}) Sent to {student_name} ({student_info['email']})")
                                        
                                        # Add delay between emails to avoid rate limiting (except after the last email)
                                        if i < total_count:
                                            time.sleep(0.8)  # 0.8 second delay between emails
                                            
                                    except Exception as e:
                                        st.warning(f"⚠️ Failed to send notification to {student_name}: {e}")
                                
                                if success_count == total_count:
                                    st.success(f"✅ All {total_count} Research Assistant(s) have been notified!")
                                elif success_count > 0:
                                    st.warning(f"⚠️ Sent {success_count}/{total_count} notifications successfully")
                                else:
                                    st.error("❌ Failed to send notifications to Research Assistants")
                                
                                # Send notification to TAP about re-assignment
                                if tap_email and tap_email.strip():
                                    tap_subject = f"Support Request Re-assigned - {date_str_display} at {request_time if request_time else 'N/A'}"
                                    tap_body = f"""
Dear {tap_name},

Your support request has been re-assigned by the previously assigned research assistant.

Request Details:
- Request Type: {request_type}
- Date: {date_str_display}
- Time: {request_time if request_time else 'N/A'}
- Time Commitment: {request_time_commitment}
- Anticipated Deadline: {request_anticipated_deadline}
- Request Description: {request_description}
- Anticipated Deliverable: {anticipated_delivery}

Previous Research Assistant: {ga_support_name}
Email: {st.session_state.user_email}

The request is now unassigned and notifications have been sent to all Research Assistants. A new research assistant will be able to assign themselves to this request.

GU-TAP System: https://hrsagutap.streamlit.app/

Best regards,
GU-TAP System
                                    """
                                    
                                    try:
                                        send_email_mailjet(
                                            to_email=tap_email,
                                            subject=tap_subject,
                                            body=tap_body.strip()
                                        )
                                        st.success(f"📧 Re-assignment notification sent to TAP: {tap_name}")
                                    except Exception as e:
                                        st.warning(f"⚠️ Failed to send re-assignment notification to TAP {tap_name}: {e}")
                                
                                time.sleep(3)
                                st.rerun()

                            except Exception as e:
                                st.error(f"Error updating Google Sheets: {str(e)}")

                # --- Section: View Completed Requests
                st.markdown("<hr style='margin:2em 0; border:1px solid #dee2e6;'>", unsafe_allow_html=True)
                with st.expander("✅ **COMPLETED REQUESTS**"):
                    st.markdown("""
                        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 20px; box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3); padding: 2em 1.5em 1.5em 1.5em; margin-bottom: 2em; margin-top: 1em;'>
                            <div style='color: white; font-family: "Segoe UI", "Arial", sans-serif; font-weight: 800; font-size: 1.6em; margin-bottom: 0.5em; text-align: center;'>
                                ✅ Completed Requests Archive
                            </div>
                            <div style='color: rgba(255,255,255,0.9); font-size: 1.1em; margin-bottom: 0.8em; text-align: center; line-height: 1.4;'>
                                View all your completed support requests. Track your accomplishments and review past work for reference.
                            </div>
                        </div>
                    """, unsafe_allow_html=True)

                    # Get my completed requests
                    completed_requests = df_support[
                        (df_support["Student assigned"] == ga_support_name) & 
                        (df_support["Request status"] == "Completed")
                    ].copy()

                    if completed_requests.empty:
                        st.info("You have no completed requests yet.")
                    else:
                        # Convert date column
                        completed_requests["Date"] = pd.to_datetime(completed_requests["Date"], errors="coerce")
                        completed_requests["Date"] = completed_requests["Date"].dt.strftime("%Y-%m-%d")

                        st.markdown("#### ✅ My Completed Requests")
                        st.dataframe(completed_requests[[
                            "Date", "Time request needed", "Request description", "Anticipated Deliverable", 
                            "TAP Name", "TAP email", "Request status"
                        ]].sort_values(by="Date", ascending=True).reset_index(drop=True))

